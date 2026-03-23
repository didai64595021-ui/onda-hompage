#!/usr/bin/env python3
"""
fill_crawl_gui.py — 네이버 플레이스 크롤러 GUI (tkinter)
모던 다크 테마 v3.3 — 반응형 가로 2컬럼 레이아웃

기능:
  - 키워드 입력 + 큐 등록 (쉼표로 다중 추가)
  - 키워드 큐 순차 자동 크롤링
  - 딜레이/타임아웃/프록시/API 설정
  - 시작/중지 버튼
  - 진행률 바 + 실시간 통계
  - 실시간 로그 창
  - 결과 다운로드 (Excel)
  - 크롤링은 별도 Thread로 실행 (GUI 멈춤 방지)
  - 반응형 가로 2컬럼 ↔ 세로 1컬럼 자동 전환
"""
import json
import os
import sys
import threading
import tkinter as tk
from tkinter import filedialog, messagebox

CONFIG_FILE = os.path.join(os.getcwd(), "crawler_config.json")

# 같은 디렉토리에서 엔진 임포트
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from crawler_engine import CrawlerEngine

# ── 컬러 팔레트 ──
C_BG = "#1a1a2e"
C_CARD = "#16213e"
C_ACCENT = "#0f3460"
C_POINT = "#e94560"
C_SUCCESS = "#00d2ff"
C_TEXT = "#e8e8e8"
C_SUBTEXT = "#a0a0b0"
C_INPUT_BG = "#0d1b2a"
C_INPUT_FG = "#e8e8e8"
C_LOG_BG = "#0a0a0a"
C_LOG_FG = "#00d2ff"
C_DISABLED = "#6a7a8e"
C_BORDER = "#2a2a4e"

# ── 폰트 ──
FONT_TITLE = ("Malgun Gothic", 16, "bold")
FONT_BODY = ("Malgun Gothic", 10)
FONT_BODY_BOLD = ("Malgun Gothic", 10, "bold")
FONT_SMALL = ("Malgun Gothic", 9)
FONT_LOG = ("Consolas", 9)
FONT_STAT_NUM = ("Malgun Gothic", 14, "bold")
FONT_STAT_LABEL = ("Malgun Gothic", 8)
FONT_BTN = ("Malgun Gothic", 11, "bold")


def create_rounded_rect(canvas, x1, y1, x2, y2, radius=15, **kwargs):
    """Canvas에 둥근 모서리 사각형 그리기"""
    points = [
        x1 + radius, y1,
        x2 - radius, y1,
        x2, y1,
        x2, y1 + radius,
        x2, y2 - radius,
        x2, y2,
        x2 - radius, y2,
        x1 + radius, y2,
        x1, y2,
        x1, y2 - radius,
        x1, y1 + radius,
        x1, y1,
    ]
    return canvas.create_polygon(points, smooth=True, **kwargs)


class CardFrame(tk.Frame):
    """다크 테마 카드 스타일 프레임"""
    def __init__(self, parent, title=None, **kwargs):
        super().__init__(parent, bg=C_CARD, highlightbackground=C_BORDER,
                         highlightthickness=1, **kwargs)
        if title:
            lbl = tk.Label(self, text=title, font=FONT_BODY_BOLD,
                           fg=C_SUBTEXT, bg=C_CARD, anchor="w")
            lbl.pack(fill="x", padx=12, pady=(4, 1))


class StatCard(tk.Frame):
    """통계 카드 위젯"""
    def __init__(self, parent, icon, label, color=C_TEXT):
        super().__init__(parent, bg=C_CARD, highlightbackground=C_BORDER,
                         highlightthickness=1)
        self.color = color
        self.icon_label = tk.Label(self, text=icon, font=("", 13), bg=C_CARD, fg=color)
        self.icon_label.pack(pady=(4, 0))
        self.num_label = tk.Label(self, text="0", font=FONT_STAT_NUM, bg=C_CARD, fg=color)
        self.num_label.pack()
        self.text_label = tk.Label(self, text=label, font=FONT_STAT_LABEL, bg=C_CARD, fg=C_SUBTEXT)
        self.text_label.pack(pady=(0, 4))

    def set_value(self, val):
        self.num_label.config(text=str(val))


class SegmentButton(tk.Frame):
    """세그먼트 컨트롤 (2개 버튼 토글)"""
    def __init__(self, parent, options, variable, command=None):
        super().__init__(parent, bg=C_CARD)
        self.variable = variable
        self.command = command
        self.buttons = {}
        for i, opt in enumerate(options):
            btn = tk.Label(self, text=opt, font=FONT_BODY, cursor="hand2",
                           padx=16, pady=5, bg=C_ACCENT, fg=C_SUBTEXT)
            btn.pack(side="left", padx=(0 if i > 0 else 0, 0))
            btn.bind("<Button-1>", lambda e, o=opt: self._select(o))
            self.buttons[opt] = btn
        self._select(variable.get(), trigger=False)

    def _select(self, option, trigger=True):
        self.variable.set(option)
        for opt, btn in self.buttons.items():
            if opt == option:
                btn.config(bg=C_POINT, fg="white")
            else:
                btn.config(bg=C_ACCENT, fg=C_SUBTEXT)
        if trigger and self.command:
            self.command()


class GradientProgressBar(tk.Canvas):
    """커스텀 그라데이션 프로그레스 바"""
    def __init__(self, parent, height=22, **kwargs):
        super().__init__(parent, height=height, bg=C_ACCENT,
                         highlightthickness=0, **kwargs)
        self._pct = 0
        self.bind("<Configure>", self._draw)

    def set(self, pct):
        self._pct = max(0, min(100, pct))
        self._draw()

    def _draw(self, event=None):
        self.delete("all")
        w = self.winfo_width()
        h = self.winfo_height()
        if w < 2:
            return
        # 배경
        self.create_rectangle(0, 0, w, h, fill=C_ACCENT, outline="")
        # 채워진 부분
        fill_w = int(w * self._pct / 100)
        if fill_w > 0:
            steps = max(1, fill_w)
            for i in range(steps):
                ratio = i / max(1, fill_w - 1)
                r = int(233 * (1 - ratio) + 0 * ratio)
                g = int(69 * (1 - ratio) + 210 * ratio)
                b = int(96 * (1 - ratio) + 255 * ratio)
                color = f"#{r:02x}{g:02x}{b:02x}"
                self.create_line(i, 0, i, h, fill=color)
        # 퍼센트 텍스트 오버레이
        text = f"{self._pct:.0f}%"
        self.create_text(w // 2, h // 2, text=text, fill="white",
                         font=("Malgun Gothic", 9, "bold"))


class ToolTip:
    """마우스 호버 시 설명 팝업 표시"""
    def __init__(self, widget, text, delay=500):
        self.widget = widget
        self.text = text
        self.delay = delay
        self._tip_window = None
        self._after_id = None
        widget.bind("<Enter>", self._on_enter)
        widget.bind("<Leave>", self._on_leave)

    def _on_enter(self, event=None):
        self._after_id = self.widget.after(self.delay, self._show_tip)

    def _on_leave(self, event=None):
        if self._after_id:
            self.widget.after_cancel(self._after_id)
            self._after_id = None
        self._hide_tip()

    def _show_tip(self):
        if self._tip_window:
            return
        x = self.widget.winfo_rootx() + 20
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 4
        self._tip_window = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        tw.configure(bg="#e94560")
        inner = tk.Frame(tw, bg="#2d2d44", padx=8, pady=4,
                         highlightbackground="#e94560", highlightthickness=1)
        inner.pack()
        lbl = tk.Label(inner, text=self.text, font=("Malgun Gothic", 9),
                       fg="#e8e8e8", bg="#2d2d44", wraplength=300, justify="left")
        lbl.pack()

    def _hide_tip(self):
        if self._tip_window:
            self._tip_window.destroy()
            self._tip_window = None


class CrawlerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("네이버 플레이스 크롤러 v3.3")
        screen_w = self.root.winfo_screenwidth()
        screen_h = self.root.winfo_screenheight()
        win_w = max(700, int(screen_w * 0.8))
        win_h = max(500, int(screen_h * 0.8))
        self.root.geometry(f"{win_w}x{win_h}")
        self.root.minsize(700, 500)
        self.root.configure(bg=C_BG)

        # 아이콘 색상 (타이틀바)
        try:
            self.root.tk.call("wm", "iconphoto", self.root._w, tk.PhotoImage(data=""))
        except Exception:
            pass

        self.engine = None
        self.thread = None
        self.total_rows = 0
        self._settings_visible = False

        # 키워드 큐 상태
        self._queue_items = []  # [{"keyword": str, "status": "pending"|"running"|"done"|"failed", "count": 0}]
        self._queue_running = False
        self._queue_current_idx = -1
        self._queue_stop_requested = False
        self._keyword_results = {}  # {keyword: temp_file_path} 키워드별 결과 보존

        self._build_ui()

        # 종료 시 임시파일+progress 자동 정리
        self.root.protocol("WM_DELETE_WINDOW", self._on_closing)
        self._load_config()

    def _build_ui(self):
        # ── 메인 컨테이너 ──
        self._main = tk.Frame(self.root, bg=C_BG)
        self._main.pack(fill="both", expand=True)

        # ── 상단 타이틀 바 ──
        title_bar = tk.Frame(self._main, bg=C_CARD, height=36)
        title_bar.pack(fill="x")
        title_bar.pack_propagate(False)
        tk.Label(title_bar,
                 text="\U0001f50d 네이버 플레이스 크롤러 v3.3",
                 font=FONT_TITLE, fg=C_SUCCESS, bg=C_CARD
                 ).pack(side="left", padx=12, pady=4)
        tk.Label(title_bar, text="빈칸채우기 자동 크롤링",
                 font=FONT_SMALL, fg=C_SUBTEXT, bg=C_CARD
                 ).pack(side="left", pady=4)

        # ── 바디 ──
        self._body = tk.Frame(self._main, bg=C_BG)
        self._body.pack(fill="both", expand=True)

        # ── 콘텐츠 영역 (좌/우 패널) — 고정 높이, 로그가 나머지 차지 ──
        self._content = tk.Frame(self._body, bg=C_BG)
        self._content.pack(fill="x", padx=0, pady=0)
        self._content.columnconfigure(0, weight=7, minsize=300)
        self._content.columnconfigure(1, weight=2, minsize=160)
        self._content.rowconfigure(0, weight=1)

        # ══════════════════════════════
        # 좌측 패널 (컨트롤 영역) — pack 순서가 중요!
        # 순서: 입력 → 액션바 → 큐 → 설정
        # ══════════════════════════════
        self._left_panel = tk.Frame(self._content, bg=C_BG)

        # ── 키워드 입력 카드 ──
        input_card = CardFrame(self._left_panel, title="키워드 입력")
        input_card.pack(fill="x", pady=(4, 2))

        _input_hint = tk.Label(input_card,
                 text="쉼표(,)로 여러 개 한번에 추가 가능. 예: 수원 피부과, 안양 성형외과",
                 font=FONT_SMALL, fg=C_SUBTEXT, bg=C_CARD, anchor="w",
                 justify="left")
        _input_hint.pack(fill="x", padx=12, pady=(0, 2))
        _input_hint.bind("<Configure>", lambda e: e.widget.config(wraplength=max(1, e.width - 4)))

        input_inner = tk.Frame(input_card, bg=C_CARD)
        input_inner.pack(fill="x", padx=12, pady=(2, 6))

        # 키워드 입력 행
        self.kw_frame = tk.Frame(input_inner, bg=C_CARD)
        self.kw_frame.pack(fill="x", pady=2)
        self.keyword_var = tk.StringVar()
        self.kw_entry = tk.Entry(self.kw_frame, textvariable=self.keyword_var,
                                 bg=C_INPUT_BG, fg=C_INPUT_FG, insertbackground=C_SUCCESS,
                                 font=FONT_BODY, relief="flat", highlightthickness=1,
                                 highlightbackground=C_BORDER, highlightcolor=C_SUCCESS)
        self.kw_entry.pack(side="left", fill="x", expand=True, padx=(0, 6), ipady=4)
        self.kw_entry.bind("<Return>", lambda e: self._queue_add())
        ToolTip(self.kw_entry, "예: 수원 피부과, 강남 성형외과 (Enter로 큐에 추가)")

        self.queue_add_btn = tk.Button(self.kw_frame, text="+ 추가", font=FONT_BODY_BOLD,
                                        bg=C_SUCCESS, fg="#000000", relief="flat",
                                        activebackground="#33e0ff", activeforeground="#000000",
                                        cursor="hand2", command=self._queue_add,
                                        padx=12, pady=2)
        self.queue_add_btn.pack(side="left")
        ToolTip(self.queue_add_btn, "키워드를 큐에 추가합니다 (쉼표로 여러 개 가능)")

        # 페이지 설정 행
        self.kw_page_frame = tk.Frame(input_inner, bg=C_CARD)
        self.kw_page_frame.pack(fill="x", pady=2)
        self.kw_startpage_label = tk.Label(self.kw_page_frame, text="시작 페이지",
                                           font=FONT_BODY, fg=C_TEXT, bg=C_CARD, width=10, anchor="w")
        self.kw_startpage_label.pack(side="left")
        self.start_page_var = tk.IntVar(value=1)
        self.start_page_spin = tk.Spinbox(self.kw_page_frame, from_=1, to=999,
                                          textvariable=self.start_page_var, width=5,
                                          bg=C_INPUT_BG, fg=C_INPUT_FG, font=FONT_BODY,
                                          relief="flat", buttonbackground=C_ACCENT,
                                          highlightthickness=1, highlightbackground=C_BORDER)
        self.start_page_spin.pack(side="left", padx=(0, 16))
        self.kw_maxpage_label = tk.Label(self.kw_page_frame, text="최대 페이지 (0=무제한)",
                                         font=FONT_BODY, fg=C_SUBTEXT, bg=C_CARD)
        self.kw_maxpage_label.pack(side="left")
        self.max_pages_var = tk.IntVar(value=0)
        self.max_pages_spin = tk.Spinbox(self.kw_page_frame, from_=0, to=999,
                                         textvariable=self.max_pages_var, width=5,
                                         bg=C_INPUT_BG, fg=C_INPUT_FG, font=FONT_BODY,
                                         relief="flat", buttonbackground=C_ACCENT,
                                         highlightthickness=1, highlightbackground=C_BORDER)
        self.max_pages_spin.pack(side="left", padx=(4, 0))

        # 출력 파일 (내부 임시파일 사용)
        self.output_var = tk.StringVar()
        self._temp_output = ""

        # ── 키워드 큐 패널 ──
        # ── 키워드 큐 ──
        self.queue_card = CardFrame(self._left_panel, title="키워드 큐")
        self.queue_card.pack(fill="both", expand=True, pady=(2, 2))

        _queue_hint = tk.Label(self.queue_card,
                 text="여러 키워드를 등록하면 순서대로 자동 크롤링됩니다",
                 font=FONT_SMALL, fg=C_SUBTEXT, bg=C_CARD, anchor="w",
                 justify="left")
        _queue_hint.pack(fill="x", padx=12, pady=(0, 2))
        _queue_hint.bind("<Configure>", lambda e: e.widget.config(wraplength=max(1, e.width - 4)))

        queue_inner = tk.Frame(self.queue_card, bg=C_CARD)
        queue_inner.pack(fill="both", expand=True, padx=12, pady=(2, 6))

        self.queue_progress_label = tk.Label(queue_inner, text="0/0 키워드 완료",
                                              font=FONT_BODY_BOLD, fg=C_SUCCESS, bg=C_CARD)
        self.queue_progress_label.pack(fill="x", pady=(0, 4))

        q_list_frame = tk.Frame(queue_inner, bg="#0d1117")
        q_list_frame.pack(fill="both", expand=True, pady=(4, 4))

        self.queue_listbox = tk.Listbox(q_list_frame, height=3,
                                         bg="#0d1117", fg=C_TEXT, font=FONT_BODY,
                                         selectbackground=C_POINT, selectforeground="white",
                                         relief="flat", highlightthickness=1,
                                         highlightbackground=C_BORDER, activestyle="none")
        q_scrollbar = tk.Scrollbar(q_list_frame, command=self.queue_listbox.yview,
                                    bg=C_CARD, troughcolor="#0d1117",
                                    activebackground=C_ACCENT, highlightthickness=0)
        self.queue_listbox.config(yscrollcommand=q_scrollbar.set)
        q_scrollbar.pack(side="right", fill="y")
        self.queue_listbox.pack(side="left", fill="both", expand=True)

        q_btn_row = tk.Frame(queue_inner, bg=C_CARD)
        q_btn_row.pack(fill="x", pady=(2, 0))
        q_btn_row.columnconfigure(0, weight=0)
        q_btn_row.columnconfigure(1, weight=0)
        q_btn_row.columnconfigure(2, weight=1)
        q_btn_row.columnconfigure(3, weight=1)

        self.queue_up_btn = tk.Button(q_btn_row, text="▲", font=FONT_BODY_BOLD,
                                       bg=C_ACCENT, fg=C_TEXT, relief="flat",
                                       activebackground=C_POINT, cursor="hand2",
                                       command=self._queue_move_up, padx=6, pady=1)
        self.queue_up_btn.grid(row=0, column=0, padx=(0, 2))

        self.queue_down_btn = tk.Button(q_btn_row, text="▼", font=FONT_BODY_BOLD,
                                         bg=C_ACCENT, fg=C_TEXT, relief="flat",
                                         activebackground=C_POINT, cursor="hand2",
                                         command=self._queue_move_down, padx=6, pady=1)
        self.queue_down_btn.grid(row=0, column=1, padx=(0, 8))

        self.queue_del_btn = tk.Button(q_btn_row, text="선택삭제", font=FONT_SMALL,
                                        bg=C_POINT, fg="white", relief="flat",
                                        activebackground="#ff5a7a", cursor="hand2",
                                        command=self._queue_remove, pady=1)
        self.queue_del_btn.grid(row=0, column=2, sticky="ew", padx=(0, 2))
        ToolTip(self.queue_del_btn, "선택한 키워드를 큐에서 삭제합니다")

        self.queue_clear_btn = tk.Button(q_btn_row, text="전체삭제", font=FONT_SMALL,
                                          bg=C_POINT, fg="white", relief="flat",
                                          activebackground="#ff5a7a", cursor="hand2",
                                          command=self._queue_clear, pady=1)
        self.queue_clear_btn.grid(row=0, column=3, sticky="ew")
        ToolTip(self.queue_clear_btn, "큐의 모든 키워드를 삭제합니다")

        # ── 설정 영역 (접이식) ──
        self._settings_frame = tk.Frame(self._left_panel, bg=C_BG)
        self._settings_frame.pack(fill="x", pady=(2, 0))
        self._settings_header = tk.Frame(self._settings_frame, bg=C_BG)
        self._settings_header.pack(fill="x")
        self._settings_toggle_var = tk.StringVar(value="\u25b6 설정")
        self._settings_toggle_btn = tk.Label(self._settings_header,
                                              textvariable=self._settings_toggle_var,
                                              font=FONT_BODY_BOLD, fg=C_SUBTEXT, bg=C_BG,
                                              cursor="hand2")
        self._settings_toggle_btn.pack(side="left")
        self._settings_toggle_btn.bind("<Button-1>", self._toggle_settings)
        self._settings_desc = tk.Label(self._settings_header,
                                        text="⚙️ 딜레이, 프록시, API 등 상세 설정",
                                        font=FONT_SMALL, fg=C_SUBTEXT, bg=C_BG)
        self._settings_desc.pack(side="left", padx=(8, 0))

        self.settings_card = CardFrame(self._settings_frame)
        self._settings_visible = False

        settings_inner = tk.Frame(self.settings_card, bg=C_CARD)
        settings_inner.pack(fill="x", padx=12, pady=8)

        # 딜레이 설정
        delay_row = tk.Frame(settings_inner, bg=C_CARD)
        delay_row.pack(fill="x", pady=2)

        tk.Label(delay_row, text="최소 딜레이 (초)", font=FONT_BODY,
                 fg=C_TEXT, bg=C_CARD, width=14, anchor="w").pack(side="left")
        self.delay_min_var = tk.DoubleVar(value=0.5)
        delay_min_scale = tk.Scale(delay_row, from_=0.3, to=15.0, variable=self.delay_min_var,
                                   orient="horizontal", length=140, resolution=0.1,
                                   bg=C_CARD, fg=C_TEXT, troughcolor=C_ACCENT,
                                   highlightthickness=0, sliderrelief="flat",
                                   activebackground=C_POINT, font=FONT_SMALL,
                                   showvalue=False)
        delay_min_scale.pack(side="left", padx=(0, 4))
        ToolTip(delay_min_scale, "요청 사이 최소 대기 시간 (낮을수록 빠르지만 차단 위험 ↑)")
        self.delay_min_label = tk.Label(delay_row, text="0.5s", font=FONT_BODY,
                                        fg=C_SUCCESS, bg=C_CARD, width=5)
        self.delay_min_label.pack(side="left")
        self.delay_min_var.trace_add("write", lambda *_: self.delay_min_label.config(
            text=f"{self.delay_min_var.get():.1f}s"))

        tk.Label(delay_row, text="최대 딜레이", font=FONT_BODY,
                 fg=C_TEXT, bg=C_CARD, width=9, anchor="e").pack(side="left", padx=(16, 0))
        self.delay_max_var = tk.DoubleVar(value=1.5)
        delay_max_scale = tk.Scale(delay_row, from_=2.0, to=30.0, variable=self.delay_max_var,
                                   orient="horizontal", length=140, resolution=0.1,
                                   bg=C_CARD, fg=C_TEXT, troughcolor=C_ACCENT,
                                   highlightthickness=0, sliderrelief="flat",
                                   activebackground=C_POINT, font=FONT_SMALL,
                                   showvalue=False)
        delay_max_scale.pack(side="left", padx=(4, 4))
        ToolTip(delay_max_scale, "요청 사이 최대 대기 시간")
        self.delay_max_label = tk.Label(delay_row, text="1.5s", font=FONT_BODY,
                                        fg=C_SUCCESS, bg=C_CARD, width=5)
        self.delay_max_label.pack(side="left")
        self.delay_max_var.trace_add("write", lambda *_: self.delay_max_label.config(
            text=f"{self.delay_max_var.get():.1f}s"))

        # 타임아웃
        timeout_row = tk.Frame(settings_inner, bg=C_CARD)
        timeout_row.pack(fill="x", pady=2)
        tk.Label(timeout_row, text="타임아웃 (초)", font=FONT_BODY,
                 fg=C_TEXT, bg=C_CARD, width=14, anchor="w").pack(side="left")
        self.timeout_var = tk.IntVar(value=8)
        timeout_spin = tk.Spinbox(timeout_row, from_=3, to=30, textvariable=self.timeout_var,
                                  width=5, bg=C_INPUT_BG, fg=C_INPUT_FG, font=FONT_BODY,
                                  relief="flat", buttonbackground=C_ACCENT,
                                  highlightthickness=1, highlightbackground=C_BORDER)
        timeout_spin.pack(side="left", padx=(0, 4))
        ToolTip(timeout_spin, "응답 대기 최대 시간 (초과 시 다음으로 넘어감)")

        # 프록시 모드
        proxy_row = tk.Frame(settings_inner, bg=C_CARD)
        proxy_row.pack(fill="x", pady=2)
        tk.Label(proxy_row, text="프록시 모드", font=FONT_BODY,
                 fg=C_TEXT, bg=C_CARD, width=14, anchor="w").pack(side="left")
        self.proxy_mode_var = tk.StringVar(value="없음 (직접연결)")
        self._proxy_seg = SegmentButton(proxy_row, ["없음 (직접연결)", "프록시 사용"],
                                        self.proxy_mode_var, command=self._on_proxy_mode_change)
        self._proxy_seg.pack(side="left", padx=(0, 8))
        ToolTip(self._proxy_seg, "IP 차단 방지용 프록시 설정 (하루 5000건 이상 시 권장)")

        self.proxy_file_var = tk.StringVar()
        self.proxy_file_entry = tk.Entry(proxy_row, textvariable=self.proxy_file_var,
                                         bg=C_INPUT_BG, fg=C_INPUT_FG, font=FONT_SMALL,
                                         relief="flat", width=24, state="disabled",
                                         disabledbackground=C_DISABLED, disabledforeground=C_SUBTEXT,
                                         highlightthickness=1, highlightbackground=C_BORDER)
        self.proxy_file_entry.pack(side="left", padx=(0, 4), ipady=2)
        self.proxy_file_btn = tk.Button(proxy_row, text="파일선택", font=FONT_SMALL,
                                        bg=C_ACCENT, fg=C_TEXT, relief="flat",
                                        activebackground=C_POINT, state="disabled",
                                        disabledforeground=C_SUBTEXT, command=self._browse_proxy,
                                        padx=8, pady=1)
        self.proxy_file_btn.pack(side="left")

        # API 모드
        api_row = tk.Frame(settings_inner, bg=C_CARD)
        api_row.pack(fill="x", pady=2)
        tk.Label(api_row, text="API 모드", font=FONT_BODY,
                 fg=C_TEXT, bg=C_CARD, width=14, anchor="w").pack(side="left")
        self.api_mode_var = tk.StringVar(value="없음 (HTML 크롤링)")
        self._api_seg = SegmentButton(api_row, ["없음 (HTML 크롤링)", "네이버 검색 API"],
                                      self.api_mode_var, command=self._on_api_mode_change)
        self._api_seg.pack(side="left")
        ToolTip(self._api_seg, "네이버 검색 API 사용 여부 (없어도 크롤링 가능)")

        # API 키 입력
        api_keys_frame = tk.Frame(settings_inner, bg=C_CARD)
        api_keys_frame.pack(fill="x", pady=(2, 4))

        api_key1_row = tk.Frame(api_keys_frame, bg=C_CARD)
        api_key1_row.pack(fill="x", pady=1)
        tk.Label(api_key1_row, text="Client ID", font=FONT_SMALL,
                 fg=C_SUBTEXT, bg=C_CARD, width=14, anchor="w").pack(side="left")
        self.api_id1_var = tk.StringVar(value="j2jMCfsKAStJ10rb0Q2A")
        self.api_id1_entry = tk.Entry(api_key1_row, textvariable=self.api_id1_var,
                                      bg=C_INPUT_BG, fg=C_INPUT_FG, font=FONT_SMALL,
                                      relief="flat", width=18, state="disabled",
                                      disabledbackground=C_DISABLED, disabledforeground=C_SUBTEXT,
                                      highlightthickness=1, highlightbackground=C_BORDER)
        self.api_id1_entry.pack(side="left", padx=(0, 8), ipady=2)
        tk.Label(api_key1_row, text="Secret", font=FONT_SMALL,
                 fg=C_SUBTEXT, bg=C_CARD).pack(side="left", padx=(0, 4))
        self.api_secret1_var = tk.StringVar(value="nv10elHEYJ")
        self.api_secret1_entry = tk.Entry(api_key1_row, textvariable=self.api_secret1_var,
                                          bg=C_INPUT_BG, fg=C_INPUT_FG, font=FONT_SMALL,
                                          relief="flat", width=18, show="*", state="disabled",
                                          disabledbackground=C_DISABLED, disabledforeground=C_SUBTEXT,
                                          highlightthickness=1, highlightbackground=C_BORDER)
        self.api_secret1_entry.pack(side="left", ipady=2)

        api_key2_row = tk.Frame(api_keys_frame, bg=C_CARD)
        api_key2_row.pack(fill="x", pady=1)
        tk.Label(api_key2_row, text="Client ID 2", font=FONT_SMALL,
                 fg=C_SUBTEXT, bg=C_CARD, width=14, anchor="w").pack(side="left")
        self.api_id2_var = tk.StringVar()
        self.api_id2_entry = tk.Entry(api_key2_row, textvariable=self.api_id2_var,
                                      bg=C_INPUT_BG, fg=C_INPUT_FG, font=FONT_SMALL,
                                      relief="flat", width=18, state="disabled",
                                      disabledbackground=C_DISABLED, disabledforeground=C_SUBTEXT,
                                      highlightthickness=1, highlightbackground=C_BORDER)
        self.api_id2_entry.pack(side="left", padx=(0, 8), ipady=2)
        tk.Label(api_key2_row, text="Secret 2", font=FONT_SMALL,
                 fg=C_SUBTEXT, bg=C_CARD).pack(side="left", padx=(0, 4))
        self.api_secret2_var = tk.StringVar()
        self.api_secret2_entry = tk.Entry(api_key2_row, textvariable=self.api_secret2_var,
                                          bg=C_INPUT_BG, fg=C_INPUT_FG, font=FONT_SMALL,
                                          relief="flat", width=18, show="*", state="disabled",
                                          disabledbackground=C_DISABLED, disabledforeground=C_SUBTEXT,
                                          highlightthickness=1, highlightbackground=C_BORDER)
        self.api_secret2_entry.pack(side="left", ipady=2)

        # ── 액션 바 (입력 바로 다음에 삽입) ──
        action_card = tk.Frame(self._left_panel, bg=C_BG)
        action_card.pack(fill="x", pady=(2, 2), after=input_card)

        btn_row = tk.Frame(action_card, bg=C_BG)
        btn_row.pack(fill="x")
        btn_row.columnconfigure(0, weight=2)
        btn_row.columnconfigure(1, weight=2)
        btn_row.columnconfigure(2, weight=3)

        self.start_btn = tk.Button(btn_row, text="\u25b6 시작", font=FONT_BTN,
                                   bg=C_POINT, fg="white", relief="flat",
                                   activebackground="#ff5a7a", activeforeground="white",
                                   cursor="hand2", command=self._start_crawl,
                                   pady=6)
        self.start_btn.grid(row=0, column=0, sticky="ew", padx=(0, 3))
        ToolTip(self.start_btn, "크롤링을 시작합니다")

        self.stop_btn = tk.Button(btn_row, text="\u25a0 중지", font=FONT_BTN,
                                  bg=C_ACCENT, fg=C_SUBTEXT, relief="flat",
                                  activebackground="#1a4a80", activeforeground=C_TEXT,
                                  cursor="hand2", command=self._stop_crawl,
                                  state="disabled", disabledforeground=C_DISABLED,
                                  pady=6)
        self.stop_btn.grid(row=0, column=1, sticky="ew", padx=3)
        ToolTip(self.stop_btn, "현재 작업을 중지합니다 (진행중 데이터는 보존)")

        self.download_btn = tk.Button(btn_row, text="\u2b07 다운로드", font=FONT_BTN,
                                      bg="#005577", fg=C_SUCCESS, relief="flat",
                                      activebackground="#007799", activeforeground="white",
                                      cursor="hand2", command=self._download_result,
                                      state="disabled", disabledforeground=C_DISABLED,
                                      pady=6)
        self.download_btn.grid(row=0, column=2, sticky="ew", padx=(3, 0))
        ToolTip(self.download_btn, "크롤링 결과를 Excel 파일로 저장합니다")

        # 프로그레스 바 + 라벨
        progress_frame = tk.Frame(action_card, bg=C_BG)
        progress_frame.pack(fill="x", pady=(6, 0))

        self.progress_var = tk.DoubleVar(value=0)
        self.progress_bar = GradientProgressBar(progress_frame)
        self.progress_bar.pack(fill="x", side="left", expand=True, padx=(0, 8))
        ToolTip(self.progress_bar, "현재 크롤링 진행률")

        self.progress_label = tk.Label(progress_frame, text="0/0 (0%)", font=FONT_BODY,
                                       fg=C_SUBTEXT, bg=C_BG, width=16, anchor="e")
        self.progress_label.pack(side="right")

        # ══════════════════════════════
        # 우측 패널 (통계 대시보드)
        # ══════════════════════════════
        self._right_panel = tk.Frame(self._content, bg=C_BG)

        stats_header = tk.Frame(self._right_panel, bg=C_BG)
        stats_header.pack(fill="x", pady=(4, 0))
        tk.Label(stats_header, text="통계", font=FONT_BODY_BOLD,
                 fg=C_SUBTEXT, bg=C_BG, anchor="w").pack(side="left")
        tk.Label(stats_header, text="실시간 수집 현황", font=FONT_SMALL,
                 fg=C_SUBTEXT, bg=C_BG, anchor="w").pack(side="left", padx=(8, 0))

        self._stats_frame = tk.Frame(self._right_panel, bg=C_BG)
        self._stats_frame.pack(fill="both", expand=True, pady=(2, 4))

        stats_data = [
            ("\U0001f4de", "안심번호", C_POINT),
            ("\U0001f310", "홈페이지", C_SUCCESS),
            ("\U0001f4e7", "이메일", "#ffb347"),
            ("\U0001f194", "네이버ID", "#bb86fc"),
            ("\u2705", "성공", "#00e676"),
            ("\U0001f6ab", "차단", "#ff5252"),
        ]
        stat_tooltips = {
            "안심번호": "수집된 안심번호(050/070) 수",
            "홈페이지": "수집된 홈페이지 URL 수",
            "이메일": "수집된 이메일 주소 수",
            "네이버ID": "수집된 네이버 블로그/아이디 수",
            "성공": "상세 크롤링 성공 건수",
            "차단": "네이버 차단(403/429)으로 실패한 건수",
        }
        self._stat_card_list = []
        self.stat_cards = {}
        for i, (icon, label, color) in enumerate(stats_data):
            card = StatCard(self._stats_frame, icon, label, color)
            self._stat_card_list.append(card)
            self.stat_cards[label] = card
            if label in stat_tooltips:
                ToolTip(card, stat_tooltips[label])

        # backward compat aliases for _update_stats
        self.stat_phone = type('', (), {'config': lambda s, **kw: None})()
        self.stat_hp = type('', (), {'config': lambda s, **kw: None})()
        self.stat_email = type('', (), {'config': lambda s, **kw: None})()
        self.stat_success = type('', (), {'config': lambda s, **kw: None})()
        self.stat_blocked = type('', (), {'config': lambda s, **kw: None})()

        # ══════════════════════════════
        # 하단 로그 영역 (전체 폭)
        # ══════════════════════════════
        log_card = CardFrame(self._body, title="로그")
        log_card.pack(fill="both", expand=True, padx=12, pady=(4, 10))

        log_inner = tk.Frame(log_card, bg=C_LOG_BG)
        log_inner.pack(fill="both", expand=True, padx=8, pady=(2, 8))

        self.log_text = tk.Text(log_inner, wrap="word", font=FONT_LOG,
                                bg=C_LOG_BG, fg=C_LOG_FG, height=6,
                                insertbackground=C_SUCCESS, relief="flat",
                                highlightthickness=0, padx=8, pady=6,
                                selectbackground=C_ACCENT, selectforeground=C_TEXT)
        log_scrollbar = tk.Scrollbar(log_inner, command=self.log_text.yview,
                                     bg=C_CARD, troughcolor=C_LOG_BG,
                                     activebackground=C_ACCENT, highlightthickness=0)
        self.log_text.config(yscrollcommand=log_scrollbar.set)
        log_scrollbar.pack(side="right", fill="y")
        self.log_text.pack(side="left", fill="both", expand=True)
        self.log_text.config(state="disabled")

        # 로그 태그 색상
        self.log_text.tag_configure("INFO", foreground=C_SUCCESS)
        self.log_text.tag_configure("ERROR", foreground="#ff5252")
        self.log_text.tag_configure("SUCCESS", foreground="#00e676")
        self.log_text.tag_configure("WARN", foreground="#ffb347")
        self.log_text.tag_configure("WELCOME", foreground=C_SUBTEXT)

        # 초기 안내 메시지
        welcome = (
            "Welcome to 네이버 플레이스 크롤러 v3.3\n"
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            "📋 사용법:\n"
            "  1. 키워드 입력 → 시작 (단일 키워드 실행)\n"
            "  2. 키워드 여러 개 추가 → 시작 (큐 순차 자동 실행)\n"
            "\n"
            "💡 팁: 쉼표(,)로 여러 키워드 한번에 추가 가능\n"
            "💡 팁: 딜레이 0.5~1.5초면 차단 없이 빠르게 수집 가능\n"
            "📥 완료 후 '결과 다운로드' 버튼으로 Excel 저장\n"
        )
        self.log_text.config(state="normal")
        self.log_text.insert("end", welcome, "WELCOME")
        self.log_text.config(state="disabled")

        # ── 반응형 레이아웃 초기화 ──
        self._current_layout = None
        self._stats_cols = 0
        self._resize_after_id = None
        self.root.bind("<Configure>", self._on_resize)
        self.root.after(50, self._apply_layout)

    # ═══════════════════════════════════════════
    # 반응형 레이아웃
    # ═══════════════════════════════════════════

    def _on_resize(self, event=None):
        if event and event.widget != self.root:
            return
        if self._resize_after_id:
            self.root.after_cancel(self._resize_after_id)
        self._resize_after_id = self.root.after(80, self._apply_layout)

    def _apply_layout(self):
        self._resize_after_id = None
        width = self.root.winfo_width()
        new_layout = "horizontal" if width >= 800 else "vertical"

        if new_layout != self._current_layout:
            self._current_layout = new_layout
            self._left_panel.grid_forget()
            self._right_panel.grid_forget()

            if new_layout == "horizontal":
                self._content.columnconfigure(1, weight=2, minsize=160)
                self._content.rowconfigure(1, weight=0)
                self._left_panel.grid(row=0, column=0, sticky="nsew", padx=(12, 6), pady=4)
                self._right_panel.grid(row=0, column=1, sticky="nsew", padx=(6, 12), pady=4)
            else:
                self._content.columnconfigure(1, weight=0, minsize=0)
                self._content.rowconfigure(0, weight=1)
                self._content.rowconfigure(1, weight=0)
                self._left_panel.grid(row=0, column=0, sticky="nsew", padx=12, pady=4)
                self._right_panel.grid(row=1, column=0, sticky="nsew", padx=12, pady=4)

        # 통계 카드 그리드 재배치
        if new_layout == "horizontal":
            stat_cols = 2  # 우측 패널이 좁으므로 2열x3행
        else:
            stat_cols = 6 if width >= 800 else 3
        self._rearrange_stats(stat_cols)

    def _rearrange_stats(self, cols):
        if self._stats_cols == cols:
            return
        self._stats_cols = cols
        for card in self._stat_card_list:
            card.grid_forget()
            card.pack_forget()
        if cols <= 2:
            # 가로 모드: 2열 grid, uniform으로 균등 분배
            for i, card in enumerate(self._stat_card_list):
                r, c = divmod(i, cols)
                card.grid(row=r, column=c, sticky="nsew", padx=2, pady=2)
            for c in range(cols):
                self._stats_frame.columnconfigure(c, weight=1, uniform="stat")
            for c in range(cols, 6):
                self._stats_frame.columnconfigure(c, weight=0, minsize=0)
        else:
            # 세로 모드: cols열 grid
            for i, card in enumerate(self._stat_card_list):
                r, c = divmod(i, cols)
                card.grid(row=r, column=c, sticky="nsew", padx=2, pady=2)
            for c in range(cols):
                self._stats_frame.columnconfigure(c, weight=1, uniform="stat")
            for c in range(cols, 6):
                self._stats_frame.columnconfigure(c, weight=0, minsize=0)

    # ═══════════════════════════════════════════
    # 접이식 설정 패널
    # ═══════════════════════════════════════════

    def _toggle_settings(self, event=None):
        if self._settings_visible:
            self.settings_card.pack_forget()
            self._settings_toggle_var.set("\u25b6 설정")
            self._settings_visible = False
        else:
            self.settings_card.pack(fill="x", pady=(0, 4),
                                    after=self._settings_header)
            self._settings_toggle_var.set("\u25bc 설정")
            self._settings_visible = True

    # ═══════════════════════════════════════════
    # 설정 저장/로드
    # ═══════════════════════════════════════════

    def _load_config(self):
        """시작 시 crawler_config.json 로드 -> 각 입력란에 값 채움"""
        if not os.path.isfile(CONFIG_FILE):
            return
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                cfg = json.load(f)
        except (json.JSONDecodeError, OSError):
            return

        # 딜레이/타임아웃
        if cfg.get("delay_min") is not None:
            self.delay_min_var.set(cfg["delay_min"])
        if cfg.get("delay_max") is not None:
            self.delay_max_var.set(cfg["delay_max"])
        if cfg.get("timeout") is not None:
            self.timeout_var.set(cfg["timeout"])

        # 프록시 모드
        if cfg.get("proxy_mode"):
            self.proxy_mode_var.set(cfg["proxy_mode"])
            self._on_proxy_mode_change()
        if cfg.get("proxy_file"):
            self.proxy_file_var.set(cfg["proxy_file"])

        # API 모드
        if cfg.get("api_mode"):
            self.api_mode_var.set(cfg["api_mode"])
            self._on_api_mode_change()
        if cfg.get("api_key1_id"):
            self.api_id1_var.set(cfg["api_key1_id"])
        if cfg.get("api_key1_secret"):
            self.api_secret1_var.set(cfg["api_key1_secret"])
        if cfg.get("api_key2_id"):
            self.api_id2_var.set(cfg["api_key2_id"])
        if cfg.get("api_key2_secret"):
            self.api_secret2_var.set(cfg["api_key2_secret"])

        if cfg.get("keyword"):
            self.keyword_var.set(cfg["keyword"])
        if cfg.get("start_page") is not None:
            self.start_page_var.set(cfg["start_page"])
        if cfg.get("max_pages") is not None:
            self.max_pages_var.set(cfg["max_pages"])

        # 세그먼트 버튼 시각 동기화 (프록시/API)
        if cfg.get("proxy_mode"):
            self._proxy_seg._select(cfg["proxy_mode"], trigger=False)
        if cfg.get("api_mode"):
            self._api_seg._select(cfg["api_mode"], trigger=False)

        # 키워드 큐 로드
        if cfg.get("keyword_queue"):
            self._queue_items = []
            for kw in cfg["keyword_queue"]:
                if isinstance(kw, str) and kw.strip():
                    self._queue_items.append({"keyword": kw.strip(), "status": "pending", "count": 0})
            self._queue_refresh_listbox()

    def _save_config(self):
        """현재 GUI 값 -> crawler_config.json 저장"""
        # 큐에서 pending 키워드만 저장 (done/failed 제외)
        queue_keywords = [item["keyword"] for item in self._queue_items
                          if item["status"] in ("pending", "running")]
        cfg = {
            "proxy_mode": self.proxy_mode_var.get(),
            "proxy_file": self.proxy_file_var.get(),
            "api_mode": self.api_mode_var.get(),
            "api_key1_id": self.api_id1_var.get(),
            "api_key1_secret": self.api_secret1_var.get(),
            "api_key2_id": self.api_id2_var.get(),
            "api_key2_secret": self.api_secret2_var.get(),
            "delay_min": self.delay_min_var.get(),
            "delay_max": self.delay_max_var.get(),
            "timeout": self.timeout_var.get(),
            "keyword": self.keyword_var.get(),
            "start_page": self.start_page_var.get(),
            "max_pages": self.max_pages_var.get(),
            "keyword_queue": queue_keywords,
        }
        try:
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(cfg, f, ensure_ascii=False, indent=2)
        except OSError:
            pass

    # ═══════════════════════════════════════════
    # 파일 선택 다이얼로그
    # ═══════════════════════════════════════════

    def _on_proxy_mode_change(self, event=None):
        if self.proxy_mode_var.get() == "프록시 사용":
            self.proxy_file_entry.config(state="normal")
            self.proxy_file_btn.config(state="normal")
        else:
            self.proxy_file_entry.config(state="disabled")
            self.proxy_file_btn.config(state="disabled")

    def _on_api_mode_change(self, event=None):
        state = "normal" if self.api_mode_var.get() == "네이버 검색 API" else "disabled"
        for w in (self.api_id1_entry, self.api_secret1_entry,
                  self.api_id2_entry, self.api_secret2_entry):
            w.config(state=state)

    def _browse_proxy(self):
        path = filedialog.askopenfilename(
            title="프록시 파일 선택",
            filetypes=[("텍스트 파일", "*.txt"), ("모든 파일", "*.*")],
        )
        if path:
            self.proxy_file_var.set(path)

    def _on_closing(self):
        """종료 시 임시파일 + progress 파일 자동 정리"""
        import glob
        import tempfile
        # progress 파일 삭제
        tmp_dir = tempfile.gettempdir()
        for f in glob.glob(os.path.join(tmp_dir, "crawl_result*progress*")):
            try:
                os.remove(f)
            except Exception:
                pass
        # 임시 출력 파일 삭제 (다운로드 안 한 경우)
        if hasattr(self, "_temp_output") and self._temp_output and os.path.isfile(self._temp_output):
            try:
                os.remove(self._temp_output)
            except Exception:
                pass
        # 키워드별 임시 결과 파일 삭제
        if hasattr(self, "_keyword_results"):
            for kw_path in self._keyword_results.values():
                try:
                    if kw_path and os.path.isfile(kw_path):
                        os.remove(kw_path)
                except Exception:
                    pass
            self._keyword_results.clear()
        # Playwright 브라우저 정리
        if self.engine:
            try:
                self.engine._close_playwright()
            except Exception:
                pass
        self.root.destroy()

    def _download_result(self):
        """크롤링 완료 후 결과 파일 다운로드 (저장 위치 선택)"""
        if not self._temp_output or not os.path.isfile(self._temp_output):
            messagebox.showerror("오류", "다운로드할 결과 파일이 없습니다.\n먼저 크롤링을 실행하세요.")
            return

        # 큐 모드(키워드별 결과가 있을 때) → 선택 다운로드 팝업
        if self._keyword_results:
            self._download_select_dialog()
            return

        # 단일 키워드 모드 → 기존처럼 바로 다운로드
        self._save_file_dialog(self._temp_output)

    def _save_file_dialog(self, src_file):
        """파일 저장 다이얼로그"""
        import shutil
        path = filedialog.asksaveasfilename(
            title="결과 파일 저장",
            filetypes=[("엑셀 파일", "*.xlsx"), ("CSV 파일", "*.csv"), ("모든 파일", "*.*")],
            defaultextension=".xlsx",
            initialfile="크롤링결과.xlsx",
        )
        if path:
            try:
                shutil.copy2(src_file, path)
                messagebox.showinfo("완료", f"저장 완료!\n\n{path}")
                self._append_log(f"파일 저장: {path}")
            except Exception as e:
                messagebox.showerror("오류", f"저장 실패: {e}")

    def _download_select_dialog(self):
        """키워드별 선택 다운로드 팝업"""
        dialog = tk.Toplevel(self.root)
        dialog.title("결과 다운로드 — 키워드 선택")
        dialog.configure(bg=C_BG)
        dialog.transient(self.root)
        dialog.grab_set()

        # 크기/위치
        dw, dh = 420, 400
        x = self.root.winfo_x() + (self.root.winfo_width() - dw) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - dh) // 2
        dialog.geometry(f"{dw}x{dh}+{x}+{y}")
        dialog.minsize(350, 300)

        # 타이틀
        tk.Label(dialog, text="다운로드할 키워드를 선택하세요",
                 font=FONT_BODY_BOLD, fg=C_TEXT, bg=C_BG).pack(pady=(12, 6))

        # 전체선택 / 선택해제 버튼 프레임
        btn_frame = tk.Frame(dialog, bg=C_BG)
        btn_frame.pack(fill="x", padx=16, pady=(0, 4))

        check_vars = {}  # {keyword: BooleanVar}

        def select_all():
            for v in check_vars.values():
                v.set(True)

        def deselect_all():
            for v in check_vars.values():
                v.set(False)

        tk.Button(btn_frame, text="전체선택", font=FONT_SMALL, fg=C_TEXT, bg=C_ACCENT,
                  relief="flat", cursor="hand2", command=select_all).pack(side="left", padx=(0, 6))
        tk.Button(btn_frame, text="선택해제", font=FONT_SMALL, fg=C_TEXT, bg=C_ACCENT,
                  relief="flat", cursor="hand2", command=deselect_all).pack(side="left")

        # 체크박스 목록 (스크롤 가능)
        list_frame = tk.Frame(dialog, bg=C_CARD, highlightbackground=C_BORDER, highlightthickness=1)
        list_frame.pack(fill="both", expand=True, padx=16, pady=4)

        canvas = tk.Canvas(list_frame, bg=C_CARD, highlightthickness=0)
        scrollbar = tk.Scrollbar(list_frame, orient="vertical", command=canvas.yview)
        inner = tk.Frame(canvas, bg=C_CARD)

        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        for item in self._queue_items:
            kw = item["keyword"]
            if item["status"] == "done" and kw in self._keyword_results:
                var = tk.BooleanVar(value=True)
                check_vars[kw] = var
                cb = tk.Checkbutton(inner, text=f"{kw}  ({item['count']}건)",
                                    variable=var, font=FONT_BODY,
                                    bg=C_CARD, fg=C_TEXT, selectcolor=C_INPUT_BG,
                                    activebackground=C_CARD, activeforeground=C_TEXT,
                                    anchor="w")
                cb.pack(fill="x", padx=8, pady=2)

        if not check_vars:
            tk.Label(inner, text="다운로드 가능한 결과가 없습니다.",
                     font=FONT_BODY, fg=C_SUBTEXT, bg=C_CARD).pack(padx=8, pady=8)

        # 하단 버튼
        bottom = tk.Frame(dialog, bg=C_BG)
        bottom.pack(fill="x", padx=16, pady=(4, 12))

        def do_download():
            selected = [kw for kw, var in check_vars.items() if var.get()]
            if not selected:
                messagebox.showwarning("알림", "키워드를 1개 이상 선택하세요.", parent=dialog)
                return

            path = filedialog.asksaveasfilename(
                parent=dialog,
                title="결과 파일 저장",
                filetypes=[("엑셀 파일", "*.xlsx"), ("모든 파일", "*.*")],
                defaultextension=".xlsx",
                initialfile="크롤링결과.xlsx",
            )
            if not path:
                return

            try:
                import openpyxl
                dst_wb = openpyxl.Workbook()
                # 기본 시트 제거
                if dst_wb.sheetnames:
                    dst_wb.remove(dst_wb.active)

                for kw in selected:
                    src_path = self._keyword_results.get(kw)
                    if not src_path or not os.path.isfile(src_path):
                        continue
                    try:
                        src_wb = openpyxl.load_workbook(src_path)
                        src_ws = src_wb.active
                    except Exception:
                        continue
                    # 시트명 안전처리
                    safe_name = kw[:31]
                    for ch in ['\\', '/', '*', '?', ':', '[', ']']:
                        safe_name = safe_name.replace(ch, '_')
                    # 중복 시트명 처리
                    base_name = safe_name
                    n = 1
                    while safe_name in dst_wb.sheetnames:
                        suffix = f"_{n}"
                        safe_name = base_name[:31 - len(suffix)] + suffix
                        n += 1
                    dst_ws = dst_wb.create_sheet(title=safe_name)
                    for row in src_ws.iter_rows(values_only=True):
                        dst_ws.append(row)
                    src_wb.close()

                dst_wb.save(path)
                dst_wb.close()
                messagebox.showinfo("완료", f"선택한 {len(selected)}개 키워드 결과 저장 완료!\n\n{path}", parent=dialog)
                self._append_log(f"선택 다운로드 저장: {path} ({len(selected)}개 키워드)")
                dialog.destroy()
            except Exception as e:
                messagebox.showerror("오류", f"저장 실패: {e}", parent=dialog)

        # 전체 다운로드 (기존 병합파일 그대로)
        def do_download_all():
            dialog.destroy()
            self._save_file_dialog(self._temp_output)

        tk.Button(bottom, text="선택 다운로드", font=FONT_BTN, fg=C_TEXT, bg=C_POINT,
                  relief="flat", cursor="hand2", padx=16, pady=6,
                  command=do_download).pack(side="right", padx=(6, 0))
        tk.Button(bottom, text="전체 다운로드", font=FONT_BODY_BOLD, fg=C_TEXT, bg=C_ACCENT,
                  relief="flat", cursor="hand2", padx=12, pady=6,
                  command=do_download_all).pack(side="right")

    # ═══════════════════════════════════════════
    # 키워드 큐 관리
    # ═══════════════════════════════════════════

    def _queue_add(self):
        """입력창에서 키워드를 큐에 추가 (쉼표 구분 지원)"""
        text = self.keyword_var.get().strip()
        if not text:
            return
        keywords = [kw.strip() for kw in text.split(",") if kw.strip()]
        for kw in keywords:
            # 중복 체크
            if any(item["keyword"] == kw for item in self._queue_items):
                continue
            self._queue_items.append({"keyword": kw, "status": "pending", "count": 0})
        self.keyword_var.set("")
        self._queue_refresh_listbox()

    def _queue_remove(self):
        """선택한 키워드 삭제 (진행중이 아닌 것만)"""
        sel = self.queue_listbox.curselection()
        if not sel:
            return
        idx = sel[0]
        if idx < len(self._queue_items):
            if self._queue_items[idx]["status"] == "running":
                messagebox.showwarning("경고", "진행중인 키워드는 삭제할 수 없습니다.")
                return
            self._queue_items.pop(idx)
            self._queue_refresh_listbox()

    def _queue_clear(self):
        """대기중인 키워드 전체 삭제 (진행중/완료 제외)"""
        self._queue_items = [item for item in self._queue_items
                             if item["status"] == "running"]
        self._queue_refresh_listbox()

    def _queue_move_up(self):
        """선택 항목 위로 이동"""
        sel = self.queue_listbox.curselection()
        if not sel or sel[0] == 0:
            return
        idx = sel[0]
        self._queue_items[idx - 1], self._queue_items[idx] = \
            self._queue_items[idx], self._queue_items[idx - 1]
        self._queue_refresh_listbox()
        self.queue_listbox.selection_set(idx - 1)

    def _queue_move_down(self):
        """선택 항목 아래로 이동"""
        sel = self.queue_listbox.curselection()
        if not sel or sel[0] >= len(self._queue_items) - 1:
            return
        idx = sel[0]
        self._queue_items[idx + 1], self._queue_items[idx] = \
            self._queue_items[idx], self._queue_items[idx + 1]
        self._queue_refresh_listbox()
        self.queue_listbox.selection_set(idx + 1)

    def _queue_refresh_listbox(self):
        """큐 리스트박스 갱신"""
        self.queue_listbox.delete(0, "end")
        if not self._queue_items:
            self.queue_listbox.insert("end", "  키워드를 추가하세요 (예: 수원 피부과)")
            self.queue_listbox.itemconfig(0, fg=C_SUBTEXT)
            self.queue_progress_label.config(text="0/0 키워드 완료")
            return
        done_count = 0
        for item in self._queue_items:
            kw = item["keyword"]
            status = item["status"]
            count = item["count"]
            if status == "pending":
                prefix = "⏳"
                suffix = ""
            elif status == "running":
                prefix = "🔄"
                suffix = ""
            elif status == "done":
                prefix = "✅"
                suffix = f" ({count}건)"
                done_count += 1
            else:  # failed
                prefix = "❌"
                suffix = ""
            self.queue_listbox.insert("end", f" {prefix}  {kw}{suffix}")

            # 색상 설정
            idx = self.queue_listbox.size() - 1
            if status == "done":
                self.queue_listbox.itemconfig(idx, fg=C_SUCCESS)
            elif status == "failed":
                self.queue_listbox.itemconfig(idx, fg="#ff5252")
            elif status == "running":
                self.queue_listbox.itemconfig(idx, fg="#ffb347")

        total = len(self._queue_items)
        self.queue_progress_label.config(text=f"{done_count}/{total} 키워드 완료")
        self._update_start_btn_text()

    def _update_start_btn_text(self):
        """큐 상태에 따라 시작 버튼 텍스트 동적 변경"""
        if self.start_btn.cget("state") == "disabled":
            return
        has_pending = any(i["status"] == "pending" for i in self._queue_items)
        has_done_or_failed = any(i["status"] in ("done", "failed") for i in self._queue_items)
        if has_pending and has_done_or_failed:
            self.start_btn.config(text="\u25b6  이어하기")
        else:
            self.start_btn.config(text="\u25b6  시작")

    # ═══════════════════════════════════════════
    # 크롤링 제어
    # ═══════════════════════════════════════════

    def _get_engine_config(self):
        """프록시/API/딜레이 설정을 검증하고 반환. 실패 시 None."""
        proxy_file = None
        if self.proxy_mode_var.get() == "프록시 사용":
            pf = self.proxy_file_var.get().strip()
            if pf and os.path.isfile(pf):
                proxy_file = pf
            else:
                messagebox.showerror("오류", "프록시 파일을 선택하세요.")
                return None

        api_keys = None
        if self.api_mode_var.get() == "네이버 검색 API":
            id1 = self.api_id1_var.get().strip()
            sec1 = self.api_secret1_var.get().strip()
            if not id1 or not sec1:
                messagebox.showerror("오류", "API Client ID와 Secret을 입력하세요.")
                return None
            api_keys = [{"client_id": id1, "client_secret": sec1}]
            id2 = self.api_id2_var.get().strip()
            sec2 = self.api_secret2_var.get().strip()
            if id2 and sec2:
                api_keys.append({"client_id": id2, "client_secret": sec2})

        delay_min = self.delay_min_var.get()
        delay_max = self.delay_max_var.get()
        timeout = self.timeout_var.get()
        if delay_max <= delay_min:
            delay_max = delay_min * 2.5

        return {
            "proxy_file": proxy_file,
            "api_keys": api_keys,
            "delay_min": delay_min,
            "delay_max": delay_max,
            "timeout": timeout,
        }

    def _start_crawl(self):
        self._save_config()

        # 큐에 pending 키워드가 있으면 큐 모드, 없으면 입력창 키워드로 단일 실행
        has_queue = len([i for i in self._queue_items if i["status"] == "pending"]) > 0
        if not has_queue:
            keyword = self.keyword_var.get().strip()
            if not keyword:
                messagebox.showerror("오류", "키워드를 입력하세요.\n또는 큐에 키워드를 추가하세요.")
                return

        cfg = self._get_engine_config()
        if cfg is None:
            return

        # 이어하기 여부 판단 (큐에 done/failed가 이미 있으면 이어하기)
        is_resume = has_queue and any(i["status"] in ("done", "failed") for i in self._queue_items)

        # 임시 출력 파일: 이어하기면 기존 파일 재사용
        if is_resume and self._temp_output and os.path.isfile(self._temp_output):
            output_file = self._temp_output
        else:
            import tempfile
            fd, temp_path = tempfile.mkstemp(suffix=".xlsx", prefix="crawl_result_")
            os.close(fd)
            self._temp_output = temp_path
            output_file = temp_path

        # UI 상태 전환
        self.start_btn.config(state="disabled")
        self.stop_btn.config(state="normal", fg=C_TEXT)
        self.progress_var.set(0)
        self.progress_bar.set(0)

        if is_resume:
            self._append_log(f"\n{'='*50}")
            self._append_log("\u25b6 이어하기 \u2014 이전 결과에 이어서 진행합니다.")
            self._append_log(f"{'='*50}")
        else:
            self.download_btn.config(state="disabled")
            self._clear_log()
            self._update_stats({})

        if has_queue:
            # 큐 모드
            self._queue_running = True
            self._queue_stop_requested = False
            self.thread = threading.Thread(
                target=self._run_queue_crawl,
                args=(output_file, cfg),
                daemon=True,
            )
        else:
            # 단일 키워드 모드
            self.engine = CrawlerEngine(
                proxy_file=cfg["proxy_file"],
                delay_min=cfg["delay_min"],
                delay_max=cfg["delay_max"],
                timeout=cfg["timeout"],
                callback=self._engine_callback,
                api_keys=cfg["api_keys"],
            )
            keyword = self.keyword_var.get().strip()
            start_page = self.start_page_var.get()
            max_pages = self.max_pages_var.get()
            self.thread = threading.Thread(
                target=self._run_keyword_crawl,
                args=(keyword, output_file, start_page, max_pages),
                daemon=True,
            )
        self.thread.start()

    def _run_keyword_crawl(self, keyword, output_file, start_page, max_pages):
        try:
            self.engine.run_keyword_search(keyword, output_file, start_page, max_pages)
        except Exception as e:
            self._safe_callback("log", f"오류 발생: {e}")
        finally:
            self._safe_callback("finished", None)

    def _run_queue_crawl(self, output_file, cfg):
        """큐의 키워드를 순차적으로 크롤링, 결과를 시트별로 하나의 xlsx에 저장"""
        import tempfile
        try:
            import openpyxl
        except ImportError:
            openpyxl = None

        start_page = self.start_page_var.get()
        max_pages = self.max_pages_var.get()
        pending_items = [i for i, item in enumerate(self._queue_items) if item["status"] == "pending"]

        for queue_idx in pending_items:
            if self._queue_stop_requested:
                self._safe_callback("log", "큐 중지됨. 남은 키워드는 대기 상태 유지.")
                break

            item = self._queue_items[queue_idx]
            keyword = item["keyword"]

            # 상태 업데이트: 진행중
            item["status"] = "running"
            self._queue_current_idx = queue_idx
            self.root.after(0, self._queue_refresh_listbox)

            self._safe_callback("log", f"\n{'='*50}")
            self._safe_callback("log", f"큐 [{queue_idx+1}/{len(self._queue_items)}] 키워드: {keyword}")
            self._safe_callback("log", f"{'='*50}")

            # 각 키워드별 임시 파일
            fd, kw_temp = tempfile.mkstemp(suffix=".xlsx", prefix=f"crawl_kw_")
            os.close(fd)

            # 엔진 생성 (매 키워드마다 새로 생성)
            self.engine = CrawlerEngine(
                proxy_file=cfg["proxy_file"],
                delay_min=cfg["delay_min"],
                delay_max=cfg["delay_max"],
                timeout=cfg["timeout"],
                callback=self._engine_callback,
                api_keys=cfg["api_keys"],
            )

            # 진행률 리셋
            self.root.after(0, lambda: self.progress_bar.set(0))
            self.root.after(0, lambda: self.progress_label.config(text="0/0 (0%)"))
            self.total_rows = 0

            try:
                self.engine.run_keyword_search(keyword, kw_temp, start_page, max_pages)
                item["status"] = "done"
                # 수집 건수 파악
                count = self._count_xlsx_rows(kw_temp)
                item["count"] = count
                self._safe_callback("log", f"✅ '{keyword}' 완료 — {count}건 수집")
            except Exception as e:
                if self._queue_stop_requested:
                    # 중지에 의한 중단은 pending으로 복원 (이어하기 가능)
                    item["status"] = "pending"
                    self._safe_callback("log", f"⏸ '{keyword}' 일시중지됨 — 다음 시작 시 이어서 진행")
                else:
                    item["status"] = "failed"
                    self._safe_callback("log", f"❌ '{keyword}' 실패: {e}")

            self.root.after(0, self._queue_refresh_listbox)

            # 중간 결과를 메인 xlsx에 시트로 병합
            self._merge_to_output(kw_temp, output_file, keyword)

            # 키워드 완료마다 중간 다운로드 활성화
            self.root.after(0, lambda: self.download_btn.config(
                state="normal", fg=C_SUCCESS,
                text="\u2b07 중간결과"))

            # 키워드별 결과 파일 보존 (선택 다운로드용)
            self._keyword_results[keyword] = kw_temp

        self._queue_running = False
        self._queue_current_idx = -1

        done_count = sum(1 for i in self._queue_items if i["status"] == "done")
        total_count = len(self._queue_items)
        total_items = sum(i["count"] for i in self._queue_items if i["status"] == "done")
        pending_count = sum(1 for i in self._queue_items if i["status"] == "pending")
        self._safe_callback("log", f"\n{'='*50}")
        if self._queue_stop_requested and pending_count > 0:
            self._safe_callback("log", f"큐 일시정지: {done_count}/{total_count} 키워드 완료, "
                                f"{pending_count}건 대기 중, 총 {total_items}건 수집")
        else:
            self._safe_callback("log", f"큐 완료: {done_count}/{total_count} 키워드, 총 {total_items}건 수집")
        self._safe_callback("log", f"{'='*50}")
        self._safe_callback("done", None)
        self._safe_callback("finished", None)

    def _count_xlsx_rows(self, filepath):
        """xlsx 파일의 데이터 행 수 반환"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True)
            ws = wb.active
            count = max(0, ws.max_row - 1) if ws.max_row else 0
            wb.close()
            return count
        except Exception:
            return 0

    def _merge_to_output(self, src_file, dst_file, sheet_name):
        """src_file의 데이터를 dst_file에 시트로 추가"""
        try:
            import openpyxl
        except ImportError:
            # openpyxl 없으면 단순 복사
            import shutil
            shutil.copy2(src_file, dst_file)
            return

        # 시트명 31자 제한
        safe_name = sheet_name[:31]
        # 시트명에 사용 불가 문자 제거
        for ch in ['\\', '/', '*', '?', ':', '[', ']']:
            safe_name = safe_name.replace(ch, '_')

        try:
            src_wb = openpyxl.load_workbook(src_file)
            src_ws = src_wb.active
        except Exception:
            return

        try:
            dst_wb = openpyxl.load_workbook(dst_file)
        except Exception:
            dst_wb = openpyxl.Workbook()
            # 기본 시트 삭제
            if "Sheet" in dst_wb.sheetnames:
                del dst_wb["Sheet"]

        # 중복 시트명 처리
        if safe_name in dst_wb.sheetnames:
            del dst_wb[safe_name]

        dst_ws = dst_wb.create_sheet(title=safe_name)

        for row in src_ws.iter_rows(values_only=True):
            dst_ws.append(list(row))

        try:
            dst_wb.save(dst_file)
        except Exception as e:
            self._safe_callback("log", f"시트 저장 오류: {e}")
        finally:
            src_wb.close()
            dst_wb.close()

    def _stop_crawl(self):
        if self.engine:
            self.engine.stop()
        if self._queue_running:
            self._queue_stop_requested = True
        self.stop_btn.config(state="disabled")
        self._append_log("⏸ 일시정지 요청됨... 현재 작업 중단 후 대기합니다.")

    # ═══════════════════════════════════════════
    # 콜백 (엔진 → GUI, thread-safe)
    # ═══════════════════════════════════════════

    def _engine_callback(self, event, data):
        """엔진에서 호출 — 메인 스레드로 전달"""
        self.root.after(0, self._handle_event, event, data)

    def _safe_callback(self, event, data):
        """쓰레드에서 안전하게 GUI 업데이트"""
        self.root.after(0, self._handle_event, event, data)

    def _handle_event(self, event, data):
        if event == "log":
            self._append_log(str(data))
        elif event == "total":
            self.total_rows = data
        elif event == "progress":
            if self.total_rows > 0:
                pct = data / self.total_rows * 100
                self.progress_var.set(pct)
                self.progress_bar.set(pct)
                self.progress_label.config(text=f"{data}/{self.total_rows} ({pct:.0f}%)")
        elif event == "stats":
            self._update_stats(data)
        elif event == "done":
            if not self._queue_running:
                # 단일 키워드 모드에서만 팝업
                self._append_log(f"\n크롤링 완료!")
                self.download_btn.config(state="normal", fg=C_SUCCESS,
                                          text="\u2b07 다운로드")
                messagebox.showinfo("완료", "크롤링 완료!\n\n'결과 다운로드' 버튼을 눌러 저장하세요.")
            else:
                # 큐 모드 — 팝업 없이 다운로드 버튼만 활성화
                self.download_btn.config(state="normal", fg=C_SUCCESS,
                                          text="\u2b07 다운로드")
        elif event == "finished":
            self.start_btn.config(state="normal")
            self.stop_btn.config(state="disabled")
            self._update_start_btn_text()

    def _append_log(self, text):
        self.log_text.config(state="normal")
        # 태그 자동 분류
        tag = None
        lower = text.lower()
        if "오류" in lower or "error" in lower or "실패" in lower:
            tag = "ERROR"
        elif "완료" in lower or "success" in lower or "성공" in lower:
            tag = "SUCCESS"
        elif "경고" in lower or "warn" in lower:
            tag = "WARN"
        else:
            tag = "INFO"
        self.log_text.insert("end", text + "\n", tag)
        self.log_text.see("end")
        self.log_text.config(state="disabled")

    def _clear_log(self):
        self.log_text.config(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.config(state="disabled")

    def _update_stats(self, stats):
        phone = stats.get("phone", 0)
        hp = stats.get("hp", 0)
        email = stats.get("email", 0)
        naver_id = stats.get("naver_id", 0)
        success = stats.get("success", 0)
        blocked = stats.get("blocked", 0)

        self.stat_cards["안심번호"].set_value(phone)
        self.stat_cards["홈페이지"].set_value(hp)
        self.stat_cards["이메일"].set_value(email)
        self.stat_cards["네이버ID"].set_value(naver_id)
        self.stat_cards["성공"].set_value(success)
        self.stat_cards["차단"].set_value(blocked)


def main():
    root = tk.Tk()

    # 다크 테마이므로 ttk 스타일 최소 설정
    root.option_add("*TCombobox*Listbox.background", C_INPUT_BG)
    root.option_add("*TCombobox*Listbox.foreground", C_INPUT_FG)

    app = CrawlerGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
