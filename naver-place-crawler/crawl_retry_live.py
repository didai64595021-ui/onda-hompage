"""Retry failed keywords — 4-slot parallel + real-time Google Sheets + Telegram."""
import sys, os, json, glob, time, random, threading

sys.stdout.reconfigure(encoding='utf-8', errors='replace', line_buffering=True)
sys.stderr.reconfigure(encoding='utf-8', errors='replace')
os.chdir(os.path.dirname(os.path.abspath(__file__)))

from concurrent.futures import ThreadPoolExecutor, as_completed
import gspread
from crawler_engine import CrawlerEngine

# === Config ===
PARALLEL_KEYWORDS = 4
SA_PATH = os.path.join(os.path.dirname(__file__), "google-sa.json")
SHEET_ID = "1mwavPlXeZtPXBD-DgixyfS61a8P9wlVIx09L5V64QbY"
WS_NAME = "플레이스크롤링_실시간"
BOT_TOKEN = "8574880668:AAHb75dmkFchbjBNj7VgPZuKrptFgIjQ_es"
CHAT_ID = "7383805736"
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "retry_results")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# === Google Sheets ===
gc = gspread.service_account(filename=SA_PATH)
sh = gc.open_by_key(SHEET_ID)
ws = sh.worksheet(WS_NAME)
sheet_lock = threading.Lock()
SHEET_COLS = ['업종(키워드)', '상호명', '네이버아이디', '업체이메일', '안심번호', '010전화', '업체주소',
              '홈페이지URL', '홈페이지반응형', '전화버튼없음', '웹빌더', '톡톡', '톡톡URL',
              '카카오톡', '인스타그램', '방문자리뷰수', '블로그리뷰수', '리뷰합계', '신규업체',
              '고유번호', '플레이스URL', '업데이트날짜']
print(f"Sheet: {sh.url}#gid={ws.id}", flush=True)

def push_to_sheet(rows_data):
    rows = []
    for r in rows_data:
        # 빈 행/불완전 행 스킵
        name = r.get("상호명", "").strip()
        pid = r.get("고유번호", "").strip()
        if not name or not pid:
            continue
        # 칼럼 정합성 검증: 안심번호에 주소가 들어있으면 행이 밀린 것 → 스킵
        tel = r.get("안심번호", "").strip()
        if tel and len(tel) > 5 and not any(c.isdigit() for c in tel[:4]):
            continue
        rows.append([str(r.get(c, "")).strip() for c in SHEET_COLS])
    if rows:
        with sheet_lock:
            for attempt in range(3):
                try:
                    ws.append_rows(rows, value_input_option='RAW')
                    break
                except Exception as e:
                    if attempt < 2:
                        import time as _t
                        _t.sleep(5 * (attempt + 1))
                    else:
                        print(f"  [SHEET ERROR] {e}", flush=True)

# === Telegram ===
import urllib.request
def tg(text):
    try:
        payload = json.dumps({"chat_id": CHAT_ID, "text": text}).encode("utf-8")
        req = urllib.request.Request(
            f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
            data=payload, headers={"Content-Type": "application/json"})
        urllib.request.urlopen(req, timeout=10)
    except:
        pass

# === Collect failed keywords ===
pfiles = sorted(
    glob.glob("C:/Users/anytwo/AppData/Local/Temp/crawl_kw_*-progress.json"),
    key=os.path.getmtime, reverse=True)
recent = [pf for pf in pfiles if os.path.getmtime(pf) > 1744300000]

failed_keywords = []
seen = set()
for pf in recent:
    try:
        with open(pf, "r", encoding="utf-8") as f:
            data = json.load(f)
        kw = data.get("keyword", "").strip().replace("\n", " ").replace("  ", " ")
        rows = len(data.get("rows", []))
        if kw and kw not in seen and rows <= 10:
            seen.add(kw)
            failed_keywords.append(kw)
    except:
        pass

total = len(failed_keywords)
print(f"Failed keywords: {total}, Parallel: {PARALLEL_KEYWORDS}", flush=True)

tg(f"🔄 플레이스 크롤러 재시도 시작\n\n"
   f"대상: {total}개 키워드 / {PARALLEL_KEYWORDS}병렬\n"
   f"Sheet: https://docs.google.com/spreadsheets/d/{SHEET_ID}#gid={ws.id}")

# === Shared state ===
results_lock = threading.Lock()
results = {"success": 0, "partial": 0, "fail": 0, "total_rows": 0, "total_010": 0, "done": 0}

def process_keyword(idx, keyword):
    """Process a single keyword — called from thread pool."""
    slot = threading.current_thread().name
    print(f"[{idx+1}/{total}] [{slot}] '{keyword}' 시작", flush=True)

    output_file = os.path.join(OUTPUT_DIR, f"{keyword.replace(' ', '_')}.xlsx")

    def cb(event, data=None):
        if event == "log":
            if "talk-debug" in str(data):
                return
            s = str(data)
            # Only log important events
            if any(k in s for k in ['PHASE', '완료', '차단', '010:', 'Selenium']):
                try:
                    print(f"  [{slot}] {data}", flush=True)
                except:
                    pass

    engine = CrawlerEngine(callback=cb, delay_min=0.5, delay_max=1.5)
    # 이전 progress 파일 삭제 (오래된 포맷 복구 방지)
    _pf = output_file.rsplit(".", 1)[0] + "-progress.json"
    if os.path.isfile(_pf):
        try: os.remove(_pf)
        except: pass
    try:
        engine.run_keyword_search(keyword=keyword, output_file=output_file, start_page=1, max_pages=1)
    except Exception as e:
        print(f"  [{slot}] ENGINE ERROR: {e}", flush=True)

    # Always try to read result and push to sheet (even after error)
    count = 0
    count_010 = 0
    try:
        if os.path.isfile(output_file) and os.path.getsize(output_file) > 0:
            import openpyxl
            wb = openpyxl.load_workbook(output_file, read_only=True)
            for sn in wb.sheetnames:
                ws_xl = wb[sn]
                rows = list(ws_xl.iter_rows(min_row=1, values_only=True))
                if len(rows) < 2:
                    continue
                header = [str(c).strip() if c else "" for c in rows[0]]
                data_rows = []
                for r in rows[1:]:
                    row_dict = {header[j]: str(r[j]).strip() if j < len(r) and r[j] else "" for j in range(len(header))}
                    data_rows.append(row_dict)
                    if row_dict.get("010전화") and row_dict["010전화"] not in ("", "X", "None"):
                        count_010 += 1
                count = len(data_rows)
                push_to_sheet(data_rows)
            wb.close()
    except Exception as e:
        print(f"  [{slot}] READ/PUSH ERROR: {e}", flush=True)

    status = "success" if count >= 10 else ("partial" if count > 0 else "fail")
    with results_lock:
        results[status] += 1
        results["total_rows"] += count
        results["total_010"] += count_010
        results["done"] += 1
        done = results["done"]

    print(f"  [{slot}] '{keyword}' -> {count}건, 010={count_010} [{status}]", flush=True)

    if done % 10 == 0:
        with results_lock:
            r = dict(results)
        tg(f"📊 크롤링 진행: {r['done']}/{total}\n"
           f"성공={r['success']}, 부분={r['partial']}, 실패={r['fail']}\n"
           f"누적: {r['total_rows']}건, 010={r['total_010']}건")

# === Run 4-parallel ===
print(f"\n🔥 {PARALLEL_KEYWORDS}병렬 크롤링 시작", flush=True)

with ThreadPoolExecutor(max_workers=PARALLEL_KEYWORDS, thread_name_prefix="S") as executor:
    futures = {executor.submit(process_keyword, i, kw): kw for i, kw in enumerate(failed_keywords)}
    for future in as_completed(futures):
        kw = futures[future]
        try:
            future.result()
        except Exception as e:
            print(f"  THREAD ERROR [{kw}]: {e}", flush=True)

# === Final report ===
r = results
summary = (
    f"✅ 플레이스 크롤러 재시도 완료\n\n"
    f"대상: {total}개 / {PARALLEL_KEYWORDS}병렬\n"
    f"성공: {r['success']}개\n"
    f"부분: {r['partial']}개\n"
    f"실패: {r['fail']}개\n"
    f"총 수집: {r['total_rows']}건\n"
    f"010 추출: {r['total_010']}건 ({r['total_010']*100//max(r['total_rows'],1)}%)\n\n"
    f"Sheet: https://docs.google.com/spreadsheets/d/{SHEET_ID}#gid={ws.id}")
tg(summary)
print(f"\n{summary}", flush=True)
