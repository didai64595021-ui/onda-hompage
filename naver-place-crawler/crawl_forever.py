#!/usr/bin/env python
"""24시간 무중단 크롤링 — 업종별 시트 분리, 자동 디버깅, 텔레그램 보고.
멈추라고 할 때까지 계속 돌림. 키워드 소진 시 새 업종 자동 시작."""
import sys, os, json, time, threading, glob, random

sys.stdout.reconfigure(encoding='utf-8', errors='replace', line_buffering=True)
sys.stderr.reconfigure(encoding='utf-8', errors='replace')
os.chdir(os.path.dirname(os.path.abspath(__file__)))

from concurrent.futures import ThreadPoolExecutor, as_completed
import gspread
from crawler_engine import CrawlerEngine
import urllib.request

# ═══════════════════════════════════════════
# CONFIG
# ═══════════════════════════════════════════
PARALLEL = 4
SA_PATH = "google-sa.json"
SHEET_ID = "1T3zN0mecMdfxuaNQLa3wGP2sFgCWI38wrcU8iq3Gz9U"
BOT_TOKEN = "8574880668:AAHb75dmkFchbjBNj7VgPZuKrptFgIjQ_es"
CHAT_ID = "-1003806737505"
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 구/동 세분화 데이터
import json as _json
with open(os.path.join(BASE_DIR, "districts_data.json"), "r", encoding="utf-8") as _f:
    DISTRICTS = _json.load(_f)

# ═══════════════════════════════════════════
# KEYWORD SETS (업종별)
# ═══════════════════════════════════════════
REGIONS_FULL = [
    "강남","서초","마포","송파","강서","영등포","성동","용산","종로","노원","관악","강북","서대문","동작",
    "수원","성남","용인","화성","고양","부천","안양","평택","의정부","김포","광명","하남","파주","분당","일산","판교",
    "인천","부평","송도","청라",
    "부산","해운대","서면",
    "대구","대전","광주","울산","창원",
    "세종","천안","전주","청주","제주","원주","춘천",
]

JOBS = [
    {
        "name": "헬스PT_전국",
        "sheet_tab": "헬스PT_전국",
        "keywords": [f"{r} 헬스장" for r in REGIONS_FULL] + [f"{r} PT" for r in REGIONS_FULL],
        "output_dir": os.path.join(BASE_DIR, "results_health"),
    },
    {
        "name": "네일_전국",
        "sheet_tab": "네일_전국",
        "keywords": [f"{r} 네일샵" for r in REGIONS_FULL],
        "output_dir": os.path.join(BASE_DIR, "results_nail"),
    },
    {
        "name": "피부과_전국",
        "sheet_tab": "피부과_전국",
        "keywords": [f"{r} 피부과" for r in REGIONS_FULL],
        "output_dir": os.path.join(BASE_DIR, "results_derma"),
    },
    {
        "name": "미용실_전국",
        "sheet_tab": "미용실_전국",
        "keywords": [f"{r} 미용실" for r in REGIONS_FULL],
        "output_dir": os.path.join(BASE_DIR, "results_hair"),
    },
    {
        "name": "필라테스_전국",
        "sheet_tab": "필라테스_전국",
        "keywords": [f"{r} 필라테스" for r in REGIONS_FULL],
        "output_dir": os.path.join(BASE_DIR, "results_pilates"),
    },
    {
        "name": "맛집_재시도",
        "sheet_tab": "맛집_재시도",
        "keywords": None,
        "output_dir": os.path.join(BASE_DIR, "results_retry"),
    },
    {
        "name": "인테리어_전국",
        "sheet_tab": "인테리어_전국",
        "keywords": [f"{r} 인테리어" for r in REGIONS_FULL],
        "output_dir": os.path.join(BASE_DIR, "results_interior"),
    },
    {
        "name": "속눈썹_전국",
        "sheet_tab": "속눈썹_전국",
        "keywords": [f"{r} 속눈썹" for r in REGIONS_FULL],
        "output_dir": os.path.join(BASE_DIR, "results_lashes"),
    },
    {
        "name": "왁싱_전국",
        "sheet_tab": "왁싱_전국",
        "keywords": [f"{r} 왁싱" for r in REGIONS_FULL],
        "output_dir": os.path.join(BASE_DIR, "results_waxing"),
    },
    {
        "name": "마사지_전국",
        "sheet_tab": "마사지_전국",
        "keywords": [f"{r} 마사지" for r in REGIONS_FULL],
        "output_dir": os.path.join(BASE_DIR, "results_massage"),
    },
    {
        "name": "에스테틱_전국",
        "sheet_tab": "에스테틱_전국",
        "keywords": [f"{r} 에스테틱" for r in REGIONS_FULL],
        "output_dir": os.path.join(BASE_DIR, "results_esthetic"),
    },
    {
        "name": "반영구_전국",
        "sheet_tab": "반영구_전국",
        "keywords": [f"{r} 반영구" for r in REGIONS_FULL],
        "output_dir": os.path.join(BASE_DIR, "results_semi"),
    },
    {
        "name": "치과_전국",
        "sheet_tab": "치과_전국",
        "keywords": [f"{r} 치과" for r in REGIONS_FULL],
        "output_dir": os.path.join(BASE_DIR, "results_dental"),
    },
    {
        "name": "한의원_전국",
        "sheet_tab": "한의원_전국",
        "keywords": [f"{r} 한의원" for r in REGIONS_FULL],
        "output_dir": os.path.join(BASE_DIR, "results_oriental"),
    },
    {
        "name": "학원_전국",
        "sheet_tab": "학원_전국",
        "keywords": [f"{r} 학원" for r in REGIONS_FULL] + [f"{r} 보습학원" for r in REGIONS_FULL],
        "output_dir": os.path.join(BASE_DIR, "results_academy"),
    },
    {
        "name": "꽃집_전국",
        "sheet_tab": "꽃집_전국",
        "keywords": [f"{r} 꽃집" for r in REGIONS_FULL],
        "output_dir": os.path.join(BASE_DIR, "results_flower"),
    },
    {
        "name": "사진관_전국",
        "sheet_tab": "사진관_전국",
        "keywords": [f"{r} 사진관" for r in REGIONS_FULL] + [f"{r} 스튜디오" for r in REGIONS_FULL],
        "output_dir": os.path.join(BASE_DIR, "results_photo"),
    },
]

# ═══════════════════════════════════════════
# GOOGLE SHEETS
# ═══════════════════════════════════════════
gc = gspread.service_account(filename=SA_PATH)
sh = gc.open_by_key(SHEET_ID)
sheet_lock = threading.Lock()
SHEET_COLS = ['지역','업종(키워드)','상호명','네이버아이디','업체이메일','안심번호','010전화','업체주소',
              '홈페이지URL','홈페이지반응형','전화버튼없음','웹빌더','톡톡','톡톡URL',
              '카카오톡','인스타그램','방문자리뷰수','블로그리뷰수','리뷰합계','신규업체',
              '고유번호','플레이스URL','업데이트날짜']

def get_or_create_tab(tab_name):
    try:
        return sh.worksheet(tab_name)
    except:
        ws = sh.add_worksheet(title=tab_name, rows=10000, cols=25)
        ws.append_row(SHEET_COLS)
        ws.format('1:1', {'textFormat': {'bold': True}})
        ws.freeze(rows=1)
        return ws

def push_to_sheet(ws, rows_data, region=""):
    rows = []
    for r in rows_data:
        if not r.get("상호명", "").strip() or not r.get("고유번호", "").strip():
            continue
        tel = r.get("안심번호", "").strip()
        if tel and len(tel) > 5 and not any(c.isdigit() for c in tel[:4]):
            continue
        row = [region] + [str(r.get(c, "")).strip() for c in SHEET_COLS[1:]]
        rows.append(row)
    if rows:
        with sheet_lock:
            for attempt in range(3):
                try:
                    ws.append_rows(rows, value_input_option='RAW')
                    break
                except Exception as e:
                    if attempt < 2:
                        time.sleep(5 * (attempt + 1))
                    else:
                        print(f"  [SHEET ERROR] {e}", flush=True)

# ═══════════════════════════════════════════
# TELEGRAM
# ═══════════════════════════════════════════
def tg(text):
    try:
        payload = json.dumps({"chat_id": CHAT_ID, "text": text}).encode("utf-8")
        req = urllib.request.Request(f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
                                     data=payload, headers={"Content-Type": "application/json"})
        urllib.request.urlopen(req, timeout=10)
    except: pass

# ═══════════════════════════════════════════
# CRAWL ONE JOB
# ═══════════════════════════════════════════
def run_job(job):
    name = job["name"]
    tab_name = job["sheet_tab"]
    output_dir = job["output_dir"]
    os.makedirs(output_dir, exist_ok=True)

    ws = get_or_create_tab(tab_name)

    # Collect keywords
    if job["keywords"] is None:
        # 맛집 재시도: progress.json에서 수집
        keywords = []
        seen = set()
        for pf in sorted(glob.glob("C:/Users/anytwo/AppData/Local/Temp/crawl_kw_*-progress.json"),
                         key=os.path.getmtime, reverse=True):
            try:
                with open(pf, "r", encoding="utf-8") as f:
                    data = json.load(f)
                kw = data.get("keyword", "").strip().replace("\n", " ").replace("  ", " ")
                rows = len(data.get("rows", []))
                if kw and kw not in seen and rows <= 10:
                    seen.add(kw)
                    keywords.append(kw)
            except: pass
    else:
        keywords = list(job["keywords"])

    total = len(keywords)
    if total == 0:
        print(f"[{name}] 키워드 없음, 스킵", flush=True)
        return

    print(f"\n{'='*60}", flush=True)
    print(f"[{name}] {total}개 키워드, {PARALLEL}병렬", flush=True)
    print(f"{'='*60}", flush=True)

    tg(f"🚀 [{name}] 크롤링 시작\n대상: {total}개 / {PARALLEL}병렬\nSheet: {tab_name} 탭")

    results_lock = threading.Lock()
    stats = {"success": 0, "partial": 0, "fail": 0, "total_rows": 0, "total_010": 0, "done": 0}

    def process_kw(idx, keyword):
        slot = threading.current_thread().name
        parts = keyword.split(" ", 1)
        region = parts[0]
        biz_type = parts[1] if len(parts) > 1 else keyword

        def cb(event, data=None):
            if event == "log":
                if "talk-debug" in str(data): return
                s = str(data)
                if any(k in s for k in ['PHASE', '010:', '차단']):
                    try: print(f"  [{slot}] {data}", flush=True)
                    except: pass

        def crawl_one(kw, out):
            pf = out.rsplit(".", 1)[0] + "-progress.json"
            if os.path.isfile(pf):
                try: os.remove(pf)
                except: pass
            eng = CrawlerEngine(callback=cb, delay_min=0.5, delay_max=1.5)
            try:
                eng.run_keyword_search(keyword=kw, output_file=out, start_page=1, max_pages=0)
            except Exception as e:
                print(f"  [{slot}] ERROR: {e}", flush=True)

        # 1차 크롤링
        output_file = os.path.join(output_dir, f"{keyword.replace(' ', '_')}.xlsx")
        crawl_one(keyword, output_file)

        # 300건 cap 감지 → 구 세분화
        first_count = 0
        try:
            if os.path.isfile(output_file) and os.path.getsize(output_file) > 0:
                import openpyxl as _xl
                _wb = _xl.load_workbook(output_file, read_only=True)
                for _sn in _wb.sheetnames:
                    first_count = _wb[_sn].max_row - 1
                    break
                _wb.close()
        except: pass

        if first_count >= 270:
            for city, districts in DISTRICTS.items():
                if region in city or city in region:
                    sub_gus = list(districts.keys()) if isinstance(districts, dict) else districts
                    print(f"  [{slot}] ⚠️ {first_count}건 cap → {len(sub_gus)}개 구 세분화", flush=True)
                    for gu in sub_gus:
                        sub_kw = f"{gu} {biz_type}"
                        sub_out = os.path.join(output_dir, f"{sub_kw.replace(' ', '_')}.xlsx")
                        crawl_one(sub_kw, sub_out)
                    break

        # 모든 xlsx 결과 합산 + PID 중복제거
        count = 0
        count_010 = 0
        seen_pids = set()
        all_data = []
        try:
            import openpyxl
            # output_dir에서 이 키워드 관련 모든 xlsx 수집
            xlsx_pattern = keyword.replace(' ', '_')
            all_files = [output_file]
            if first_count >= 270:
                import glob as _g
                all_files += _g.glob(os.path.join(output_dir, f"*{biz_type.replace(' ', '_')}*.xlsx"))
            for xf in set(all_files):
                if not os.path.isfile(xf) or os.path.getsize(xf) == 0:
                    continue
                wb = openpyxl.load_workbook(xf, read_only=True)
                for sn in wb.sheetnames:
                    ws_xl = wb[sn]
                    rows = list(ws_xl.iter_rows(min_row=1, values_only=True))
                    if len(rows) < 2: continue
                    header = [str(c).strip() if c else "" for c in rows[0]]
                    for r in rows[1:]:
                        rd = {header[j]: str(r[j]).strip() if j < len(r) and r[j] else "" for j in range(len(header))}
                        pid_val = rd.get("고유번호", "")
                        if pid_val and pid_val in seen_pids:
                            continue  # 중복 제거
                        if pid_val:
                            seen_pids.add(pid_val)
                        all_data.append(rd)
                        if rd.get("010전화") and rd["010전화"] not in ("", "X", "None"):
                            count_010 += 1
                wb.close()
            count = len(all_data)
            if all_data:
                push_to_sheet(ws, all_data, region)
        except Exception as e:
            print(f"  [{slot}] READ ERROR: {e}", flush=True)

        status = "success" if count >= 10 else ("partial" if count > 0 else "fail")
        with results_lock:
            stats[status] += 1
            stats["total_rows"] += count
            stats["total_010"] += count_010
            stats["done"] += 1
            done = stats["done"]

        p = count_010 * 100 // max(count, 1)
        print(f"[{done}/{total}] '{keyword}' -> {count}건, 010={count_010}({p}%) [{status}]", flush=True)

        if done % 10 == 0:
            with results_lock:
                r = dict(stats)
            rate = r['total_010'] * 100 // max(r['total_rows'], 1)
            tg(f"📊 [{name}] {r['done']}/{total}\n"
               f"수집={r['total_rows']}건 / 010={r['total_010']}건({rate}%)")

    with ThreadPoolExecutor(max_workers=PARALLEL, thread_name_prefix="W") as ex:
        futures = {ex.submit(process_kw, i, kw): kw for i, kw in enumerate(keywords)}
        for f in as_completed(futures):
            try: f.result()
            except Exception as e:
                print(f"  THREAD ERROR: {e}", flush=True)

    r = stats
    rate = r['total_010'] * 100 // max(r['total_rows'], 1)
    tg(f"✅ [{name}] 완료\n"
       f"성공={r['success']} / 부분={r['partial']} / 실패={r['fail']}\n"
       f"총 {r['total_rows']}건 / 010={r['total_010']}건({rate}%)")
    return stats

# ═══════════════════════════════════════════
# MAIN — 무한 루프
# ═══════════════════════════════════════════
print("=" * 60, flush=True)
print("ONDA 크롤러 — 24시간 무중단 모드", flush=True)
print(f"Sheet: https://docs.google.com/spreadsheets/d/{SHEET_ID}", flush=True)
print(f"병렬: {PARALLEL}슬롯", flush=True)
print("=" * 60, flush=True)

tg(f"🔥 ONDA 크롤러 24시간 모드 시작\n\n"
   f"병렬: {PARALLEL}슬롯\n"
   f"업종: {', '.join(j['name'] for j in JOBS)}\n"
   f"Sheet: https://docs.google.com/spreadsheets/d/{SHEET_ID}")

cycle = 0
while True:
    cycle += 1
    print(f"\n{'#'*60}", flush=True)
    print(f"CYCLE {cycle} — {time.strftime('%Y-%m-%d %H:%M:%S')}", flush=True)
    print(f"{'#'*60}", flush=True)

    for job in JOBS:
        try:
            run_job(job)
        except Exception as e:
            print(f"[JOB ERROR: {job['name']}] {e}", flush=True)
            tg(f"⚠️ [{job['name']}] 에러: {str(e)[:200]}")
            time.sleep(30)

    # 한 사이클 끝나면 30분 대기 후 반복
    print(f"\n사이클 {cycle} 완료. 30분 대기 후 다시 시작...", flush=True)
    tg(f"🔄 사이클 {cycle} 완료. 30분 후 재시작.")
    time.sleep(1800)
