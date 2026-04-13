"""전국 인테리어 업종 크롤링 — 010 추출 최대화 목표."""
import sys, os, json, time, threading

sys.stdout.reconfigure(encoding='utf-8', errors='replace', line_buffering=True)
sys.stderr.reconfigure(encoding='utf-8', errors='replace')
os.chdir(os.path.dirname(os.path.abspath(__file__)))

from concurrent.futures import ThreadPoolExecutor, as_completed
import gspread
from crawler_engine import CrawlerEngine
import urllib.request

# === Config ===
PARALLEL = 2  # 맛집 4병렬이 이미 돌고 있으므로 2병렬로
SA_PATH = os.path.join(os.path.dirname(__file__), "google-sa.json")
SHEET_ID = "1mwavPlXeZtPXBD-DgixyfS61a8P9wlVIx09L5V64QbY"
BOT_TOKEN = "8574880668:AAHb75dmkFchbjBNj7VgPZuKrptFgIjQ_es"
CHAT_ID = "7383805736"
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "interior_results")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# === Keywords: 전국 주요 지역 x 인테리어 ===
REGIONS = [
    # 서울
    "강남", "서초", "마포", "송파", "강서", "영등포", "성동", "용산", "종로", "노원", "관악",
    # 경기
    "수원", "성남", "용인", "화성", "고양", "부천", "안양", "평택", "의정부", "김포", "광명", "하남", "파주",
    "분당", "일산", "판교",
    # 인천
    "인천", "부평", "송도", "청라",
    # 광역시
    "부산", "대구", "대전", "광주", "울산", "창원",
    # 기타
    "세종", "천안", "전주", "청주", "제주", "원주", "춘천",
]
KEYWORDS = [f"{r} 인테리어" for r in REGIONS]

# === Google Sheets ===
gc = gspread.service_account(filename=SA_PATH)
sh = gc.open_by_key(SHEET_ID)
WS_NAME = "인테리어_전국_010"
try:
    ws = sh.worksheet(WS_NAME)
except:
    ws = sh.add_worksheet(title=WS_NAME, rows=5000, cols=25)
    headers = ['업종(키워드)', '상호명', '네이버아이디', '업체이메일', '안심번호', '010전화', '업체주소',
               '홈페이지URL', '홈페이지반응형', '전화버튼없음', '웹빌더', '톡톡', '톡톡URL',
               '카카오톡', '인스타그램', '방문자리뷰수', '블로그리뷰수', '리뷰합계', '신규업체',
               '고유번호', '플레이스URL', '업데이트날짜']
    ws.append_row(headers)
    ws.format('1:1', {'textFormat': {'bold': True}})
    ws.freeze(rows=1)
sheet_lock = threading.Lock()
SHEET_COLS = ['업종(키워드)', '상호명', '네이버아이디', '업체이메일', '안심번호', '010전화', '업체주소',
              '홈페이지URL', '홈페이지반응형', '전화버튼없음', '웹빌더', '톡톡', '톡톡URL',
              '카카오톡', '인스타그램', '방문자리뷰수', '블로그리뷰수', '리뷰합계', '신규업체',
              '고유번호', '플레이스URL', '업데이트날짜']

print(f"Sheet: {sh.url}#gid={ws.id}", flush=True)
print(f"Keywords: {len(KEYWORDS)}, Parallel: {PARALLEL}", flush=True)

def tg(text):
    try:
        payload = json.dumps({"chat_id": CHAT_ID, "text": text}).encode("utf-8")
        req = urllib.request.Request(f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
                                     data=payload, headers={"Content-Type": "application/json"})
        urllib.request.urlopen(req, timeout=10)
    except: pass

def push_to_sheet(rows_data):
    rows = []
    for r in rows_data:
        if not r.get("상호명", "").strip() or not r.get("고유번호", "").strip():
            continue
        tel = r.get("안심번호", "").strip()
        if tel and len(tel) > 5 and not any(c.isdigit() for c in tel[:4]):
            continue
        rows.append([str(r.get(c, "")).strip() for c in SHEET_COLS])
    if rows:
        with sheet_lock:
            for attempt in range(3):
                try:
                    ws.append_rows(rows, value_input_option='RAW')
                    break
                except Exception as e:
                    if attempt < 2:
                        time.sleep(5 * (attempt + 1))
                    else:
                        print(f"  [SHEET ERROR] {e}", flush=True)

tg(f"🏠 인테리어 전국 크롤링 시작\n\n"
   f"목적: 010 추출률 테스트\n"
   f"대상: {len(KEYWORDS)}개 지역 / {PARALLEL}병렬\n"
   f"Sheet: https://docs.google.com/spreadsheets/d/{SHEET_ID}#gid={ws.id}")

# === Shared state ===
results_lock = threading.Lock()
results = {"success": 0, "partial": 0, "fail": 0, "total_rows": 0, "total_010": 0, "done": 0}
total = len(KEYWORDS)

def process_keyword(idx, keyword):
    slot = threading.current_thread().name
    print(f"[{idx+1}/{total}] [{slot}] '{keyword}'", flush=True)
    output_file = os.path.join(OUTPUT_DIR, f"{keyword.replace(' ', '_')}.xlsx")
    _pf = output_file.rsplit(".", 1)[0] + "-progress.json"
    if os.path.isfile(_pf):
        try: os.remove(_pf)
        except: pass

    def cb(event, data=None):
        if event == "log":
            if "talk-debug" in str(data): return
            s = str(data)
            if any(k in s for k in ['PHASE', '010:', 'Selenium', '차단']):
                try: print(f"  [{slot}] {data}", flush=True)
                except: pass

    engine = CrawlerEngine(callback=cb, delay_min=0.5, delay_max=1.5)
    try:
        engine.run_keyword_search(keyword=keyword, output_file=output_file, start_page=1, max_pages=1)
    except Exception as e:
        print(f"  [{slot}] ERROR: {e}", flush=True)

    count = 0
    count_010 = 0
    try:
        if os.path.isfile(output_file) and os.path.getsize(output_file) > 0:
            import openpyxl
            wb = openpyxl.load_workbook(output_file, read_only=True)
            for sn in wb.sheetnames:
                ws_xl = wb[sn]
                rows = list(ws_xl.iter_rows(min_row=1, values_only=True))
                if len(rows) < 2: continue
                header = [str(c).strip() if c else "" for c in rows[0]]
                data_rows = []
                for r in rows[1:]:
                    row_dict = {header[j]: str(r[j]).strip() if j < len(r) and r[j] else "" for j in range(len(header))}
                    data_rows.append(row_dict)
                    if row_dict.get("010전화") and row_dict["010전화"] not in ("", "X", "None"):
                        count_010 += 1
                count = len(data_rows)
                push_to_sheet(data_rows)
            wb.close()
    except Exception as e:
        print(f"  [{slot}] READ ERROR: {e}", flush=True)

    status = "success" if count >= 10 else ("partial" if count > 0 else "fail")
    with results_lock:
        results[status] += 1
        results["total_rows"] += count
        results["total_010"] += count_010
        results["done"] += 1
        done = results["done"]

    p010_pct = count_010 * 100 // max(count, 1)
    print(f"  [{slot}] '{keyword}' -> {count}건, 010={count_010}건({p010_pct}%) [{status}]", flush=True)

    if done % 10 == 0:
        with results_lock:
            r = dict(results)
        rate = r['total_010'] * 100 // max(r['total_rows'], 1)
        tg(f"🏠 인테리어 크롤링: {r['done']}/{total}\n"
           f"누적: {r['total_rows']}건, 010={r['total_010']}건({rate}%)")

# === Run ===
print(f"\n🔥 {PARALLEL}병렬 크롤링 시작", flush=True)
with ThreadPoolExecutor(max_workers=PARALLEL, thread_name_prefix="I") as executor:
    futures = {executor.submit(process_keyword, i, kw): kw for i, kw in enumerate(KEYWORDS)}
    for future in as_completed(futures):
        try: future.result()
        except Exception as e:
            print(f"  THREAD ERROR: {e}", flush=True)

r = results
rate = r['total_010'] * 100 // max(r['total_rows'], 1)
summary = (f"✅ 인테리어 전국 크롤링 완료\n\n"
           f"지역: {total}개 / {PARALLEL}병렬\n"
           f"성공: {r['success']}개\n"
           f"총 수집: {r['total_rows']}건\n"
           f"010 추출: {r['total_010']}건 ({rate}%)\n\n"
           f"Sheet: https://docs.google.com/spreadsheets/d/{SHEET_ID}#gid={ws.id}")
tg(summary)
print(f"\n{summary}", flush=True)
