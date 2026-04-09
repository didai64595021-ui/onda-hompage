#!/usr/bin/env python3
"""
crawler_engine.py — 12개 항목 전체 크롤링 엔진

입력 CSV에 "상호명"만 있으면 나머지 11개 항목 전부 자동 크롤링.

STEP 0: 고유번호 + 플레이스URL (네이버 검색 → placeId 추출)
STEP 1: 플레이스 HTML 파싱 (업종/상호명/주소/안심번호/방문자리뷰수/블로그리뷰수)
STEP 2: 홈페이지 (API 우선 → 네이버 검색 폴백)
STEP 3: 이메일 (홈페이지 HTML + 네이버 검색 폴백)
STEP 4: 네이버아이디 → @naver.com
STEP 5: 업데이트날짜 = 오늘

  - Fingerprint 로테이션 (13개 UA, Referer 위장, 랜덤 딜레이)
  - 차단 감지 (403/캡차/빈응답) → 자동 프록시 전환
  - CSV/XLSX 입출력 (openpyxl)
"""
import re
import os
import csv
import json
import time
import random
import logging
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from io import StringIO
from urllib.parse import urlencode, quote

import requests
from bs4 import BeautifulSoup

logger = logging.getLogger(__name__)

# ═══════════════════════════════════════════
# Fingerprint 데이터베이스
# ═══════════════════════════════════════════
FINGERPRINTS = [
    {
        "ua": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
        "headers": {
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
            "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
            "Sec-Ch-Ua": '"Chromium";v="122", "Not(A:Brand";v="24", "Google Chrome";v="122"',
            "Sec-Ch-Ua-Mobile": "?0",
            "Sec-Ch-Ua-Platform": '"Windows"',
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "none",
            "Sec-Fetch-User": "?1",
            "Upgrade-Insecure-Requests": "1",
        },
    },
    {
        "ua": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
        "headers": {
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
            "Accept-Language": "ko-KR,ko;q=0.9,en;q=0.8",
            "Sec-Ch-Ua": '"Chromium";v="122", "Not(A:Brand";v="24", "Google Chrome";v="122"',
            "Sec-Ch-Ua-Mobile": "?0",
            "Sec-Ch-Ua-Platform": '"macOS"',
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "none",
            "Sec-Fetch-User": "?1",
            "Upgrade-Insecure-Requests": "1",
        },
    },
    {
        "ua": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
        "headers": {
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
            "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8",
            "Sec-Ch-Ua": '"Not A(Brand";v="99", "Google Chrome";v="121", "Chromium";v="121"',
            "Sec-Ch-Ua-Mobile": "?0",
            "Sec-Ch-Ua-Platform": '"Linux"',
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "none",
            "Sec-Fetch-User": "?1",
            "Upgrade-Insecure-Requests": "1",
        },
    },
    {
        "ua": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:123.0) Gecko/20100101 Firefox/123.0",
        "headers": {
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
            "Accept-Language": "ko-KR,ko;q=0.8,en-US;q=0.5,en;q=0.3",
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "none",
            "Sec-Fetch-User": "?1",
            "Upgrade-Insecure-Requests": "1",
            "DNT": "1",
        },
    },
    {
        "ua": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:123.0) Gecko/20100101 Firefox/123.0",
        "headers": {
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
            "Accept-Language": "ko,en-US;q=0.7,en;q=0.3",
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "none",
            "Sec-Fetch-User": "?1",
            "Upgrade-Insecure-Requests": "1",
        },
    },
    {
        "ua": "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:122.0) Gecko/20100101 Firefox/122.0",
        "headers": {
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
            "Accept-Language": "ko-KR,ko;q=0.8,en;q=0.5",
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "none",
            "Sec-Fetch-User": "?1",
            "Upgrade-Insecure-Requests": "1",
        },
    },
    {
        "ua": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.3 Safari/605.1.15",
        "headers": {
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "ko-KR,ko;q=0.9",
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "none",
        },
    },
    {
        "ua": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36 Edg/122.0.0.0",
        "headers": {
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
            "Accept-Language": "ko,en;q=0.9,en-US;q=0.8",
            "Sec-Ch-Ua": '"Chromium";v="122", "Not(A:Brand";v="24", "Microsoft Edge";v="122"',
            "Sec-Ch-Ua-Mobile": "?0",
            "Sec-Ch-Ua-Platform": '"Windows"',
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "none",
            "Sec-Fetch-User": "?1",
            "Upgrade-Insecure-Requests": "1",
        },
    },
    {
        "ua": "Mozilla/5.0 (Linux; Android 14; SM-S918B) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.6261.64 Mobile Safari/537.36",
        "headers": {
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
            "Accept-Language": "ko-KR,ko;q=0.9",
            "Sec-Ch-Ua": '"Chromium";v="122", "Not(A:Brand";v="24", "Google Chrome";v="122"',
            "Sec-Ch-Ua-Mobile": "?1",
            "Sec-Ch-Ua-Platform": '"Android"',
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "none",
            "Sec-Fetch-User": "?1",
            "Upgrade-Insecure-Requests": "1",
        },
    },
    {
        "ua": "Mozilla/5.0 (iPhone; CPU iPhone OS 17_3_1 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.3 Mobile/15E148 Safari/604.1",
        "headers": {
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "ko-KR,ko;q=0.9",
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "none",
        },
    },
    {
        "ua": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "headers": {
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
            "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
            "Sec-Ch-Ua": '"Not_A Brand";v="8", "Chromium";v="120", "Google Chrome";v="120"',
            "Sec-Ch-Ua-Mobile": "?0",
            "Sec-Ch-Ua-Platform": '"Windows"',
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "none",
            "Sec-Fetch-User": "?1",
            "Upgrade-Insecure-Requests": "1",
        },
    },
    {
        "ua": "Mozilla/5.0 (Linux; Android 13; SM-G991B) AppleWebKit/537.36 (KHTML, like Gecko) SamsungBrowser/23.0 Chrome/115.0.0.0 Mobile Safari/537.36",
        "headers": {
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
            "Accept-Language": "ko-KR,ko;q=0.9",
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "none",
            "Sec-Fetch-User": "?1",
            "Upgrade-Insecure-Requests": "1",
        },
    },
    {
        "ua": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
        "headers": {
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
            "Accept-Language": "ko,en-US;q=0.9,en;q=0.8",
            "Sec-Ch-Ua": '"Not A(Brand";v="99", "Google Chrome";v="121", "Chromium";v="121"',
            "Sec-Ch-Ua-Mobile": "?0",
            "Sec-Ch-Ua-Platform": '"macOS"',
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "none",
            "Sec-Fetch-User": "?1",
            "Upgrade-Insecure-Requests": "1",
        },
    },
]

# ═══════════════════════════════════════════
# 지역명 제거 패턴
# ═══════════════════════════════════════════
REGION_NAMES = sorted(
    [
        "부산서면역", "부산서면점", "부산서면", "부산해운대", "부산남포동", "부산광안리",
        "부산동래", "부산사하", "부산연제", "부산수영", "부산진구", "부산사상", "부산북구",
        "강남역", "강남점", "강남구", "강남", "서초구", "서초", "잠실", "홍대입구", "홍대",
        "신촌", "이태원", "명동", "종로", "강동", "강서", "송파", "마포", "영등포",
        "신림", "건대", "왕십리", "성수", "압구정", "청담", "삼성", "논현", "역삼",
        "선릉", "대치", "도곡", "방배", "사당", "신사", "합정", "상수", "연남",
        "성북", "동대문", "중구", "용산", "광화문", "여의도", "목동", "노원", "은평",
        "수원", "성남", "분당", "일산", "고양", "용인", "부천", "안양", "평택",
        "안산", "의정부", "파주", "김포", "광명", "하남", "남양주",
        "대구", "대전", "광주", "인천", "울산", "세종",
        "제주", "창원", "포항", "천안", "청주", "전주", "춘천", "원주", "강릉",
        "여수", "순천", "목포", "군산", "익산", "경주", "거제", "진주", "통영",
        "부산", "서면역", "서면점", "서면",
    ],
    key=len,
    reverse=True,
)

REGION_PATTERN = re.compile(r"\s*(" + "|".join(re.escape(r) for r in REGION_NAMES) + r")\s*")

# ═══════════════════════════════════════════
# 블랙리스트
# ═══════════════════════════════════════════
EMAIL_BLACKLIST = [
    "naver", "google", "kakao", "apple", "facebook", "instagram", "twitter", "youtube",
    "microsoft", "w3.org", "schema", "cloudflare", "jsdelivr", "unpkg", "webpack", "babel",
    "example", "noreply", "no-reply", "sentry", "wixpress", "placeholder", "fontawesome",
    "bootstrap", "jquery", "react", "angular", "localhost", "gabia", "cafe24", "hosting",
    "whois", "dothome", "iwinv", "vultr", "aws", "azure", "godo", "sixshop", "shopify",
    "wix.com", "squarespace", "godaddy", "bluehost", "namecheap", "imweb", "modoo",
    "adsense", "analytics", "mailchimp", "sendgrid", "hubspot", "intercom", "privacy",
    "webmaster", "postmaster", "abuse@", "root@", "imaeil", "alisonbrodmc",
]

HP_BLACKLIST = [
    "naver", "google", "gstatic", "pstatic", "facebook", "instagram", "youtube", "kakao",
    "twitter", "tiktok", "modoodoc", "goodoc", "hidoc", "medinavi", "pervsi", "k-info",
    "w3.org", "saramin", "jobkorea", "incruit", "daum", "wikipedia", "apple.com", "microsoft",
    "cloudflare", "jsdelivr", "unpkg", "schema.org", "goodhosrank", "cdn.", "ssl.",
    "polyfill", "fonts.", "jquery", "occupational", "line-culture",
    "cultureexpo", "invitanku", "recruit.kpb", "intojob",
    "tmailor", "himalayantouch", "ebs.co.kr", "psj.kr", "songpa.go", "cha.go",
    "mid.ebs", "kbs.co.kr", "mbc.co.kr", "sbs.co.kr",
]

# go.kr/or.kr 도메인은 조건부 차단 (검색 제목에 업체명 포함 시만 허용)
HP_GOV_DOMAINS = ["go.kr", "or.kr"]


class CrawlerEngine:
    """빈칸 채우기 크롤링 엔진"""

    def __init__(
        self,
        proxy_file=None,
        delay_min=0.5,
        delay_max=1.5,
        timeout=5,
        callback=None,
        api_keys=None,
        detail_workers=12,
        save_every=10,
    ):
        """
        Args:
            proxy_file: 프록시 IP:PORT 파일 경로 (없으면 직접 연결)
            delay_min: 최소 딜레이 (초)
            delay_max: 최대 딜레이 (초)
            timeout: 요청 타임아웃 (초)
            callback: 진행 상황 콜백 fn(event, data)
            api_keys: 네이버 API 키 리스트 [{"client_id": "...", "client_secret": "..."}, ...]
            detail_workers: PHASE 2 상세 크롤링 동시 워커 수 (기본 12, IP차단 없을 때 3배 병렬)
            save_every: xlsx/progress 저장 주기 — N건마다 1회 (기본 10)
        """
        self.delay_min = delay_min
        self.delay_max = delay_max
        self.timeout = timeout
        self.callback = callback or (lambda e, d: None)
        self.running = False
        self.detail_workers = max(1, int(detail_workers or 1))
        self.save_every = max(1, int(save_every or 1))
        self.stats = {
            "place_id": 0, "category": 0, "name": 0, "address": 0,
            "phone": 0, "visitor_review": 0, "blog_review": 0,
            "hp": 0, "email": 0, "naver_id": 0,
            "responsive": 0, "talktalk": 0, "kakao": 0, "instagram": 0,
            "no_phone_btn": 0, "webbuilder": 0, "new_biz": 0,
            "blocked": 0, "success": 0,
        }
        self._consecutive_blocks = 0
        # ═ 동시성 락 (PHASE 2 병렬 처리용) ═
        self._stats_lock = threading.Lock()
        self._save_lock = threading.Lock()
        self._cb_lock = threading.Lock()

        # 네이버 API 키 (로테이션) — 기본 내장 키
        _default_keys = [
            {"client_id": "yoBUbNSW9MGSPH36zaHN", "client_secret": "wt0HPPOVKA"},
        ]
        self.api_keys = api_keys if api_keys else _default_keys
        self._api_key_idx = 0

        # 프록시 로드
        self.proxies = []
        self.proxy_idx = 0
        self.fingerprint_map = {}
        if proxy_file and os.path.isfile(proxy_file):
            with open(proxy_file, "r") as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith("#"):
                        parts = line.split(":")
                        ip = parts[0].strip()
                        port = int(parts[1]) if len(parts) > 1 else 8080
                        user = parts[2].strip() if len(parts) > 2 else None
                        pw = parts[3].strip() if len(parts) > 3 else None
                        proxy = {"ip": ip, "port": port, "user": user, "pw": pw, "blocked": False, "block_until": 0}
                        self._assign_fingerprint(proxy)
                        self.proxies.append(proxy)
            self.callback("log", f"프록시 {len(self.proxies)}개 로드됨")

        # 직접연결용 fingerprint
        self._direct_fp_idx = random.randint(0, len(FINGERPRINTS) - 1)

        # requests 세션 (쿠키 없음, 연결 풀 강화)
        self.session = requests.Session()
        self.session.cookies.clear()
        # 연결 풀 크기 확대 → 병렬 워커 + 재연결 오버헤드 제거
        _pool_size = max(20, self.detail_workers * 4)
        adapter = requests.adapters.HTTPAdapter(
            pool_connections=_pool_size, pool_maxsize=_pool_size, max_retries=1
        )
        self.session.mount("http://", adapter)
        self.session.mount("https://", adapter)

    def _assign_fingerprint(self, proxy):
        key = f"{proxy['ip']}:{proxy['port']}"
        if key not in self.fingerprint_map:
            self.fingerprint_map[key] = random.randint(0, len(FINGERPRINTS) - 1)
        proxy["fp_idx"] = self.fingerprint_map[key]

    def _reassign_fingerprint(self, proxy):
        key = f"{proxy['ip']}:{proxy['port']}"
        old_idx = self.fingerprint_map.get(key, 0)
        new_idx = old_idx
        while new_idx == old_idx and len(FINGERPRINTS) > 1:
            new_idx = random.randint(0, len(FINGERPRINTS) - 1)
        self.fingerprint_map[key] = new_idx
        proxy["fp_idx"] = new_idx

    def _get_fingerprint(self, proxy=None):
        if proxy:
            idx = proxy.get("fp_idx", 0) % len(FINGERPRINTS)
        else:
            idx = self._direct_fp_idx
        return FINGERPRINTS[idx]

    def _get_headers(self, proxy=None, referer=None):
        fp = self._get_fingerprint(proxy)
        headers = {"User-Agent": fp["ua"], **fp["headers"]}
        if referer:
            headers["Referer"] = referer
            headers["Sec-Fetch-Site"] = "cross-site"
        # 쿠키 없음
        headers.pop("Cookie", None)
        return headers

    def _get_proxy_dict(self, proxy):
        if not proxy:
            return None
        if proxy.get("user") and proxy.get("pw"):
            url = f"http://{proxy['user']}:{proxy['pw']}@{proxy['ip']}:{proxy['port']}"
        else:
            url = f"http://{proxy['ip']}:{proxy['port']}"
        return {"http": url, "https": url}

    def _get_next_proxy(self):
        if not self.proxies:
            return None
        now = time.time()
        for i in range(len(self.proxies)):
            idx = (self.proxy_idx + i) % len(self.proxies)
            p = self.proxies[idx]
            if p["blocked"] and now > p["block_until"]:
                p["blocked"] = False
            if not p["blocked"]:
                self.proxy_idx = (idx + 1) % len(self.proxies)
                return p
        # 전부 차단 → 가장 오래된 복구
        oldest = min(self.proxies, key=lambda p: p["block_until"])
        oldest["blocked"] = False
        return oldest

    def _mark_blocked(self, proxy):
        proxy["blocked"] = True
        proxy["block_until"] = time.time() + 300  # 5분 쿨다운
        self._reassign_fingerprint(proxy)
        self.stats["blocked"] += 1
        self._consecutive_blocks += 1

    def _is_blocked(self, status_code, text):
        if status_code in (403, 429):
            return True
        if not text or len(text) < 500:
            return True
        # 실제 차단 페이지 감지 (captchaApi 등 설정 URL은 정상이므로 제외)
        block_signs = ["서비스 이용이 제한", "자동입력 방지", "이 페이지를 찾을 수 없습니다"]
        if any(s in text for s in block_signs):
            return True
        # captcha 체크는 실제 캡차 입력 폼이 있는 경우만 (API URL은 제외)
        if ("captcha" in text.lower() and "captcha-input" in text.lower()):
            return True
        return False

    def _random_delay(self, min_s=None, max_s=None):
        backoff = min(1.5 ** self._consecutive_blocks, 10)
        lo = (min_s if min_s is not None else self.delay_min) * backoff
        hi = (max_s if max_s is not None else self.delay_max) * backoff
        time.sleep(random.uniform(lo, hi))

    def _fetch(self, url, referer=None, timeout=None):
        """HTTP GET with fingerprint, proxy rotation, retry"""
        timeout = timeout or self.timeout
        for attempt in range(3):
            if not self.running:
                return ""
            try:
                proxy = self._get_next_proxy()
                headers = self._get_headers(proxy, referer)
                proxy_dict = self._get_proxy_dict(proxy)

                # 새 세션마다 쿠키 초기화
                self.session.cookies.clear()

                resp = self.session.get(
                    url, headers=headers, proxies=proxy_dict, timeout=timeout, allow_redirects=True
                )

                if self._is_blocked(resp.status_code, resp.text):
                    if proxy:
                        self._mark_blocked(proxy)
                        self.callback("log", f"차단 감지 → IP 로테이션 (시도 {attempt+1}/3)")
                        continue
                    else:
                        self.stats["blocked"] += 1
                        self._consecutive_blocks += 1
                        time.sleep(3 * (attempt + 1))
                        continue

                self.stats["success"] += 1
                self._consecutive_blocks = 0
                resp.encoding = "utf-8"
                return resp.text

            except Exception as e:
                logger.debug(f"fetch error: {e}")
                if attempt < 2:
                    time.sleep(1 * (attempt + 1))
        return ""

    # ═══════════════════════════════════════════
    # 네이버 API
    # ═══════════════════════════════════════════

    def _search_naver_api(self, query):
        """네이버 지역검색 API로 업체 정보 조회. 결과 dict 리스트 반환."""
        if not self.api_keys:
            return []
        key = self.api_keys[self._api_key_idx % len(self.api_keys)]
        self._api_key_idx += 1
        try:
            resp = self.session.get(
                "https://openapi.naver.com/v1/search/local.json",
                params={"query": query, "display": 5},
                headers={
                    "X-Naver-Client-Id": key["client_id"],
                    "X-Naver-Client-Secret": key["client_secret"],
                },
                timeout=self.timeout,
            )
            if resp.status_code == 200:
                data = resp.json()
                return data.get("items", [])
            else:
                logger.debug(f"Naver API {resp.status_code}: {resp.text[:200]}")
        except Exception as e:
            logger.debug(f"Naver API error: {e}")
        return []

    # ═══════════════════════════════════════════
    # 추출 함수들
    # ═══════════════════════════════════════════

    @staticmethod
    def extract_emails(html):
        """HTML에서 이메일 추출 (블랙리스트 필터링)"""
        raw = set(re.findall(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", html))
        result = []
        for e in raw:
            lo = e.lower()
            if re.search(r"\.(png|jpg|jpeg|gif|svg|webp|ico|css|js|woff|ttf|eot|mp4|pdf|map|min)$", lo):
                continue
            if any(w in lo for w in EMAIL_BLACKLIST):
                continue
            local, domain = lo.split("@", 1) if "@" in lo else ("", "")
            if not local or not domain or len(local) < 2 or len(domain) < 4:
                continue
            if re.match(r"^\d+$", local):
                continue
            if not re.search(r"\.(com|co\.kr|kr|net|org|gmail|nate|hanmail|daum)$", domain):
                continue
            result.append(e)
        return result

    @staticmethod
    def extract_homepage(html):
        """HTML에서 홈페이지 URL 추출 (블랙리스트 필터링)"""
        urls = re.findall(
            r"https?://(?:www\.)?[a-zA-Z0-9][a-zA-Z0-9.-]+\.(?:co\.kr|com|kr|net|clinic|hospital)[^\s\"'<>&;)}\]]{0,80}",
            html,
            re.IGNORECASE,
        )
        domain_map = {}
        for u in urls:
            lo = u.lower()
            if any(b in lo for b in HP_BLACKLIST) or len(u) > 120:
                continue
            try:
                from urllib.parse import urlparse
                parsed = urlparse(u)
                domain = parsed.hostname.replace("www.", "") if parsed.hostname else ""
                if domain and domain not in domain_map:
                    domain_map[domain] = f"{parsed.scheme}://{parsed.hostname}/"
            except Exception:
                pass
        return list(domain_map.values())

    @staticmethod
    def clean_business_name(name):
        """업체명에서 지역명 제거"""
        return REGION_PATTERN.sub(" ", name).strip()

    # ═══════════════════════════════════════════
    # 신규 보조 헬퍼 (반응형 / 톡톡 / 신규업체 / 리뷰합계)
    # ═══════════════════════════════════════════

    @staticmethod
    def detect_responsive(html):
        """홈페이지 HTML이 반응형인지 판정.
        - viewport meta + (media query OR 반응형 프레임워크) → "O"
        - 그 외 → "X"
        반환: "O" / "X"
        """
        if not html or len(html) < 50:
            return "X"
        lo = html.lower()
        # 1) viewport meta 필수 (width=device-width)
        has_viewport = bool(re.search(
            r'<meta[^>]+name=["\']viewport["\'][^>]*content=["\'][^"\']*width\s*=\s*device-width',
            lo,
        ))
        if not has_viewport:
            return "X"
        # 2) 보조 단서 — 미디어쿼리 / 반응형 프레임워크
        clues = (
            "@media",
            "max-width",
            "min-width",
            "bootstrap",
            "tailwind",
            "responsive",
            "col-md-",
            "col-sm-",
            "col-lg-",
            "flex-",
            "grid-",
        )
        if any(c in lo for c in clues):
            return "O"
        # viewport만 있고 단서 없으면 그래도 O로 간주 (모바일 대응 의도가 있음)
        return "O"

    # 네이버 톡톡 시스템/공용 채널 블랙리스트 (업체 톡톡 아님)
    # - w9ni795: "네이버 공유누리 예약 AI" — 예약 모듈 placeholder, 모든 페이지에 박힘
    TALKTALK_BLACKLIST = frozenset({
        "w9ni795",
    })

    @staticmethod
    def _normalize_talktalk_id(raw_id):
        """톡톡 채널 ID 정규화 + 블랙리스트 검사.
        반환: 통과 시 정규화된 ID(소문자), 차단 시 ""
        """
        if not raw_id:
            return ""
        cid = raw_id.lower().strip("/")
        # 길이 검증 (4자 미만 = 잡음)
        if len(cid) < 4 or len(cid) > 30:
            return ""
        # 블랙리스트
        if cid in CrawlerEngine.TALKTALK_BLACKLIST:
            return ""
        return cid

    # 디버그용: 마지막 톡톡 추출 시도 정보 (빈 결과 원인 추적)
    _last_talktalk_debug = None

    @staticmethod
    def extract_talktalk_url(html):
        """HTML/JSON 텍스트에서 네이버 톡톡 URL 추출 (관대한 매칭 + 블랙리스트).

        매칭 대상:
        A) URL 형식
           - https://talk.naver.com/ct/<id>     (비즈니스 톡톡 표준)
           - https://talk.naver.com/W<id>       (대문자 W 단축)
           - https://talk.naver.com/<id>        (소문자/영숫자 단축)
        B) JSON 필드 (네이버 GraphQL/플레이스 API 응답에 박힌 ID)
           - "talktalkId":"<id>"
           - "talkChannelId":"<id>"
           - "chatChannelId":"<id>"
           - "talkId":"<id>"
           - "naverTalkId":"<id>"

        검증: ID 4~30자 + 블랙리스트 (w9ni795 등 시스템 placeholder 차단)
        모든 출력은 ct/소문자 형식.
        """
        if not html:
            return ""
        # JSON 인코딩된 슬래시(\u002F) 디코딩
        text = html.replace("\\u002F", "/").replace("\\/", "/")

        debug = {"input_len": len(text), "talk_substring": "talk.naver.com" in text or "talktalk" in text.lower(),
                 "pattern_a_hits": 0, "pattern_b_hits": 0, "rejected_ids": []}

        # === Pattern A: talk.naver.com URL (관대) ===
        for m in re.finditer(r'https?://talk\.naver\.com/([A-Za-z0-9_/\-]+)', text):
            debug["pattern_a_hits"] += 1
            raw_path = m.group(1).rstrip('"\'<>),;.?#&')
            if raw_path.lower().startswith('ct/'):
                cid_raw = raw_path[3:].split('/')[0]
            else:
                cid_raw = raw_path.split('/')[0]
            cid = CrawlerEngine._normalize_talktalk_id(cid_raw)
            if cid:
                CrawlerEngine._last_talktalk_debug = debug
                return f"https://talk.naver.com/ct/{cid}"
            else:
                debug["rejected_ids"].append(cid_raw)

        # === Pattern B: JSON 필드 안의 ID ===
        json_field_patterns = [
            r'"talktalkId"\s*:\s*"([A-Za-z0-9_\-]{4,30})"',
            r'"talkChannelId"\s*:\s*"([A-Za-z0-9_\-]{4,30})"',
            r'"chatChannelId"\s*:\s*"([A-Za-z0-9_\-]{4,30})"',
            r'"talkId"\s*:\s*"([A-Za-z0-9_\-]{4,30})"',
            r'"naverTalkId"\s*:\s*"([A-Za-z0-9_\-]{4,30})"',
            r'"channelId"\s*:\s*"([A-Za-z0-9_\-]{4,30})"',  # 네이버 booking 응답
        ]
        for p in json_field_patterns:
            for m in re.finditer(p, text):
                debug["pattern_b_hits"] += 1
                cid = CrawlerEngine._normalize_talktalk_id(m.group(1))
                if cid:
                    CrawlerEngine._last_talktalk_debug = debug
                    return f"https://talk.naver.com/ct/{cid}"
                else:
                    debug["rejected_ids"].append(m.group(1))

        CrawlerEngine._last_talktalk_debug = debug
        return ""

    @staticmethod
    def normalize_talktalk_url(url):
        """이미 갖고 있는 URL을 검증/정규화 (GraphQL repr/etc 출처용).
        통과 시 ct/ 형식 URL 반환, 실패 시 ""
        """
        if not url:
            return ""
        # 도메인 빠른 체크
        if "talk.naver.com" not in url:
            return ""
        return CrawlerEngine.extract_talktalk_url(url)

    @staticmethod
    def extract_kakao_url(html):
        """HTML 텍스트에서 카카오톡 채널/플러스친구 URL 추출.
        지원 패턴:
        - https://pf.kakao.com/_xxxxx
        - https://plus.kakao.com/...
        - https://open.kakao.com/...
        - https://kakaochannel.com/...
        반환: URL 문자열 또는 ""
        """
        if not html:
            return ""
        text = html.replace("\\u002F", "/").replace("\\/", "/")
        patterns = [
            r'https?://pf\.kakao\.com/[A-Za-z0-9_\-/]+',
            r'https?://plus\.kakao\.com/[A-Za-z0-9_\-/]+',
            r'https?://open\.kakao\.com/[A-Za-z0-9_\-/]+',
            r'https?://kakaochannel\.com/[A-Za-z0-9_\-/]+',
        ]
        for p in patterns:
            m = re.search(p, text)
            if m:
                url = m.group(0).rstrip('"\'<>)?,;')
                # 최소 길이 검증
                if len(url) > 22:
                    return url
        return ""

    @staticmethod
    def extract_instagram_url(html):
        """HTML 텍스트에서 인스타그램 프로필 URL 추출.
        지원 패턴:
        - https://www.instagram.com/<username>
        - https://instagram.com/<username>
        주의: /p/<postid>, /reel/<id>, /explore 등 비프로필 경로는 제외.
        반환: URL 문자열 또는 ""
        """
        if not html:
            return ""
        text = html.replace("\\u002F", "/").replace("\\/", "/")
        # username만 매칭 (소문자/숫자/언더스코어/마침표, 30자 이하)
        # lookahead로 끝 경계 검사 — 슬래시/따옴표/태그/쿼리/공백/콤마/세미콜론/문자열끝
        m = re.search(
            r'https?://(?:www\.)?instagram\.com/([A-Za-z0-9_.]{1,30})(?=$|[/"\'<>\s\?,;)\]])',
            text,
        )
        if not m:
            return ""
        username = m.group(1)
        # 비프로필 경로/예약어 제외
        if username.lower() in (
            "p", "reel", "reels", "tv", "explore", "stories", "accounts",
            "developer", "about", "directory", "legal", "privacy", "terms",
            "web", "fragment", "challenge",
        ):
            return ""
        return f"https://www.instagram.com/{username}"

    @staticmethod
    def detect_phone_button(html):
        """홈페이지 HTML에 모바일 바로 전화 버튼이 있는지 검사.
        지원 패턴:
        - <a href="tel:01012345678">
        - href='tel:0212345678'
        - onclick="location.href='tel:...'"
        - data-href="tel:..."
        반환: True (전화버튼 있음) / False (없음)
        """
        if not html:
            return False
        # tel: 링크 (가장 신뢰할 수 있는 단서)
        if re.search(r'''(?:href|data-href|src)\s*=\s*["']\s*tel:\s*[\d+\-\s()]{6,}''', html, re.IGNORECASE):
            return True
        if re.search(r'''location\.href\s*=\s*["']\s*tel:''', html, re.IGNORECASE):
            return True
        if re.search(r'''window\.open\s*\(\s*["']\s*tel:''', html, re.IGNORECASE):
            return True
        return False

    # SaaS 웹빌더 시그니처 (도메인/HTML marker 양쪽 검사)
    # 영업 정책: O = 웹빌더 사용 (코드 수정 불가), X = 일반 HTML (수정 가능 → 영업 타겟)
    _WEBBUILDER_SIGNATURES = (
        ("modoo",        ("modoo.at", "MODOO_BUILD", "modoo-resource")),
        ("imweb",        ("imweb.me", "imweb.kr", "static.imweb", "imweb-")),
        ("wix",          ("wix.com", "wixstatic", "wixsite", "wix-bolt", "_wixCIDX")),
        ("cafe24",       ("cafe24.com", "ec.cafe24.com", "cafe24cdn")),
        ("sixshop",      ("sixshop.com", "sixshopservice")),
        ("makeshop",     ("makeshop.co.kr", "makeshopcafe")),
        ("godomall",     ("godomall.com", "godo.co.kr", "godomall-")),
        ("squarespace",  ("squarespace.com", "static1.squarespace")),
        ("weebly",       ("weebly.com", "weeblycloud")),
        ("shopify",      ("myshopify.com", "cdn.shopify.com")),
        ("webflow",      ("webflow.com", "webflow.io")),
        ("jimdo",        ("jimdo.com", "jimdofree.com")),
        ("wordpress.com",("wordpress.com", "wp.com")),  # 호스팅판만 (셀프호스트 wp-content 제외)
        ("google-sites", ("sites.google.com",)),
        ("naver-blog",   ("blog.naver.com", "blog.me")),
        ("tistory",      ("tistory.com",)),
        ("homepy",       ("homepy.com",)),
        ("1px",          ("1px.kr",)),
    )

    @staticmethod
    def detect_webbuilder(html, url=None):
        """홈페이지가 SaaS 웹빌더로 만들어졌는지 검사.
        도메인(URL hostname) + HTML 내 marker 두 방향으로 검사.
        반환: "O" (웹빌더 감지 — 코드 수정 불가) / "X" (일반 HTML — 영업 타겟)
        """
        # 1) URL hostname 검사 (가장 강한 단서)
        if url:
            try:
                from urllib.parse import urlparse
                host = (urlparse(url).hostname or "").lower()
            except Exception:
                host = ""
            if host:
                for _name, markers in CrawlerEngine._WEBBUILDER_SIGNATURES:
                    for m in markers:
                        # 도메인 매칭: hostname이 marker로 끝나거나 marker 자체
                        if host == m or host.endswith("." + m) or m in host:
                            return "O"
        # 2) HTML 본문 marker 검사
        if html:
            text = html if isinstance(html, str) else str(html)
            for _name, markers in CrawlerEngine._WEBBUILDER_SIGNATURES:
                for m in markers:
                    if m in text:
                        return "O"
        return "X"

    @staticmethod
    def detect_new_business(html):
        """place HTML 또는 GraphQL 응답 텍스트에서 '신규개업' 여부 감지.
        다음 단서 중 하나라도 있으면 "O":
        - "isNewlyOpened": true / "newlyOpenBadge"
        - "newBusinessHours" 또는 "newBusiness" 키
        - apollo state 안의 "NEW_BUSINESS"
        - 한글 배지 텍스트 "신규" / "NEW" (place_blind 인접)
        반환: "O" / "X"
        """
        if not html:
            return "X"
        patterns = [
            r'"isNewlyOpened"\s*:\s*true',
            r'"isNew"\s*:\s*true',
            r'"newlyOpenBadge"',
            r'"newBusinessHours"',
            r'"newBusiness"\s*:\s*true',
            r'"NEW_BUSINESS"',
            r'"newOpen"\s*:\s*true',
            r'"isNewOpen"\s*:\s*true',
        ]
        for p in patterns:
            if re.search(p, html):
                return "O"
        # place HTML 한글/배지 휴리스틱 (오탐 줄이려고 강한 단어만)
        if re.search(r'place_blind[^<]{0,40}>[^<]*신규개업', html):
            return "O"
        if re.search(r'>NEW<\s*/[a-z]+>\s*신규', html):
            return "O"
        return "X"

    @staticmethod
    def _compute_review_total(r):
        """방문자리뷰수 + 블로그리뷰수 → 정수 합계.
        문자열/공백/None 모두 0으로 처리.
        """
        def _to_int(v):
            try:
                s = str(v or "").strip().replace(",", "")
                return int(s) if s else 0
            except (ValueError, TypeError):
                return 0
        return _to_int(r.get("방문자리뷰수")) + _to_int(r.get("블로그리뷰수"))

    # ═══════════════════════════════════════════
    # STEP 0: 고유번호(placeId) 추출
    # ═══════════════════════════════════════════

    def _find_place_id(self, name):
        """네이버 검색으로 placeId 추출. 반환: str 또는 ""."""
        clean = self.clean_business_name(name)

        # API 모드 우선
        if self.api_keys:
            items = self._search_naver_api(clean)
            for item in items:
                link = item.get("link", "")
                m = re.search(r"/place/(\d+)", link)
                if m:
                    return m.group(1)

        # HTML 폴백: 네이버 검색 결과에서 placeId 추출
        url = f"https://search.naver.com/search.naver?query={quote(clean)}"
        html = self._fetch(url, referer="https://www.naver.com/")
        if html:
            # /place/12345 패턴 매칭
            m = re.search(r"/place/(\d+)", html)
            if m:
                return m.group(1)
        return ""

    # ═══════════════════════════════════════════
    # STEP 1: 플레이스 HTML에서 여러 항목 동시 추출
    # ═══════════════════════════════════════════

    def _fetch_place_html(self, pid, name=""):
        """네이버 플레이스 HTML 가져오기 — localhost:3100 프록시 경유.
        place HTML은 CSR이라 일반 fetch로는 빈 껍데기만 옴.
        localhost:3100 프록시가 SSR 데이터를 포함한 전체 HTML을 반환.
        hospital URL 사용 (place는 리다이렉트만 반환).
        """
        import json as _json
        url = f"https://m.place.naver.com/hospital/{pid}/home"
        fp = self._get_fingerprint()
        try:
            payload = _json.dumps({
                "targetUrl": url,
                "method": "GET",
                "headers": {"User-Agent": fp["ua"]}
            })
            resp = self.session.post(
                "http://localhost:3100",
                data=payload,
                headers={
                    "Content-Type": "application/json",
                    "x-api-key": "onda-proxy-2026-secret",
                },
                timeout=self.timeout + 5,
            )
            if resp.status_code == 200 and len(resp.content) > 10000:
                self.stats["success"] += 1
                # UTF-8 강제 디코딩 (프록시가 인코딩을 잘못 보고할 수 있음)
                resp.encoding = "utf-8"
                return resp.text
        except Exception as e:
            logger.debug(f"place proxy fetch error: {e}")

        # 폴백: 직접 fetch (한국 IP면 작동할 수 있음)
        referer = f"https://m.search.naver.com/search.naver?query={quote(name)}"
        html = self._fetch(url, referer=referer)
        if html and len(html) > 10000:
            return html

        # 3차 폴백: 모바일 검색 결과에서 해당 업체 JSON 블록 추출
        mobile_data = self._fetch_place_from_mobile_search(pid, name)
        if mobile_data:
            return mobile_data  # dict 형태로 반환 (place HTML 대신)
        return ""

    def _fetch_place_from_mobile_search(self, pid, name=""):
        """모바일 네이버 검색 결과에서 업체 상세 정보 추출.
        place HTML 프록시 없는 Windows 환경용 폴백.
        반환: place HTML과 호환되는 pseudo-HTML 문자열 또는 ""
        """
        if not name:
            return ""
        clean = self.clean_business_name(name)
        url = f"https://m.search.naver.com/search.naver?where=m&query={quote(clean)}"
        # 모바일 UA로 직접 요청 (일반 _fetch의 UA가 데스크톱이라 결과가 다를 수 있음)
        try:
            self.session.cookies.clear()
            resp = self.session.get(url, headers={
                "User-Agent": "Mozilla/5.0 (Linux; Android 13; SM-G991N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Mobile Safari/537.36",
                "Referer": "https://m.naver.com/",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language": "ko-KR,ko;q=0.9",
            }, timeout=self.timeout)
            if resp.status_code == 200:
                resp.encoding = "utf-8"
                html = resp.text
            else:
                html = ""
        except Exception:
            html = ""
        if not html or len(html) < 1000:
            return ""

        # PlaceSummary:{pid} JSON 블록에서 데이터 추출
        pid_str = str(pid)
        marker = f"PlaceSummary:{pid_str}"
        idx = html.find(marker)
        if idx < 0:
            # pid 자체로 검색
            idx = html.find(pid_str)
        if idx < 0:
            return ""

        # pid 주변 3000자 범위에서 데이터 추출
        chunk = html[idx:min(len(html), idx + 3000)]

        result = {}
        # category
        m = re.search(r'"category"\s*:\s*"([^"]+)"', chunk)
        if m:
            result["category"] = m.group(1)
        # fullAddress 우선, roadAddress 폴백
        m = re.search(r'"fullAddress"\s*:\s*"([^"]+)"', chunk)
        if m:
            result["address"] = m.group(1)
        else:
            m = re.search(r'"roadAddress"\s*:\s*"([^"]+)"', chunk)
            if m:
                result["address"] = m.group(1)
        # virtualPhone
        m = re.search(r'"virtualPhone"\s*:\s*"([0-9-]+)"', chunk)
        if m:
            result["phone"] = m.group(1)
        # visitorReviewCount (모바일 검색에서는 축약값)
        m = re.search(r'"visitorReviewCount"\s*:\s*"?(\d+)', chunk)
        if m:
            result["visitor_review"] = m.group(1)
        # blogCafeReviewCount
        m = re.search(r'"blogCafeReviewCount"\s*:\s*"?(\d+)', chunk)
        if m:
            result["blog_review"] = m.group(1)
        # normalizedName 우선 (mark 태그 없음), name 폴백
        m = re.search(r'"normalizedName"\s*:\s*"([^"]+)"', chunk)
        if m:
            result["name"] = m.group(1)
        else:
            m = re.search(r'"name"\s*:\s*"([^"]+)"', chunk)
            if m:
                name_val = m.group(1)
                # <mark> 태그 제거
                name_val = re.sub(r'\\u003C/?mark\\u003E|\\u003C\\u002Fmark\\u003E', '', name_val)
                result["name"] = name_val

        if not result:
            return ""

        # _parse_place_html과 호환되는 pseudo-HTML 문자열 생성
        pseudo = " ".join(f'"{k}":"{v}"' for k, v in result.items())
        return pseudo

    def _parse_place_html(self, html):
        """
        플레이스 HTML에서 여러 항목 추출.
        반환: dict with keys: category, name, address, phone, visitor_review, blog_review
        """
        result = {}

        if not html or len(html) < 3000:
            return result

        # 업종(키워드)
        m = re.search(r'"category"\s*:\s*"([^"]+)"', html)
        if m:
            result["category"] = m.group(1)

        # 상호명
        m = re.search(r'"name"\s*:\s*"([^"]+)"', html)
        if m:
            result["name"] = m.group(1)

        # 업체주소 (roadAddress 우선)
        m = re.search(r'"roadAddress"\s*:\s*"([^"]+)"', html)
        if not m:
            m = re.search(r'"address"\s*:\s*"([^"]+)"', html)
        if m:
            result["address"] = m.group(1)

        # 안심번호 (virtualPhone 우선)
        vp = re.search(r'"virtualPhone"\s*:\s*"([^"]+)"', html)
        ph = re.search(r'"phone"\s*:\s*"([^"]+)"', html)
        p = (vp.group(1) if vp else "") or (ph.group(1) if ph else "")
        if p and re.search(r"\d", p):
            result["phone"] = p

        # 방문자리뷰수 (여러 필드명 대응)
        for pat in [r'"visitorReviewCount"\s*:\s*(\d+)', r'"visitorReviewsTotal"\s*:\s*(\d+)']:
            m = re.search(pat, html)
            if m and int(m.group(1)) > 0:
                result["visitor_review"] = m.group(1)
                break

        # 블로그리뷰수 (여러 필드명 대응)
        for pat in [
            r'FsasReviewsResult","total":(\d+)',           # __APOLLO_STATE__ 내 블로그/카페 리뷰
            r'"blogCafeReviewCount"\s*:\s*(\d+)',           # GraphQL 응답
            r'"blogCafeReviewsTotal"\s*:\s*(\d+)',          # 변형
        ]:
            m = re.search(pat, html)
            if m and int(m.group(1)) > 0:
                result["blog_review"] = m.group(1)
                break

        # 네이버 블로그 ID 추출 (3단계 폴백)
        blog_id = ""
        # 1단계: BaseNaverBlog
        m = re.search(r'BaseNaverBlog:([a-zA-Z0-9_.-]+)', html)
        if m:
            blog_id = m.group(1)
        # 2단계: blogId
        if not blog_id:
            m = re.search(r'"blogId"\s*:\s*"([a-zA-Z0-9_.-]+)"', html)
            if m:
                blog_id = m.group(1)
        # 3단계: blog.naver.com URL
        if not blog_id:
            m = re.search(r'blog\.naver\.com/([a-zA-Z0-9_.-]+)', html)
            if m and m.group(1) not in ("PostView", "PostList", "BlogTagLog", "NBlogTop", "profile", "intro"):
                blog_id = m.group(1)
        if blog_id:
            result["naver_id"] = blog_id

        # 이메일 (place HTML에서 추출, 추가 요청 0)
        email_m = re.search(r'"email"\s*:\s*"([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})"', html)
        if email_m:
            result["email"] = email_m.group(1)

        # 홈페이지 URL (place HTML에서 추출)
        # 방법1: place_blind span 뒤의 첫 <a href> (가장 안정적)
        m = re.search(r'place_blind[^<]*</span></strong>[^<]*<[^>]*<[^>]*<a\s+href="(https?://[^"]+)"', html)
        if not m:
            # 방법2: homepages 블록에서 마지막 HomepageRepr의 url (홈페이지는 보통 마지막)
            homepage_urls = re.findall(r'"HomepageRepr","url":"([^"]+)"', html)
            if homepage_urls:
                # 마지막 URL이 보통 홈페이지 (예약/인스타/유튜브/블로그 뒤)
                last_url = homepage_urls[-1].replace("\\u002F", "/")
                if not any(b in last_url.lower() for b in HP_BLACKLIST):
                    result["homepage"] = last_url
        else:
            hp_url = m.group(1)
            if hp_url and not any(b in hp_url.lower() for b in HP_BLACKLIST):
                result["homepage"] = hp_url

        # 네이버 톡톡 URL (place HTML 안의 talk.naver.com 링크 또는 JSON 필드)
        talktalk = self.extract_talktalk_url(html)
        if talktalk:
            result["talktalk"] = talktalk
        else:
            # 디버그: 입력에 talk 단서가 있었는데 추출 실패한 경우 GUI 콘솔에 노출
            dbg = CrawlerEngine._last_talktalk_debug
            if dbg and dbg.get("talk_substring") and (dbg.get("pattern_a_hits", 0) + dbg.get("pattern_b_hits", 0)) > 0:
                try:
                    rejected = ', '.join(dbg.get("rejected_ids", [])[:3])
                    self.callback("log", f"[talk-debug] place HTML에 talk 단서 있으나 추출 실패. A={dbg['pattern_a_hits']} B={dbg['pattern_b_hits']} 거절ID=[{rejected}]")
                except Exception:
                    pass

        # 카카오톡 URL
        kakao = self.extract_kakao_url(html)
        if kakao:
            result["kakao"] = kakao

        # 인스타그램 URL
        insta = self.extract_instagram_url(html)
        if insta:
            result["instagram"] = insta

        # 신규개업 여부
        if self.detect_new_business(html) == "O":
            result["is_new"] = "O"

        return result

    # ═══════════════════════════════════════════
    # 홈페이지 검증
    # ═══════════════════════════════════════════

    def _extract_homepage_with_verify(self, search_html, name):
        """검색 결과 HTML에서 홈페이지 URL 추출 + 경량 검증.
        go.kr/or.kr 및 확장 블랙리스트 차단. 추가 HTTP 요청 없이 URL만으로 판단.
        반환: URL 문자열 또는 ""
        """
        urls = self.extract_homepage(search_html)
        for url in urls:
            lo = url.lower()
            # go.kr/or.kr 정부/비영리 스킵
            if any(d in lo for d in HP_GOV_DOMAINS):
                continue
            # 확장 블랙리스트 (관공서/구인/무관 사이트)
            if any(b in lo for b in ["cultureexpo", "invitanku", "recruit.kpb", 
                                      "intojob", "songpa.go", "cha.go"]):
                continue
            return url
        # 폴백: 블랙리스트 없이 첫 번째 URL (go.kr/or.kr만 제외)
        for url in urls:
            lo = url.lower()
            if not any(d in lo for d in HP_GOV_DOMAINS):
                return url
        return ""

    # ═══════════════════════════════════════════
    # 네이버아이디 4단계 폴백
    # ═══════════════════════════════════════════

    def _fetch_place_graphql(self, pid):
        """네이버 플레이스 GraphQL API로 상세 정보 가져오기.
        반환: dict with phone, homepage, visitor_review, blog_review, category, address 등
        """
        import json as _json
        url = "https://pcmap-api.place.naver.com/graphql"
        fp = self._get_fingerprint()
        headers = {
            "User-Agent": fp["ua"],
            "Referer": "https://map.naver.com/",
            "Content-Type": "application/json",
        }
        payload = [{
            "operationName": "getPlaceDetail",
            "variables": {"input": {"deviceType": "pc", "id": str(pid), "isNx": False}},
            "query": """query getPlaceDetail($input: PlaceDetailInput!) {
              placeDetail(input: $input) {
                base {
                  name phone virtualPhone category
                  address roadAddress
                  visitorReviewsTotal
                }
                homepages {
                  repr { url type }
                  etc { url type }
                }
                fsasReviews { total }
              }
            }"""
        }]
        result = {}
        try:
            resp = self.session.post(url, json=payload, headers=headers, timeout=self.timeout)
            if resp.status_code == 200:
                raw_text = resp.text or ""
                data = resp.json()
                if data and isinstance(data, list) and data[0].get("data"):
                    detail = data[0]["data"].get("placeDetail", {})
                    base = detail.get("base", {})
                    hp = detail.get("homepages", {})
                    fsas = detail.get("fsasReviews", {})

                    # 전화번호
                    result["phone"] = base.get("virtualPhone") or base.get("phone") or ""
                    result["real_phone"] = base.get("phone") or ""

                    # 홈페이지 (repr에서)
                    repr_hp = hp.get("repr")
                    if repr_hp and repr_hp.get("url"):
                        result["homepage"] = repr_hp["url"]
                        result["homepage_type"] = repr_hp.get("type", "")
                    # etc 링크들 (인스타 등)
                    etc_links = hp.get("etc", [])
                    if etc_links:
                        result["etc_links"] = [{"url": l.get("url", ""), "type": l.get("type", "")} for l in etc_links]

                    # 리뷰
                    result["visitor_review"] = str(base.get("visitorReviewsTotal", 0) or 0)
                    result["blog_review"] = str(fsas.get("total", 0) if fsas else 0)

                    # 카테고리/주소
                    result["category"] = base.get("category", "")
                    result["address"] = base.get("roadAddress") or base.get("address") or ""

                    # 네이버 톡톡 URL (repr/etc 우선, 없으면 raw text — 모두 정규화/블랙리스트 검사)
                    tt = ""
                    kk = ""
                    ig = ""
                    for src in ([repr_hp] if repr_hp else []) + (etc_links or []):
                        u = (src.get("url") if src else "") or ""
                        if not tt and "talk.naver.com" in u:
                            normalized = self.normalize_talktalk_url(u)
                            if normalized:
                                tt = normalized
                        if not kk and ("pf.kakao.com" in u or "plus.kakao.com" in u or "kakaochannel.com" in u or "open.kakao.com" in u):
                            kk = u
                        if not ig and "instagram.com" in u:
                            ig = u
                    if not tt:
                        tt = self.extract_talktalk_url(raw_text)
                    # 디버그: 톡톡 추출 실패한 경우 GraphQL 응답에 단서가 있었는지 노출
                    if not tt:
                        dbg = CrawlerEngine._last_talktalk_debug
                        if dbg and dbg.get("talk_substring") and (dbg.get("pattern_a_hits", 0) + dbg.get("pattern_b_hits", 0)) > 0:
                            try:
                                rejected = ', '.join(dbg.get("rejected_ids", [])[:3])
                                self.callback("log", f"[talk-debug] GraphQL에 talk 단서 있으나 추출 실패. A={dbg['pattern_a_hits']} B={dbg['pattern_b_hits']} 거절ID=[{rejected}]")
                            except Exception:
                                pass
                    if not kk:
                        kk = self.extract_kakao_url(raw_text)
                    if not ig:
                        ig = self.extract_instagram_url(raw_text)
                    if tt:
                        result["talktalk"] = tt
                    if kk:
                        result["kakao"] = kk
                    if ig:
                        result["instagram"] = ig

                    # 신규개업 여부 (raw text 휴리스틱)
                    if self.detect_new_business(raw_text) == "O":
                        result["is_new"] = "O"

                    self.stats["success"] += 1
            elif resp.status_code in (429, 403):
                self.stats["blocked"] += 1
        except Exception as e:
            self.stats["fail"] += 1

        return result

    # summary API로 블로그/방문자 리뷰 수 보충
    def _fetch_place_summary(self, pid):
        """map.naver.com summary API — blogReviews, visitorReviews 가져오기"""
        fp = self._get_fingerprint()
        result = {}
        try:
            url = f"https://map.naver.com/p/api/place/summary/{pid}"
            resp = self.session.get(url, headers={
                "User-Agent": fp["ua"],
                "Referer": "https://map.naver.com/"
            }, timeout=self.timeout)
            if resp.status_code == 200:
                data = resp.json().get("data", {}).get("placeDetail", {})
                blog = data.get("blogReviews", {})
                result["blog_review"] = str(blog.get("total", 0)) if blog else "0"
                # 방문자 리뷰 (displayText에서 숫자 추출)
                vr = data.get("visitorReviews", {})
                if vr:
                    dt = vr.get("displayText", "")
                    import re
                    m = re.search(r'[\d,]+', dt)
                    if m:
                        result["visitor_review"] = m.group().replace(",", "")
                # 카테고리
                cat = data.get("category", {})
                if cat:
                    result["category"] = cat.get("category", "")
        except Exception:
            pass
        return result

    def _find_naver_id(self, name, parsed_place=None):
        """네이버 블로그 ID 추출 (최적화: 1단계 성공률 90%+).
        1단계: place HTML에서 BaseNaverBlog/blogId/blog.naver.com (이미 파싱됨, 추가 요청 0)
        2단계: site:blog.naver.com 검색 (1단계 실패 시만)
        3단계: 없으면 빈칸 (3단계 "블로그" 검색 제거 → 속도 우선)
        반환: "blogid@naver.com" 또는 ""
        """
        _SKIP = ("PostView", "PostList", "BlogTagLog", "NBlogTop", "prologue", "MologPost", "profile", "intro")

        # 1단계: place HTML 파싱 결과 (추가 HTTP 0회)
        if parsed_place and parsed_place.get("naver_id"):
            return parsed_place["naver_id"] + "@naver.com"

        clean = self.clean_business_name(name)

        # 2단계: site:blog.naver.com 검색 (1회 요청)
        url = f"https://search.naver.com/search.naver?query={quote(clean + ' site:blog.naver.com')}"
        html = self._fetch(url, referer="https://www.naver.com/", timeout=3)
        if html:
            for m in re.finditer(r'blog\.naver\.com/([a-zA-Z0-9_.-]+)', html):
                bid = m.group(1)
                if bid not in _SKIP and len(bid) >= 3:
                    return bid + "@naver.com"

        # 3단계: 없음
        return ""

    # ═══════════════════════════════════════════
    # CSV/XLSX IO
    # ═══════════════════════════════════════════

    OUT_HEADERS = [
        "업종(키워드)", "상호명", "네이버아이디", "업체이메일", "안심번호", "업체주소",
        "홈페이지URL", "홈페이지반응형", "전화버튼없음", "웹빌더",
        "톡톡", "톡톡URL", "카카오톡", "인스타그램",
        "방문자리뷰수", "블로그리뷰수", "리뷰합계",
        "신규업체",
        "고유번호", "플레이스URL", "업데이트날짜",
    ]

    @staticmethod
    def read_csv(filepath):
        """CSV 파일 읽기 (UTF-8 BOM 지원)"""
        with open(filepath, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            return list(reader)

    @staticmethod
    def save_csv(rows, filepath):
        """CSV 파일 저장 (UTF-8 BOM)"""
        with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=CrawlerEngine.OUT_HEADERS, extrasaction="ignore")
            writer.writeheader()
            for r in rows:
                # 리뷰합계 계산 (방문자 + 블로그)
                r["리뷰합계"] = CrawlerEngine._compute_review_total(r)
                out = {}
                for h in CrawlerEngine.OUT_HEADERS:
                    val = r.get(h, "")
                    if h in (
                        "홈페이지URL", "홈페이지반응형", "전화버튼없음", "웹빌더",
                        "톡톡", "카카오톡", "인스타그램",
                        "신규업체",
                    ) and (not val or not str(val).strip()):
                        val = "X"
                    if h == "업데이트날짜" and not val:
                        val = time.strftime("%Y-%m-%d")
                    out[h] = val
                writer.writerow(out)

    @staticmethod
    def save_xlsx(rows, filepath):
        """엑셀 파일 저장"""
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        ws = wb.active
        ws.title = "크롤링결과"

        # 헤더
        for col, h in enumerate(CrawlerEngine.OUT_HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)

        # 데이터
        for row_idx, r in enumerate(rows, 2):
            # 리뷰합계 계산 (방문자 + 블로그)
            r["리뷰합계"] = CrawlerEngine._compute_review_total(r)
            for col, h in enumerate(CrawlerEngine.OUT_HEADERS, 1):
                val = r.get(h, "")
                if h in (
                    "홈페이지URL", "홈페이지반응형", "전화버튼없음", "웹빌더",
                    "톡톡", "카카오톡", "인스타그램",
                    "신규업체",
                ) and (not val or not str(val).strip()):
                    val = "X"
                if h == "업데이트날짜" and not val:
                    val = time.strftime("%Y-%m-%d")
                if h in ("방문자리뷰수", "블로그리뷰수", "리뷰합계"):
                    try:
                        val = int(val) if val not in ("", None) else 0
                    except (ValueError, TypeError):
                        val = 0
                ws.cell(row=row_idx, column=col, value=val)

        # 컬럼 너비 자동 조정
        for col, h in enumerate(CrawlerEngine.OUT_HEADERS, 1):
            max_len = len(h)
            for row_idx in range(2, min(len(rows) + 2, 102)):  # 최대 100행 샘플
                cell_val = str(ws.cell(row=row_idx, column=col).value or "")
                max_len = max(max_len, len(cell_val))
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max(max_len + 2, 10), 50)

        wb.save(filepath)

    def save_output(self, rows, filepath, keyword=None):
        """확장자에 따라 CSV 또는 XLSX 저장. keyword 지정 시 업종 우선 정렬 적용."""
        save_rows = self._sort_by_category(rows, keyword) if keyword else rows
        if filepath.lower().endswith(".xlsx"):
            self.save_xlsx(save_rows, filepath)
        else:
            self.save_csv(save_rows, filepath)

    # ═══════════════════════════════════════════
    # 메인 크롤링 로직
    # ═══════════════════════════════════════════

    def run(self, input_file, output_file, progress_file=None):
        """
        12개 항목 전체 크롤링 실행.
        입력 CSV에 "상호명"만 있으면 나머지 전부 자동.

        Args:
            input_file: 입력 CSV 파일 경로
            output_file: 출력 파일 경로 (.csv 또는 .xlsx)
            progress_file: 진행 저장 파일 (None이면 자동 생성)
        """
        self.running = True
        self.stats = {
            "place_id": 0, "category": 0, "name": 0, "address": 0,
            "phone": 0, "visitor_review": 0, "blog_review": 0,
            "hp": 0, "email": 0, "naver_id": 0,
            "responsive": 0, "talktalk": 0, "kakao": 0, "instagram": 0,
            "no_phone_btn": 0, "webbuilder": 0, "new_biz": 0,
            "blocked": 0, "success": 0,
        }

        if not progress_file:
            progress_file = output_file.rsplit(".", 1)[0] + "-progress.json"

        rows = self.read_csv(input_file)
        self.callback("log", f"입력: {input_file} ({len(rows)}건)")
        self.callback("log", f"출력: {output_file}")
        self.callback("log", f"딜레이: {self.delay_min}~{self.delay_max}초 (랜덤)")
        self.callback("log", f"모드: 12개 항목 전체 크롤링 (상호명 기반)")
        if self.api_keys:
            self.callback("log", f"네이버 API 키 {len(self.api_keys)}개 (로테이션)")
        if self.proxies:
            self.callback("log", f"프록시: {len(self.proxies)}개")
        self.callback("total", len(rows))

        # 이전 진행 복구
        progress = {}
        try:
            with open(progress_file, "r") as f:
                progress = json.load(f)
        except Exception:
            pass

        for i, r in enumerate(rows):
            if not self.running:
                self.callback("log", "사용자 중지")
                break

            name = (r.get("상호명") or "").strip()
            if not name:
                self.callback("log", f"[{i+1}/{len(rows)}] (상호명 없음 - 스킵)")
                self.callback("progress", i + 1)
                continue

            pid = (r.get("고유번호") or "").strip()
            naver_id = (r.get("네이버아이디") or "").replace("@naver.com", "").strip()
            log_items = []

            # ── 이전 진행 복구 ──
            prog_key = pid or name  # 고유번호 없으면 상호명을 키로
            if prog_key in progress:
                prev = progress[prog_key]
                for csv_col, prog_col in [
                    ("고유번호", "pid"), ("플레이스URL", "place_url"),
                    ("업종(키워드)", "category"), ("상호명", "name"),
                    ("업체주소", "address"), ("안심번호", "phone"),
                    ("방문자리뷰수", "visitor_review"), ("블로그리뷰수", "blog_review"),
                    ("리뷰합계", "review_total"),
                    ("홈페이지URL", "hp"),
                    ("홈페이지반응형", "responsive"),
                    ("전화버튼없음", "no_phone_btn"),
                    ("웹빌더", "webbuilder"),
                    ("톡톡", "talktalk_flag"),
                    ("톡톡URL", "talktalk_url"),
                    ("카카오톡", "kakao"),
                    ("인스타그램", "instagram"),
                    ("신규업체", "is_new"),
                    ("업체이메일", "email"),
                ]:
                    val = prev.get(prog_col, "")
                    if val and not (r.get(csv_col) or "").strip():
                        r[csv_col] = val
                # hp 특수처리 (X도 빈값)
                if prev.get("hp") and ((r.get("홈페이지URL") or "").strip() in ("", "X")):
                    r["홈페이지URL"] = prev["hp"]

            # ══ STEP 0: 고유번호 + 플레이스URL ══
            pid = (r.get("고유번호") or "").strip()
            if not pid:
                pid = self._find_place_id(name)
                if pid:
                    r["고유번호"] = pid
                    self.stats["place_id"] += 1
                    log_items.append(f"pid:{pid}")
                self._random_delay()

            if pid and not (r.get("플레이스URL") or "").strip():
                r["플레이스URL"] = f"https://m.place.naver.com/place/{pid}"

            # ══ STEP 1: 플레이스 HTML 파싱 (한 번에 여러 항목) ══
            parsed_place = {}
            if pid:
                need_any = (
                    not (r.get("업종(키워드)") or "").strip()
                    or not (r.get("업체주소") or "").strip()
                    or not (r.get("안심번호") or "").strip()
                    or not (r.get("방문자리뷰수") or "").strip()
                    or not (r.get("블로그리뷰수") or "").strip()
                    or not (r.get("네이버아이디") or "").strip()
                )
                if need_any:
                    place_html = self._fetch_place_html(pid, name)
                    parsed_place = self._parse_place_html(place_html)

                    # 네이버 톡톡 URL (place HTML 1차)
                    if parsed_place.get("talktalk") and not (r.get("톡톡URL") or "").strip():
                        r["톡톡URL"] = parsed_place["talktalk"]
                        r["톡톡"] = "O"
                        self.stats["talktalk"] += 1
                        log_items.append("talk:O")

                    # 카카오톡 URL (place HTML 1차)
                    if parsed_place.get("kakao") and not (r.get("카카오톡") or "").strip():
                        r["카카오톡"] = "O"
                        self.stats["kakao"] += 1
                        log_items.append("kakao:O")

                    # 인스타그램 URL (place HTML 1차)
                    if parsed_place.get("instagram") and not (r.get("인스타그램") or "").strip():
                        r["인스타그램"] = "O"
                        self.stats["instagram"] += 1
                        log_items.append("ig:O")

                    # 신규업체 (place HTML 1차)
                    if parsed_place.get("is_new") == "O" and not (r.get("신규업체") or "").strip():
                        r["신규업체"] = "O"
                        self.stats["new_biz"] += 1
                        log_items.append("new:O")

                    # 업종(키워드)
                    if parsed_place.get("category") and not (r.get("업종(키워드)") or "").strip():
                        r["업종(키워드)"] = parsed_place["category"]
                        self.stats["category"] += 1
                        log_items.append(f"cat:{parsed_place['category']}")

                    # 상호명 보정 (CSV에 없으면)
                    if parsed_place.get("name") and not name:
                        r["상호명"] = parsed_place["name"]
                        name = parsed_place["name"]
                        self.stats["name"] += 1

                    # 업체주소
                    if parsed_place.get("address") and not (r.get("업체주소") or "").strip():
                        r["업체주소"] = parsed_place["address"]
                        self.stats["address"] += 1
                        log_items.append(f"addr:{parsed_place['address'][:20]}")

                    # 안심번호
                    if parsed_place.get("phone") and not (r.get("안심번호") or "").strip():
                        r["안심번호"] = parsed_place["phone"]
                        self.stats["phone"] += 1
                        log_items.append(f"tel:{parsed_place['phone']}")

                    # 방문자리뷰수
                    if parsed_place.get("visitor_review") and not (r.get("방문자리뷰수") or "").strip():
                        r["방문자리뷰수"] = parsed_place["visitor_review"]
                        self.stats["visitor_review"] += 1
                        log_items.append(f"visit:{parsed_place['visitor_review']}")

                    # 블로그리뷰수
                    if parsed_place.get("blog_review") and not (r.get("블로그리뷰수") or "").strip():
                        r["블로그리뷰수"] = parsed_place["blog_review"]
                        self.stats["blog_review"] += 1
                        log_items.append(f"blog:{parsed_place['blog_review']}")

                    self._random_delay()

            # ══ STEP 2: 홈페이지 (API 우선 → HTML 폴백) ══
            need_hp = not (r.get("홈페이지URL") or "").strip() or (r.get("홈페이지URL") or "").strip() == "X"
            if need_hp:
                clean = self.clean_business_name(name)
                found_hp = False

                # API 모드
                if self.api_keys:
                    items = self._search_naver_api(clean)
                    for item in items:
                        link = item.get("link", "").strip()
                        tel = item.get("telephone", "").strip()
                        if link and not any(b in link.lower() for b in HP_BLACKLIST):
                            r["홈페이지URL"] = link
                            self.stats["hp"] += 1
                            log_items.append(f"hp:{link}(API)")
                            found_hp = True
                            break
                        if tel and not (r.get("안심번호") or "").strip():
                            r["안심번호"] = tel
                            self.stats["phone"] += 1
                            log_items.append(f"tel:{tel}(API)")

                # HTML 폴백 (검증 포함)
                if not found_hp:
                    url = f"https://search.naver.com/search.naver?query={quote(clean + ' 홈페이지')}"
                    referer = "https://www.naver.com/"
                    html = self._fetch(url, referer=referer)
                    verified_hp = self._extract_homepage_with_verify(html, name)
                    if verified_hp:
                        r["홈페이지URL"] = verified_hp
                        self.stats["hp"] += 1
                        log_items.append(f"hp:{verified_hp}")
                self._random_delay()

            # ══ STEP 3: 이메일 + 홈페이지 분석 (반응형/전화버튼/웹빌더/카카오/인스타 — HTML 1회 재활용) ══
            need_email = not (r.get("업체이메일") or "").strip()
            need_resp = not (r.get("홈페이지반응형") or "").strip()
            need_phone_btn = not (r.get("전화버튼없음") or "").strip()
            need_webbuilder = not (r.get("웹빌더") or "").strip()
            need_kakao = not (r.get("카카오톡") or "").strip()
            need_insta = not (r.get("인스타그램") or "").strip()
            hp_url = (r.get("홈페이지URL") or "").strip()
            if hp_url and hp_url != "X" and hp_url.startswith("http") and (
                need_email or need_resp or need_phone_btn or need_webbuilder or need_kakao or need_insta
            ):
                referer = f"https://www.google.com/search?q={quote(name)}"
                html = self._fetch(hp_url, referer=referer, timeout=6)
                # 이메일
                if need_email:
                    emails = self.extract_emails(html)
                    if emails:
                        r["업체이메일"] = emails[0]
                        self.stats["email"] += 1
                        log_items.append(f"email:{emails[0]}")
                # 반응형
                if need_resp and html:
                    resp_flag = self.detect_responsive(html)
                    r["홈페이지반응형"] = resp_flag
                    if resp_flag == "O":
                        self.stats["responsive"] += 1
                    log_items.append(f"resp:{resp_flag}")
                # 전화버튼 (없음 = O = 영업타겟)
                if need_phone_btn:
                    has_btn = self.detect_phone_button(html) if html else False
                    flag = "X" if has_btn else "O"
                    r["전화버튼없음"] = flag
                    if flag == "O":
                        self.stats["no_phone_btn"] += 1
                    log_items.append(f"nophonebtn:{flag}")
                # 웹빌더 (O = 웹빌더 사용 / X = 일반 HTML, 영업 타겟)
                if need_webbuilder:
                    wb_flag = self.detect_webbuilder(html, url=hp_url)
                    r["웹빌더"] = wb_flag
                    if wb_flag == "O":
                        self.stats["webbuilder"] += 1
                    log_items.append(f"wb:{wb_flag}")
                # 카카오톡
                if need_kakao and html:
                    kk = self.extract_kakao_url(html)
                    if kk:
                        r["카카오톡"] = "O"
                        self.stats["kakao"] += 1
                        log_items.append("kakao:O(hp)")
                # 인스타그램
                if need_insta and html:
                    ig = self.extract_instagram_url(html)
                    if ig:
                        r["인스타그램"] = "O"
                        self.stats["instagram"] += 1
                        log_items.append("ig:O(hp)")
                self._random_delay()

            # 이메일 폴백 (네이버 검색)
            need_email = not (r.get("업체이메일") or "").strip()
            if need_email:
                clean = self.clean_business_name(name)
                url = f"https://search.naver.com/search.naver?query={quote(clean + ' 이메일')}"
                referer = "https://www.naver.com/"
                html = self._fetch(url, referer=referer)
                emails = self.extract_emails(html)
                if emails:
                    r["업체이메일"] = emails[0]
                    self.stats["email"] += 1
                    log_items.append(f"email:{emails[0]}(검색)")
                self._random_delay()

            # ══ STEP 4: 네이버아이디 (4단계 폴백) ══
            need_nid = not (r.get("네이버아이디") or "").strip()
            if need_nid:
                nid = self._find_naver_id(name, parsed_place)
                if nid:
                    r["네이버아이디"] = nid
                    self.stats["naver_id"] += 1
                    log_items.append(f"nid:{nid}")
            elif naver_id and "@" not in (r.get("네이버아이디") or ""):
                r["네이버아이디"] = naver_id + "@naver.com"
                self.stats["naver_id"] += 1

            # ══ STEP 4-1: 톡톡/카카오/인스타/신규업체 GraphQL 보충 (place HTML에서 못 채운 경우) ══
            need_talk = not (r.get("톡톡URL") or "").strip()
            need_kakao2 = not (r.get("카카오톡") or "").strip()
            need_insta2 = not (r.get("인스타그램") or "").strip()
            need_new = not (r.get("신규업체") or "").strip()
            if pid and (need_talk or need_kakao2 or need_insta2 or need_new):
                gql_extra = self._fetch_place_graphql(pid)
                if need_talk and gql_extra.get("talktalk"):
                    r["톡톡URL"] = gql_extra["talktalk"]
                    r["톡톡"] = "O"
                    self.stats["talktalk"] += 1
                    log_items.append("talk:O(gql)")
                if need_kakao2 and gql_extra.get("kakao"):
                    r["카카오톡"] = "O"
                    self.stats["kakao"] += 1
                    log_items.append("kakao:O(gql)")
                if need_insta2 and gql_extra.get("instagram"):
                    r["인스타그램"] = "O"
                    self.stats["instagram"] += 1
                    log_items.append("ig:O(gql)")
                if need_new and gql_extra.get("is_new") == "O":
                    r["신규업체"] = "O"
                    self.stats["new_biz"] += 1
                    log_items.append("new:O(gql)")
                self._random_delay()

            # 미감지 항목은 X로 확정 (톡톡URL은 빈값 유지 — 복사 편의)
            if not (r.get("톡톡") or "").strip():
                r["톡톡"] = "X"
            if not (r.get("카카오톡") or "").strip():
                r["카카오톡"] = "X"
            if not (r.get("인스타그램") or "").strip():
                r["인스타그램"] = "X"
            if not (r.get("신규업체") or "").strip():
                r["신규업체"] = "X"
            if not (r.get("홈페이지반응형") or "").strip():
                r["홈페이지반응형"] = "X"
            if not (r.get("전화버튼없음") or "").strip():
                # 홈페이지 자체가 없거나 검사 실패 → 영업 타겟 아님 (X)
                r["전화버튼없음"] = "X"
            if not (r.get("웹빌더") or "").strip():
                # 홈페이지 없음/검사 실패 → 웹빌더 아님 = X (영업 가능 추정)
                r["웹빌더"] = "X"

            # 리뷰합계 계산 (방문자 + 블로그)
            r["리뷰합계"] = self._compute_review_total(r)

            # ══ STEP 5: 업데이트날짜 ══
            if not (r.get("업데이트날짜") or "").strip():
                r["업데이트날짜"] = time.strftime("%Y-%m-%d")

            # ── 진행 저장 (5건마다) ──
            prog_key = pid or name
            progress[prog_key] = {
                "pid": r.get("고유번호", ""),
                "place_url": r.get("플레이스URL", ""),
                "category": r.get("업종(키워드)", ""),
                "name": r.get("상호명", ""),
                "address": r.get("업체주소", ""),
                "phone": r.get("안심번호", ""),
                "visitor_review": r.get("방문자리뷰수", ""),
                "blog_review": r.get("블로그리뷰수", ""),
                "review_total": r.get("리뷰합계", ""),
                "hp": r.get("홈페이지URL", ""),
                "responsive": r.get("홈페이지반응형", ""),
                "no_phone_btn": r.get("전화버튼없음", ""),
                "webbuilder": r.get("웹빌더", ""),
                "talktalk_flag": r.get("톡톡", ""),
                "talktalk_url": r.get("톡톡URL", ""),
                "kakao": r.get("카카오톡", ""),
                "instagram": r.get("인스타그램", ""),
                "is_new": r.get("신규업체", ""),
                "email": r.get("업체이메일", ""),
                "naver_id": r.get("네이버아이디", ""),
                "update_date": r.get("업데이트날짜", ""),
            }
            if i % 5 == 0 or i == len(rows) - 1:
                try:
                    with open(progress_file, "w") as f:
                        json.dump(progress, f, ensure_ascii=False)
                    self.save_output(rows, output_file)
                except Exception as e:
                    logger.debug(f"save error: {e}")

            log_msg = f"[{i+1}/{len(rows)}] {name}"
            if log_items:
                log_msg += " -> " + " ".join(log_items)
            self.callback("log", log_msg)
            self.callback("progress", i + 1)
            self.callback("stats", dict(self.stats))

        # 최종 저장
        self.save_output(rows, output_file)
        try:
            os.remove(progress_file)
        except Exception:
            pass

        # 최종 통계 — 12개 항목별 채움율
        total = len(rows)

        def _pct(field, empty_vals=("", "X")):
            cnt = sum(1 for r in rows if (r.get(field) or "").strip() not in empty_vals)
            return cnt, round(cnt / total * 100) if total else 0

        c_cat, p_cat = _pct("업종(키워드)")
        c_name, p_name = _pct("상호명")
        c_email, p_email = _pct("업체이메일")
        c_phone, p_phone = _pct("안심번호")
        c_addr, p_addr = _pct("업체주소")
        c_hp, p_hp = _pct("홈페이지URL")
        c_resp, p_resp = _pct("홈페이지반응형", empty_vals=("", "X"))
        c_nopb, p_nopb = _pct("전화버튼없음", empty_vals=("", "X"))
        c_wb, p_wb = _pct("웹빌더", empty_vals=("", "X"))
        c_talk, p_talk = _pct("톡톡", empty_vals=("", "X"))
        c_kakao, p_kakao = _pct("카카오톡", empty_vals=("", "X"))
        c_insta, p_insta = _pct("인스타그램", empty_vals=("", "X"))
        c_new, p_new = _pct("신규업체", empty_vals=("", "X"))
        c_visit, p_visit = _pct("방문자리뷰수")
        c_blog, p_blog = _pct("블로그리뷰수")
        c_nid, p_nid = _pct("네이버아이디")
        c_pid, p_pid = _pct("고유번호")
        c_purl, p_purl = _pct("플레이스URL")
        c_date, p_date = _pct("업데이트날짜")

        s = self.stats
        summary = (
            f"\n=== 최종 결과 ({total}건) ===\n"
            f"  업종(키워드):    {c_cat}/{total} ({p_cat}%)  +{s['category']}\n"
            f"  상호명:          {c_name}/{total} ({p_name}%)  +{s['name']}\n"
            f"  업체이메일:      {c_email}/{total} ({p_email}%)  +{s['email']}\n"
            f"  안심번호:        {c_phone}/{total} ({p_phone}%)  +{s['phone']}\n"
            f"  업체주소:        {c_addr}/{total} ({p_addr}%)  +{s['address']}\n"
            f"  홈페이지URL:     {c_hp}/{total} ({p_hp}%)  +{s['hp']}\n"
            f"  홈페이지반응형:  {c_resp}/{total} ({p_resp}%)  +{s.get('responsive', 0)}\n"
            f"  전화버튼없음:    {c_nopb}/{total} ({p_nopb}%)  +{s.get('no_phone_btn', 0)}\n"
            f"  웹빌더:          {c_wb}/{total} ({p_wb}%)  +{s.get('webbuilder', 0)}\n"
            f"  톡톡:            {c_talk}/{total} ({p_talk}%)  +{s.get('talktalk', 0)}\n"
            f"  카카오톡:        {c_kakao}/{total} ({p_kakao}%)  +{s.get('kakao', 0)}\n"
            f"  인스타그램:      {c_insta}/{total} ({p_insta}%)  +{s.get('instagram', 0)}\n"
            f"  신규업체:        {c_new}/{total} ({p_new}%)  +{s.get('new_biz', 0)}\n"
            f"  방문자리뷰수:    {c_visit}/{total} ({p_visit}%)  +{s['visitor_review']}\n"
            f"  블로그리뷰수:    {c_blog}/{total} ({p_blog}%)  +{s['blog_review']}\n"
            f"  네이버아이디:    {c_nid}/{total} ({p_nid}%)  +{s['naver_id']}\n"
            f"  고유번호:        {c_pid}/{total} ({p_pid}%)  +{s['place_id']}\n"
            f"  플레이스URL:     {c_purl}/{total} ({p_purl}%)\n"
            f"  업데이트날짜:    {c_date}/{total} ({p_date}%)\n"
            f"\n완료: {output_file}"
        )
        self.callback("log", summary)
        self.callback("done", output_file)
        self.running = False

    # ═══════════════════════════════════════════
    # 키워드 검색 모드 — 전체 페이지 크롤링
    # ═══════════════════════════════════════════

    # ═══════════════════════════════════════════
    # Selenium 순위 크롤링 (map.naver.com)
    # ═══════════════════════════════════════════

    _naver_cookies = None  # 클래스 레벨 쿠키 캐시

    def _selenium_naver_login(self, driver, naver_id, naver_pw):
        """Selenium으로 네이버 자동 로그인 → 쿠키 저장.
        
        pyperclip(클립보드)을 사용하여 입력 감지 우회.
        """
        from selenium.webdriver.common.by import By
        import pyperclip

        self.callback("log", "  🔑 네이버 로그인 시도...")
        driver.get("https://nid.naver.com/nidlogin.login?mode=form&url=https://map.naver.com")
        time.sleep(2 + random.random())

        try:
            # ID 입력 (클립보드 붙여넣기로 자동입력 감지 우회)
            id_input = driver.find_element(By.CSS_SELECTOR, "input#id")
            id_input.click()
            time.sleep(0.3)
            
            try:
                pyperclip.copy(naver_id)
                id_input.send_keys(__import__('selenium.webdriver.common.keys', fromlist=['Keys']).Keys.CONTROL, 'v')
            except Exception:
                # pyperclip 없으면 JS로 직접 값 설정
                driver.execute_script(f"document.querySelector('#id').value = '{naver_id}';")
            time.sleep(0.5 + random.random() * 0.5)

            # PW 입력
            pw_input = driver.find_element(By.CSS_SELECTOR, "input#pw")
            pw_input.click()
            time.sleep(0.3)
            
            try:
                pyperclip.copy(naver_pw)
                pw_input.send_keys(__import__('selenium.webdriver.common.keys', fromlist=['Keys']).Keys.CONTROL, 'v')
            except Exception:
                driver.execute_script(f"document.querySelector('#pw').value = '{naver_pw}';")
            time.sleep(0.5 + random.random() * 0.5)

            # 로그인 버튼 클릭
            login_btn = driver.find_element(By.CSS_SELECTOR, "button.btn_login, #log\\.login, button[type='submit']")
            login_btn.click()
            time.sleep(3 + random.random() * 2)

            # 캡차/2FA 체크
            current_url = driver.current_url
            if "nidlogin" in current_url or "captcha" in current_url.lower():
                # 캡차 또는 추가 인증 필요
                page_text = driver.find_element(By.TAG_NAME, "body").text
                if "자동입력" in page_text or "captcha" in page_text.lower():
                    self.callback("log", "  ❌ 캡차 감지 — 로그인 실패 (잠시 후 재시도)")
                    return False
                elif "새로운 기기" in page_text or "본인확인" in page_text:
                    self.callback("log", "  ❌ 2단계 인증 필요 — 쿠키 방식으로 전환 권장")
                    return False
                else:
                    self.callback("log", "  ❌ 로그인 실패 — ID/PW 확인")
                    return False

            # 로그인 성공 확인
            cookies = driver.get_cookies()
            nid_cookies = [c for c in cookies if "NID" in c.get("name", "") or "nid" in c.get("name", "")]
            if nid_cookies:
                # 쿠키 저장
                CrawlerEngine._naver_cookies = cookies
                cookie_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "naver_cookies.json")
                try:
                    with open(cookie_path, "w", encoding="utf-8") as f:
                        json.dump(cookies, f, ensure_ascii=False)
                    self.callback("log", f"  ✅ 로그인 성공! 쿠키 저장됨")
                except Exception:
                    self.callback("log", f"  ✅ 로그인 성공! (쿠키 파일 저장 실패)")
                return True
            else:
                self.callback("log", "  ❌ 로그인 실패 — 쿠키 없음")
                return False

        except Exception as e:
            self.callback("log", f"  ❌ 로그인 오류: {e}")
            return False

    def _selenium_load_cookies(self, driver):
        """저장된 네이버 쿠키를 Selenium 드라이버에 로드."""
        cookies = CrawlerEngine._naver_cookies
        if not cookies:
            cookie_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "naver_cookies.json")
            if os.path.isfile(cookie_path):
                try:
                    with open(cookie_path, "r", encoding="utf-8") as f:
                        cookies = json.load(f)
                    CrawlerEngine._naver_cookies = cookies
                except Exception:
                    return False
            else:
                return False

        # 먼저 naver.com 도메인으로 이동해야 쿠키 설정 가능
        driver.get("https://map.naver.com")
        time.sleep(2)
        for cookie in cookies:
            try:
                # 필수 필드만
                c = {
                    "name": cookie["name"],
                    "value": cookie["value"],
                    "domain": cookie.get("domain", ".naver.com"),
                }
                if cookie.get("path"):
                    c["path"] = cookie["path"]
                driver.add_cookie(c)
            except Exception:
                continue
        return True

    def _selenium_rank_search(self, keyword, max_pages=0, naver_id=None, naver_pw=None):
        """Selenium으로 네이버 지도 검색 → 실제 순위 + place ID 수집.

        naver_id/naver_pw가 있으면 로그인 후 크롤링 (차단 위험 감소).

        Returns:
            list of {"id": str, "name": str, "rank": int, "page": int} 또는 None (실패/차단 시)
        """
        try:
            from selenium import webdriver
            from selenium.webdriver.chrome.options import Options
            from selenium.webdriver.chrome.service import Service
            from selenium.webdriver.common.by import By
        except ImportError:
            self.callback("log", "  ⚠️ selenium 미설치 — pip install selenium")
            return None

        try:
            from webdriver_manager.chrome import ChromeDriverManager
            use_wdm = True
        except ImportError:
            use_wdm = False

        import re as _re
        pid_pattern = _re.compile(r'/place/(\d+)')

        # ── 프록시 재시도 루프 (최대 3회) ──
        max_proxy_retries = 3 if self.proxies else 1
        sel_proxy = self._get_next_proxy() if self.proxies else None

        for _proxy_attempt in range(max_proxy_retries):
            opts = Options()
            # 최소화 모드 — 실제 렌더링하되 창은 숨김 (네이버 봇 감지 우회)
            opts.add_argument("--window-position=-32000,-32000")
            opts.add_argument("--disable-blink-features=AutomationControlled")
            opts.add_argument("--no-sandbox")
            opts.add_argument("--disable-dev-shm-usage")
            opts.add_argument("--disable-gpu")
            opts.add_argument("--window-size=1920,1080")
            opts.add_experimental_option("excludeSwitches", ["enable-automation"])
            opts.add_experimental_option("useAutomationExtension", False)
            # Performance 로그 (GraphQL 쿼리 캡처용)
            opts.set_capability("goog:loggingPrefs", {"performance": "ALL"})
            fp = self._get_fingerprint()
            opts.add_argument(f"user-agent={fp['ua']}")

            # 프록시 적용
            if sel_proxy:
                if sel_proxy.get("user") and sel_proxy.get("pw"):
                    _proxy_url = f"http://{sel_proxy['user']}:{sel_proxy['pw']}@{sel_proxy['ip']}:{sel_proxy['port']}"
                else:
                    _proxy_url = f"http://{sel_proxy['ip']}:{sel_proxy['port']}"
                opts.add_argument(f"--proxy-server={_proxy_url}")
                self.callback("log", f"  🔀 Selenium 프록시: {sel_proxy['ip']}:{sel_proxy['port']}")

            try:
                if use_wdm:
                    service = Service(ChromeDriverManager().install())
                    driver = webdriver.Chrome(service=service, options=opts)
                else:
                    driver = webdriver.Chrome(options=opts)
            except Exception as e:
                self.callback("log", f"  ⚠️ Chrome 드라이버 실패: {e}")
                return None

            results = []
            page = 1

            try:
                # ── 키워드 띄어쓰기 강제 (POI redirect 회피) ──
                # "강남헬스장" 같은 붙여쓰기 키워드는 네이버가 단일 상호("강남헬스")로 매칭해
                # appLink?pinId=강남헬스 (POI 페이지)로 redirect 시켜 검색결과 페이지를 못 받음.
                # _parse_keyword_location 으로 분리 후 "강남 헬스장" 으로 강제 → 검색의도 명시
                _loc, _biz = self._parse_keyword_location(keyword)
                search_keyword = f"{_loc} {_biz}".strip() if _biz else keyword
                if search_keyword != keyword:
                    self.callback("log", f"  ✏️ 키워드 분리: '{keyword}' → '{search_keyword}' (POI redirect 회피)")

                search_url = f"https://map.naver.com/p/search/{quote(search_keyword)}"

                # ── 0단계: map.naver.com 메인 먼저 진입 (쿠키 수집 → appLink redirect 우회) ──
                # 일부 키워드(강남미용실/부산헬스장 등 인기 키워드)는 첫 진입 시 m.map.naver.com/appLink로
                # 강제 redirect됨. 메인 페이지를 먼저 방문해서 쿠키를 미리 받으면 redirect 회피 가능
                try:
                    driver.get("https://map.naver.com/")
                    time.sleep(2 + random.random())
                except Exception:
                    pass

                # ── 1단계: 검색 URL 직접 접속 ──
                driver.get(search_url)
                time.sleep(5 + random.random() * 2)

                # ── 1.5단계: appLink redirect 감지 → 다른 URL로 폴백 ──
                cur_url = driver.current_url or ""
                if "appLink" in cur_url or "m.map.naver.com" in cur_url:
                    self.callback("log", f"  🔁 appLink redirect 감지 — 폴백 URL 시도")
                    # 폴백 URL 후보들 (역시 띄어쓰기 키워드 사용)
                    fallback_urls = [
                        f"https://map.naver.com/v5/search/{quote(search_keyword)}",
                        f"https://map.naver.com/?query={quote(search_keyword)}",
                        f"https://pcmap.place.naver.com/place/list?query={quote(search_keyword)}",
                    ]
                    for fb_url in fallback_urls:
                        try:
                            driver.get(fb_url)
                            time.sleep(3 + random.random())
                            new_url = driver.current_url or ""
                            if "appLink" not in new_url and "m.map.naver.com" not in new_url:
                                self.callback("log", f"  ✅ 폴백 성공: {fb_url[:60]}")
                                break
                        except Exception:
                            continue
                    else:
                        self.callback("log", f"  ⚠️ 모든 폴백 URL이 appLink로 redirect")

                # ── 2단계: iframe 로딩 대기 (최대 15초) ──
                has_iframe = False
                is_blocked = False
                for _wait_iframe in range(15):
                    p_src = driver.page_source
                    has_iframe = "searchIframe" in p_src
                    is_blocked = ("서비스 이용이 제한" in p_src
                                  or "자동등록방지" in p_src)
                    # ncaptcha 오탐 방지: searchIframe 있으면 차단 아님
                    if has_iframe:
                        is_blocked = False
                        break
                    if is_blocked:
                        break
                    time.sleep(1)
                    # 팝업 닫기 시도
                    try:
                        for btn in driver.find_elements(By.CSS_SELECTOR,
                            "button.btn_close, button[aria-label='닫기'], .layer_popup button"):
                            try:
                                btn.click()
                                time.sleep(0.3)
                            except Exception:
                                pass
                    except Exception:
                        pass

                # 차단 감지 시 프록시 교체 재시도
                if is_blocked and not has_iframe:
                    self.callback("log", "  🚫 IP 차단 감지")
                    driver.quit()
                    if sel_proxy and _proxy_attempt < max_proxy_retries - 1:
                        self._mark_blocked(sel_proxy)
                        sel_proxy = self._get_next_proxy()
                        self.callback("log", f"  🔄 프록시 교체 후 재시도 ({_proxy_attempt + 2}/{max_proxy_retries})")
                        continue
                    return None

                if not has_iframe:
                    # 디버그: 어떤 iframe이 있는지 + 페이지 단서 노출
                    try:
                        all_iframes = driver.find_elements(By.TAG_NAME, "iframe")
                        iframe_ids = [(f.get_attribute('id') or f.get_attribute('name') or '?')[:40] for f in all_iframes]
                        page_title = driver.title or ''
                        cur_url = driver.current_url or ''
                        body_preview = ''
                        try:
                            body_preview = driver.find_element(By.TAG_NAME, "body").text[:200].replace('\n', ' ')
                        except Exception:
                            pass
                        self.callback("log", f"  ⚠️ searchIframe 미발견 — title='{page_title[:50]}' url='{cur_url[:80]}'")
                        self.callback("log", f"  발견된 iframe: {iframe_ids if iframe_ids else '(없음)'}")
                        if body_preview:
                            self.callback("log", f"  body미리보기: {body_preview}")
                    except Exception as _e:
                        self.callback("log", f"  ⚠️ searchIframe 미발견 (디버그 실패: {_e})")

                    # 폴백 1: 다른 iframe 이름들 시도 (네이버가 카테고리별로 다른 iframe 사용 가능성)
                    fallback_iframe = None
                    for fb_sel in [
                        "iframe#entryIframe",
                        "iframe#placeListIframe",
                        "iframe#centerSearchIframe",
                        "iframe[name='searchIframe']",
                        "iframe",
                    ]:
                        try:
                            fb_el = driver.find_element(By.CSS_SELECTOR, fb_sel)
                            if fb_el:
                                fallback_iframe = (fb_sel, fb_el)
                                self.callback("log", f"  🔁 폴백 iframe 발견: {fb_sel}")
                                break
                        except Exception:
                            continue
                    if not fallback_iframe:
                        self.callback("log", "  ⚠️ 폴백 iframe도 없음 — Selenium 스킵")
                        driver.quit()
                        return None
                    # 폴백 iframe으로 진행
                    has_iframe = True

                # ── iframe 진입 ──
                driver.switch_to.default_content()
                iframe_found = False
                last_iframe_error = None
                for _retry in range(10):
                    try:
                        # 1차 시도: searchIframe
                        try:
                            iframe = driver.find_element(By.CSS_SELECTOR, "iframe#searchIframe")
                        except Exception:
                            # 2차 시도: 어느 iframe이든
                            iframe = driver.find_element(By.CSS_SELECTOR, "iframe#entryIframe, iframe#placeListIframe, iframe")
                        driver.switch_to.frame(iframe)
                        iframe_found = True
                        break
                    except Exception as _ie:
                        last_iframe_error = str(_ie)[:100]
                        time.sleep(1)
                        try:
                            for btn in driver.find_elements(By.CSS_SELECTOR,
                                "button.btn_close, button[aria-label='닫기'], .layer_popup button"):
                                btn.click()
                                time.sleep(0.3)
                        except Exception:
                            pass

                if not iframe_found:
                    self.callback("log", f"  ⚠️ iframe 전환 실패: {last_iframe_error or 'unknown'}")
                    driver.quit()
                    return None

                time.sleep(1.5 + random.random())

                # ── GraphQL 쿼리 캡처를 위해 2페이지 클릭 ──
                from selenium.webdriver.common.action_chains import ActionChains
                body = driver.find_element(By.TAG_NAME, "body")

                # 스크롤 (페이지 버튼 로드)
                for _s in range(10):
                    driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight", body)
                    time.sleep(0.3)

                # Performance 로그 비우기
                try:
                    driver.get_log("performance")
                except Exception:
                    pass

                # 2페이지 클릭
                driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight", body)
                time.sleep(1)
                gql_clicked = False
                for btn in driver.find_elements(By.CSS_SELECTOR, "a.mBN2s"):
                    try:
                        if btn.text.strip() == "2":
                            driver.execute_script("arguments[0].scrollIntoView({block:'center'})", btn)
                            time.sleep(0.3)
                            ActionChains(driver).move_to_element(btn).pause(0.2).click().perform()
                            gql_clicked = True
                            break
                    except Exception:
                        continue

                if not gql_clicked:
                    self.callback("log", "  ℹ️ 1페이지만 존재 — HTML에서 PID 수집")
                    import json as _json

                    # HTML 소스에서 PID(7자리+ 숫자) 추출 — 이름은 PHASE 2에서 상세 크롤링으로 보정
                    src = driver.page_source
                    pids = []
                    seen_p = set()
                    for m in _re.finditer(r'"id"\s*:\s*"(\d{7,})"', src):
                        pid = m.group(1)
                        if pid not in seen_p:
                            seen_p.add(pid)
                            pids.append(pid)

                    seen_ids = {r["id"] for r in results}
                    for pid in pids:
                        if pid not in seen_ids:
                            seen_ids.add(pid)
                            results.append({"id": pid, "name": "", "rank": len(results) + 1, "page": 1})

                    self.callback("log", f"  → HTML PID {len(results)}건 수집 (이름은 상세에서 보정)")
                    driver.quit()
                    return results if results else None

                time.sleep(4)

                # ── GraphQL 원본 쿼리 캡처 ──
                import json as _json
                logs = driver.get_log("performance")
                gql_payload = None

                # 카테고리별 오퍼레이션명 변동(getNxList 외)에 견고하게:
                # 페이지네이션 가능한 GraphQL POST = variables.input.start + display 둘 다 보유
                for entry in logs:
                    try:
                        log = _json.loads(entry["message"])
                        msg = log.get("message", {})
                        if msg.get("method") != "Network.requestWillBeSent":
                            continue
                        params = msg.get("params", {})
                        req = params.get("request", {})
                        url = req.get("url", "")
                        if "graphql" not in url or req.get("method") != "POST":
                            continue
                        post_data = req.get("postData", "")
                        if not post_data:
                            continue
                        # 빠른 키워드 필터 — start/display 둘 다 있어야 후보
                        if '"start"' not in post_data or '"display"' not in post_data:
                            continue
                        try:
                            candidate = _json.loads(post_data)
                        except Exception:
                            continue
                        if isinstance(candidate, list):
                            candidate = candidate[0] if candidate else None
                        if not isinstance(candidate, dict):
                            continue
                        # variables.input.start/display 가 실제로 있는지 검증 (페이지네이션 가능)
                        v = (candidate.get("variables") or {})
                        inp = (v.get("input") or {}) if isinstance(v, dict) else {}
                        if isinstance(inp, dict) and "start" in inp and "display" in inp:
                            gql_payload = candidate
                            break
                    except Exception:
                        continue

                if not gql_payload:
                    self.callback("log", "  ⚠️ GraphQL 쿼리 캡처 실패 — 페이지 순회 폴백 (1742f6f 패턴)")
                    # 폴백: 현재 페이지에서 PID 수집 → 다음 페이지 클릭 → 반복
                    visited = 0
                    no_change_rounds = 0
                    max_pg = max_pages if max_pages > 0 else 20
                    while visited < max_pg:
                        if not self.running:
                            break
                        visited += 1
                        # 1) 현재 페이지 PID 추출 (HTML page_source)
                        try:
                            src = driver.page_source
                        except Exception:
                            break
                        existing = {r["id"] for r in results}
                        added = 0
                        for m in _re.finditer(r'"id"\s*:\s*"(\d{5,})"', src):
                            pid = m.group(1)
                            if pid not in existing:
                                existing.add(pid)
                                results.append({"id": pid, "name": "", "rank": len(results) + 1, "page": visited})
                                added += 1
                        self.callback("log", f"  📊 폴백 [{visited}회차]: +{added}건 (누적 {len(results)}건)")
                        if added == 0:
                            no_change_rounds += 1
                            if no_change_rounds >= 2:
                                break
                        else:
                            no_change_rounds = 0
                        # 2) 다음 페이지 클릭 — DOM에서 활성 페이지 +1 시도, 실패 시 "다음" 버튼
                        try:
                            driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight", body)
                            time.sleep(0.4)
                        except Exception:
                            pass
                        clicked = False
                        # 활성 페이지 번호 탐색
                        active_text = ""
                        try:
                            for sel in ["a.mBN2s[aria-current='true']", "a[aria-current='page']", "a.qxokY", "a.mBN2s.qxokY"]:
                                if active_text:
                                    break
                                for el in driver.find_elements(By.CSS_SELECTOR, sel):
                                    t = (el.text or "").strip()
                                    if t.isdigit():
                                        active_text = t
                                        break
                        except Exception:
                            pass
                        # 숫자 페이지 클릭 시도
                        target_nums = []
                        if active_text and active_text.isdigit():
                            target_nums.append(str(int(active_text) + 1))
                        # 보조: 1742f6f 처럼 단순히 visited+1 도 시도
                        if str(visited + 1) not in target_nums:
                            target_nums.append(str(visited + 1))
                        for tgt in target_nums:
                            if clicked:
                                break
                            for sel in ["a.mBN2s", "a[role='button']"]:
                                if clicked:
                                    break
                                for btn in driver.find_elements(By.CSS_SELECTOR, sel):
                                    try:
                                        if btn.text.strip() == tgt:
                                            driver.execute_script("arguments[0].scrollIntoView({block:'center'})", btn)
                                            time.sleep(0.2)
                                            ActionChains(driver).move_to_element(btn).pause(0.15).click().perform()
                                            clicked = True
                                            time.sleep(2.0 + random.random())
                                            break
                                    except Exception:
                                        continue
                        # "다음" 버튼 폴백
                        if not clicked:
                            try:
                                for btn in driver.find_elements(By.CSS_SELECTOR, "a, button"):
                                    try:
                                        aria = btn.get_attribute("aria-label") or ""
                                        cls = (btn.get_attribute("class") or "").lower()
                                        txt = (btn.text or "").strip()
                                        if "다음" in aria or txt == "다음" or "next" in cls:
                                            btn.click()
                                            clicked = True
                                            time.sleep(2.0 + random.random())
                                            break
                                    except Exception:
                                        continue
                            except Exception:
                                pass
                        if not clicked:
                            self.callback("log", f"  ℹ️ 다음 페이지 버튼 없음 → 폴백 종료 (총 {len(results)}건)")
                            break
                    driver.quit()
                    return results if results else None

                self.callback("log", "  ✅ GraphQL 쿼리 캡처 성공")

                # ── 페이지 1~N GraphQL 병렬 호출 (Promise.all) ──
                import base64 as _b64
                max_gql_pages = max_pages if max_pages > 0 else 20

                # 1차: 페이지 1~max를 display=70으로 병렬 호출
                payloads_b64 = []
                for pg in range(1, max_gql_pages + 1):
                    p = _json.loads(_json.dumps(gql_payload))
                    p["variables"]["input"]["start"] = 1 + (pg - 1) * 70
                    p["variables"]["input"]["display"] = 70
                    payloads_b64.append(_b64.b64encode(_json.dumps([p]).encode("utf-8")).decode("ascii"))

                # JS: Promise.all로 병렬 fetch
                fetches_js = ",\n".join([
                    f"fetch('https://pcmap-api.place.naver.com/graphql',{{method:'POST',headers:{{'Content-Type':'application/json'}},body:atob('{b}'),credentials:'include'}}).then(r=>r.text().then(t=>r.status+'|||'+t)).catch(e=>'ERROR:'+e.message)"
                    for b in payloads_b64
                ])
                js = f"try{{ return await Promise.all([{fetches_js}]); }}catch(e){{ return ['ERROR:'+e.message]; }}"

                try:
                    all_results = driver.execute_script(js)
                except Exception as e:
                    self.callback("log", f"  ⚠️ 병렬 fetch 에러: {e}")
                    all_results = []

                # 결과 파싱
                last_total = 0
                for pg_idx, result in enumerate(all_results or []):
                    pg = pg_idx + 1
                    if not self.running:
                        break

                    if not result or result.startswith("ERROR"):
                        self.callback("log", f"  ⚠️ 페이지 {pg} 에러: {str(result)[:100]}")
                        break

                    status, body_text = result.split("|||", 1)
                    if status != "200":
                        self.callback("log", f"  ⚠️ 페이지 {pg}: HTTP {status}")
                        break

                    try:
                        data = _json.loads(body_text)
                        if isinstance(data, list):
                            data = data[0]
                        biz = (data or {}).get("data", {}) or {}
                        biz = biz.get("businesses", {}) or {}
                        total = biz.get("total", 0) or 0
                        items = biz.get("items", []) or []
                        if total and total > 0:
                            last_total = total

                        if not items:
                            break

                        page_items = 0
                        for it in items:
                            pid = str(it.get("id", ""))
                            name = it.get("name", "")
                            if pid and pid not in {r["id"] for r in results}:
                                results.append({"id": pid, "name": name, "rank": len(results) + 1, "page": pg})
                                page_items += 1

                        self.callback("log", f"  📊 페이지 {pg}: +{page_items}건 (누적 {len(results)}건)")

                    except Exception as e:
                        self.callback("log", f"  ⚠️ 페이지 {pg} 파싱 에러: {e}")
                        break

                # 2차: 마지막 페이지 0건이면 display 줄여서 보충
                if last_total > 0 and len(results) < last_total:
                    remaining = last_total - len(results)
                    if remaining > 0 and remaining <= 70:
                        start_补 = len(results) + 1
                        for smaller_display in [remaining, 50, 30, 20, 10]:
                            if smaller_display > remaining + 10:
                                continue
                            retry_p = _json.loads(_json.dumps(gql_payload))
                            retry_p["variables"]["input"]["start"] = start_补
                            retry_p["variables"]["input"]["display"] = smaller_display
                            retry_b64 = _b64.b64encode(_json.dumps([retry_p]).encode("utf-8")).decode("ascii")
                            retry_js = f"""
                            try {{
                                const resp = await fetch('https://pcmap-api.place.naver.com/graphql', {{
                                    method:'POST', headers:{{'Content-Type':'application/json'}},
                                    body: atob('{retry_b64}'), credentials:'include'
                                }});
                                const text = await resp.text();
                                return resp.status + '|||' + text;
                            }} catch(e) {{ return 'ERROR:'+e.message; }}
                            """
                            try:
                                retry_result = driver.execute_script(retry_js)
                                if retry_result and "|||" in retry_result:
                                    r_status, r_body = retry_result.split("|||", 1)
                                    if r_status == "200":
                                        r_data = _json.loads(r_body)
                                        if isinstance(r_data, list):
                                            r_data = r_data[0]
                                        r_items = r_data.get("data", {}).get("businesses", {}).get("items", [])
                                        if r_items:
                                            extra_pg = len(results) // 70 + 1
                                            for it in r_items:
                                                pid = str(it.get("id", ""))
                                                name = it.get("name", "")
                                                if pid and pid not in {r["id"] for r in results}:
                                                    results.append({"id": pid, "name": name, "rank": len(results) + 1, "page": extra_pg})
                                            self.callback("log", f"  🔄 보충: +{len(r_items)}건 (display={smaller_display}, 총 {len(results)}건)")
                                            break
                            except Exception:
                                pass

            except Exception as e:
                self.callback("log", f"  ⚠️ Selenium 오류: {e}")
            finally:
                try:
                    driver.quit()
                except Exception:
                    pass

            return results if results else None

        # 모든 프록시 재시도 소진
        return None

    # ═══════════════════════════════════════════
    # Playwright 브라우저 크롤링 (map.naver.com)
    # ═══════════════════════════════════════════

    _pw_browser = None  # 클래스 레벨 브라우저 재사용

    def _init_playwright(self):
        """Playwright 브라우저 초기화 (최초 1회, 핑거프린트 우회)"""
        if CrawlerEngine._pw_browser:
            return True
        try:
            from playwright.sync_api import sync_playwright
            self._pw = sync_playwright().start()
            # 랜덤 UA 선택
            fp = self._get_fingerprint()
            CrawlerEngine._pw_browser = self._pw.chromium.launch(
                headless=True,
                args=[
                    "--no-sandbox",
                    "--disable-blink-features=AutomationControlled",
                    f"--user-agent={fp['ua']}",
                    "--disable-dev-shm-usage",
                ],
            )
            # webdriver 플래그 숨기기용 init script 저장
            CrawlerEngine._pw_init_script = """
                Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
                Object.defineProperty(navigator, 'languages', {get: () => ['ko-KR', 'ko', 'en-US', 'en']});
                Object.defineProperty(navigator, 'platform', {get: () => 'Win32'});
            """
            return True
        except Exception as e:
            logger.debug(f"Playwright init failed: {e}")
            self.callback("log", f"⚠️ Playwright 초기화 실패: {e}")
            return False

    def _close_playwright(self):
        """Playwright 정리"""
        if CrawlerEngine._pw_browser:
            try:
                CrawlerEngine._pw_browser.close()
            except Exception:
                pass
            CrawlerEngine._pw_browser = None
        if hasattr(self, "_pw") and self._pw:
            try:
                self._pw.stop()
            except Exception:
                pass

    def _playwright_search(self, keyword):
        """Playwright로 map.naver.com에서 업체 목록 수집 (1페이지, 최대 30건).
        반환: (items list, total int) 또는 ([], 0) on failure.
        """
        if not self._init_playwright():
            return [], 0

        try:
            page = CrawlerEngine._pw_browser.new_page()
            # 핑거프린트: 랜덤 viewport + webdriver 숨기기
            vw = random.choice([1280, 1366, 1440, 1536, 1920])
            vh = random.choice([720, 768, 800, 900, 1080])
            page.set_viewport_size({"width": vw, "height": vh})
            if hasattr(CrawlerEngine, "_pw_init_script"):
                page.add_init_script(CrawlerEngine._pw_init_script)

            # API 응답 가로채기
            captured = {"items": [], "total": 0}
            def on_response(response):
                if "allSearch" in response.url:
                    try:
                        data = response.json()
                        place = data.get("result", {}).get("place")
                        if place and place.get("list"):
                            captured["total"] = place.get("totalCount", 0)
                            for it in place["list"]:
                                cat = it.get("category", "")
                                if isinstance(cat, list):
                                    cat = cat[0] if cat else ""
                                captured["items"].append({
                                    "id": str(it.get("id", "")),
                                    "name": it.get("name", ""),
                                    "category": cat,
                                    "phone": it.get("tel", ""),
                                    "virtualPhone": it.get("virtualTel", "") or it.get("tel", ""),
                                    "address": it.get("address", ""),
                                    "roadAddress": it.get("roadAddress", "") or it.get("address", ""),
                                    "reviewCount": it.get("reviewCount", 0),
                                    "blogCafeReviewCount": it.get("blogReviewCount", 0),
                                })
                    except Exception:
                        pass

            page.on("response", on_response)

            url = f"https://map.naver.com/p/search/{quote(keyword)}"
            page.goto(url, wait_until="domcontentloaded", timeout=15000)
            page.wait_for_timeout(3000)

            page.close()
            return captured["items"], captured["total"]

        except Exception as e:
            logger.debug(f"Playwright search error: {e}")
            self.callback("log", f"⚠️ Playwright 오류: {e}")
            return [], 0

    # ═══════════════════════════════════════════
    # 네이버 지역 검색 API (구/동 분할)
    # ═══════════════════════════════════════════

    # 주요 도시 구/군 목록 (자동 세분화용)
    _CITY_DISTRICTS = {
        "서울": ["강남구","강동구","강북구","강서구","관악구","광진구","구로구","금천구",
                 "노원구","도봉구","동대문구","동작구","마포구","서대문구","서초구",
                 "성동구","성북구","송파구","양천구","영등포구","용산구","은평구",
                 "종로구","중구","중랑구"],
        "부산": ["강서구","금정구","기장군","남구","동구","동래구","부산진구","북구",
                 "사상구","사하구","서구","수영구","연제구","영도구","중구","해운대구"],
        "대구": ["남구","달서구","달성군","동구","북구","서구","수성구","중구"],
        "인천": ["강화군","계양구","남동구","동구","미추홀구","부평구","서구","연수구","옹진군","중구"],
        "광주": ["광산구","남구","동구","북구","서구"],
        "대전": ["대덕구","동구","서구","유성구","중구"],
        "울산": ["남구","동구","북구","울주군","중구"],
        "세종": ["세종시"],
        "경기": ["가평군","고양시","과천시","광명시","광주시","구리시","군포시","김포시",
                 "남양주시","동두천시","부천시","성남시","수원시","시흥시","안산시","안성시",
                 "안양시","양주시","양평군","여주시","연천군","오산시","용인시","의왕시",
                 "의정부시","이천시","파주시","평택시","포천시","하남시","화성시"],
        "강원": ["강릉시","동해시","삼척시","속초시","원주시","춘천시","태백시","홍천군","횡성군","평창군","정선군","철원군","화천군","양구군","인제군","고성군","양양군","영월군"],
        "충북": ["청주시","충주시","제천시","보은군","옥천군","영동군","증평군","진천군","괴산군","음성군","단양군"],
        "충남": ["천안시","공주시","보령시","아산시","서산시","논산시","계룡시","당진시","금산군","부여군","서천군","청양군","홍성군","예산군","태안군"],
        "전북": ["전주시","군산시","익산시","정읍시","남원시","김제시","완주군","진안군","무주군","장수군","임실군","순창군","고창군","부안군"],
        "전남": ["목포시","여수시","순천시","나주시","광양시","담양군","곡성군","구례군","고흥군","보성군","화순군","장흥군","강진군","해남군","영암군","무안군","함평군","영광군","장성군","완도군","진도군","신안군"],
        "경북": ["포항시","경주시","김천시","안동시","구미시","영주시","영천시","상주시","문경시","경산시","군위군","의성군","청송군","영양군","영덕군","청도군","고령군","성주군","칠곡군","예천군","봉화군","울진군","울릉군"],
        "경남": ["창원시","진주시","통영시","사천시","김해시","밀양시","거제시","양산시","의령군","함안군","창녕군","고성군","남해군","하동군","산청군","함양군","거창군","합천군"],
        "제주": ["제주시","서귀포시"],
    }

    # 지역명 사전 (붙여쓰기 파싱용)
    _KNOWN_LOCATIONS = [
        # 광역시/도
        "서울", "부산", "대구", "인천", "광주", "대전", "울산", "세종",
        "경기", "강원", "충북", "충남", "전북", "전남", "경북", "경남", "제주",
        # 주요 시/구/동
        "수원", "성남", "안양", "안산", "용인", "고양", "과천", "광명", "군포",
        "의왕", "하남", "이천", "평택", "시흥", "파주", "김포", "화성", "양주",
        "포천", "구리", "남양주", "오산", "안성", "동두천", "의정부", "양평",
        "여주", "가평", "연천",
        "청주", "충주", "제천", "천안", "아산", "서산", "논산", "공주", "보령", "당진",
        "전주", "군산", "익산", "정읍", "남원", "김제",
        "목포", "여수", "순천", "나주", "광양",
        "포항", "경주", "김천", "안동", "구미", "영주", "영천", "상주", "문경",
        "창원", "진주", "통영", "사천", "김해", "밀양", "거제", "양산",
        "춘천", "원주", "강릉", "동해", "태백", "속초", "삼척",
        "제주", "서귀포",
        # 서울 구
        "강남", "서초", "송파", "강동", "마포", "용산", "종로", "중구", "성북",
        "강북", "도봉", "노원", "은평", "서대문", "광진", "동대문", "중랑",
        "성동", "금천", "관악", "동작", "영등포", "양천", "강서", "구로",
        # 부산 구/동
        "해운대", "서면", "부산진", "동래", "남포", "사하", "북구", "사상",
        "연제", "수영", "금정", "기장",
        # 대구/인천/광주/대전
        "수성", "달서", "남동", "부평", "계양", "연수", "미추홀",
        "북구", "광산", "서구", "유성", "대덕",
    ]

    def _parse_keyword_location(self, keyword):
        """키워드에서 지역명과 업종 분리.
        '울산 피부과' → ('울산', '피부과')
        '안양피부과' → ('안양', '피부과')  # 붙여쓰기도 지원
        """
        parts = keyword.strip().split()
        if len(parts) >= 2:
            return parts[0], " ".join(parts[1:])
        # 붙여쓰기: 긴 지역명부터 매칭 시도
        kw = keyword.strip()
        for loc in sorted(self._KNOWN_LOCATIONS, key=len, reverse=True):
            if kw.startswith(loc) and len(kw) > len(loc):
                return loc, kw[len(loc):]
        return keyword, ""

    # 동/지역명 → 도시 매핑 (구/군 이외의 지역명)
    _LOCATION_CITY_MAP = {
        "서면": "부산", "남포": "부산", "해운대": "부산", "광안리": "부산",
        "강남": "서울", "홍대": "서울", "신촌": "서울", "이태원": "서울",
        "명동": "서울", "잠실": "서울", "건대": "서울", "신림": "서울",
        "동성로": "대구", "범어": "대구", "수성": "대구",
        "충장로": "광주", "상무": "광주",
        "둔산": "대전", "유성": "대전",
        "성산": "제주", "서귀포": "제주",
        "안양": "경기", "수원": "경기", "성남": "경기", "안산": "경기",
        "용인": "경기", "고양": "경기", "과천": "경기", "광명": "경기",
        "군포": "경기", "의왕": "경기", "하남": "경기", "이천": "경기",
        "평택": "경기", "시흥": "경기", "파주": "경기", "김포": "경기",
        "화성": "경기", "양주": "경기", "포천": "경기", "구리": "경기",
        "남양주": "경기", "오산": "경기", "의정부": "경기",
        "춘천": "강원", "원주": "강원", "강릉": "강원", "속초": "강원",
        "청주": "충북", "충주": "충북", "제천": "충북",
        "천안": "충남", "아산": "충남", "서산": "충남",
        "전주": "전북", "군산": "전북", "익산": "전북",
        "목포": "전남", "여수": "전남", "순천": "전남",
        "포항": "경북", "경주": "경북", "구미": "경북", "안동": "경북",
        "창원": "경남", "진주": "경남", "김해": "경남", "거제": "경남",
    }

    def _find_city_for_location(self, location):
        """지역명이 속한 도시 찾기. '울산' → '울산', '서면' → '부산'"""
        # 직접 매칭
        if location in self._CITY_DISTRICTS:
            return location
        # 명시적 매핑
        if location in self._LOCATION_CITY_MAP:
            return self._LOCATION_CITY_MAP[location]
        # 구/군 이름으로 역매핑
        for city, districts in self._CITY_DISTRICTS.items():
            for d in districts:
                if location in d or d.startswith(location):
                    return city
        return None

    def _api_local_search(self, query, display=5, start=1):
        """네이버 지역 검색 API 호출. 반환: items list"""
        url = "https://openapi.naver.com/v1/search/local.json"
        params = {"query": query, "display": display, "start": start}

        # API 키 — 내장 키 사용
        cid = getattr(self, "naver_client_id", "yoBUbNSW9MGSPH36zaHN")
        csecret = getattr(self, "naver_client_secret", "wt0HPPOVKA")
        headers = {
            "X-Naver-Client-Id": cid,
            "X-Naver-Client-Secret": csecret,
        }
        try:
            resp = self.session.get(url, params=params, headers=headers, timeout=self.timeout)
            if resp.status_code == 200:
                return resp.json().get("items", [])
        except Exception:
            pass
        return []

    def _api_name_to_place_id(self, name, addr_hint=""):
        """업체명+주소로 네이버 검색 → Place ID 추출"""
        q = name
        if addr_hint:
            q += " " + addr_hint[:10]
        search_url = f"https://search.naver.com/search.naver?query={quote(q)}"
        fp = self._get_fingerprint()
        headers = {"User-Agent": fp["ua"]}
        try:
            resp = self.session.get(search_url, headers=headers, timeout=self.timeout)
            resp.encoding = "utf-8"
            pids = re.findall(r'/place/(\d{5,})', resp.text)
            if pids:
                return pids[0]
        except Exception:
            pass
        return ""

    def _collect_all_places_api(self, keyword):
        """네이버 지역 검색 API + 구/군 분할로 최대한 많은 Place ID 수집.
        반환: items list (id, name, category, phone, address 등)
        """
        location, biz_type = self._parse_keyword_location(keyword)
        if not biz_type:
            # 업종이 없으면 분할 불가 → 빈 결과
            return [], 0

        city = self._find_city_for_location(location)
        seen_names = set()
        all_items = []

        # 1단계: 원본 키워드 검색
        base_results = self._api_local_search(keyword)
        for it in base_results:
            name = re.sub(r'</?b>', '', it.get("title", ""))
            if name not in seen_names:
                seen_names.add(name)
                all_items.append(it)
        self.callback("log", f"📍 API 기본 검색: {len(all_items)}건")

        # 2단계: 구/군 분할 (같은 도시일 때만)
        if city and city in self._CITY_DISTRICTS:
            districts = self._CITY_DISTRICTS[city]
            for di, dist in enumerate(districts):
                if not self.running:
                    break
                q = f"{city} {dist} {biz_type}"
                items = self._api_local_search(q)
                new_count = 0
                for it in items:
                    name = re.sub(r'</?b>', '', it.get("title", ""))
                    if name not in seen_names:
                        seen_names.add(name)
                        all_items.append(it)
                        new_count += 1
                self.callback("log", f"  [{di+1}/{len(districts)}] {dist}: +{new_count}건 (총 {len(all_items)}건)")
                self._random_delay(0.3, 0.7)  # API는 짧은 딜레이
        elif location:
            # 도시 매칭 안 되면 원래 지역 + "동" 검색 1회만
            pass

        self.callback("log", f"📍 API 총 수집: {len(all_items)}건 (중복 제거)")

        # 3단계: Place ID 매칭 (업체명 → 네이버 검색)
        result_items = []
        pid_found = 0
        for idx, it in enumerate(all_items):
            if not self.running:
                break
            name = re.sub(r'</?b>', '', it.get("title", ""))
            addr = it.get("roadAddress", "") or it.get("address", "")
            cat = it.get("category", "")
            tel = it.get("telephone", "")

            pid = self._api_name_to_place_id(name, addr)
            item = {
                "id": pid or "",
                "name": name,
                "category": cat,
                "phone": tel,
                "virtualPhone": tel,
                "address": it.get("address", ""),
                "roadAddress": addr,
                "reviewCount": 0,
                "blogCafeReviewCount": 0,
            }
            result_items.append(item)
            if pid:
                pid_found += 1
                self.callback("log", f"  [{idx+1}/{len(all_items)}] {name}: PID={pid}")
            else:
                self.callback("log", f"  [{idx+1}/{len(all_items)}] {name}: PID 미발견 (name+addr로 포함)")
            self._random_delay(0.5, 1.0)

        self.callback("log", f"📍 PID 매칭: {pid_found}/{len(all_items)}건")
        return result_items, len(result_items)

    def _map_api_search(self, keyword, page_num=1):
        """Map API 직접 HTTP 호출. 반환: (items, total) 또는 ([], 0)"""
        from urllib.parse import quote as _quote
        url = "https://map.naver.com/p/api/search/allSearch"
        fp = self._get_fingerprint()
        params = {
            "query": keyword, "type": "all",
            "searchCoord": "127.0;37.5", "page": page_num, "displayCount": 30,
        }
        headers = {
            "User-Agent": fp["ua"],
            "Referer": "https://map.naver.com/p/search/" + _quote(keyword),
        }
        try:
            resp = self.session.get(url, params=params, headers=headers, timeout=self.timeout)
            if resp.status_code == 200:
                resp.encoding = "utf-8"
                data = resp.json()
                place = data.get("result", {}).get("place")
                if not place or data.get("result", {}).get("ncaptcha"):
                    return [], 0
                if place and place.get("list"):
                    raw_items = place["list"]
                    items = []
                    for it in raw_items:
                        cat = it.get("category", "")
                        if isinstance(cat, list):
                            cat = cat[0] if cat else ""
                        items.append({
                            "id": str(it.get("id", "")),
                            "name": it.get("name", ""),
                            "category": cat,
                            "phone": it.get("tel", ""),
                            "virtualPhone": it.get("virtualTel", "") or it.get("tel", ""),
                            "address": it.get("address", ""),
                            "roadAddress": it.get("roadAddress", "") or it.get("address", ""),
                            "reviewCount": it.get("reviewCount", 0),
                            "blogCafeReviewCount": it.get("blogReviewCount", 0),
                        })
                    return items, place.get("totalCount", 0)
        except Exception as e:
            logger.debug(f"Map API error: {e}")
        return [], 0

    # JSON 기본정보 캐시 (PID → {name, tel, addr, ...})
    _place_info_cache = {}

    def _mobile_map_search_single(self, keyword):
        """m.map.naver.com/search2에서 Place ID 수집 (75건/페이지).
        JSON에서 기본정보도 캐싱 (place HTML fetch 절감).
        반환: (pids set, total int)
        """
        url = f"https://m.map.naver.com/search2/search.naver?query={quote(keyword)}&sm=hty&style=v5"
        headers = {
            "User-Agent": "Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Mobile/15E148 Safari/604.1",
            "Referer": "https://m.map.naver.com/",
        }
        try:
            resp = self.session.get(url, headers=headers, timeout=self.timeout)
            resp.encoding = "utf-8"
            if resp.status_code != 200 or len(resp.text) < 1000:
                return set(), 0

            # JSON items 추출 (기본정보 캐싱)
            html = resp.text
            m = re.search(r'"totalCount":(\d+),"items":\[', html)
            total = 0
            if m:
                total = int(m.group(1))
                try:
                    start_i = html.index("[", m.start())
                    depth, end_i = 0, start_i
                    for i in range(start_i, min(start_i + 500000, len(html))):
                        if html[i] == "[": depth += 1
                        elif html[i] == "]": depth -= 1
                        if depth == 0:
                            end_i = i + 1
                            break
                    items = json.loads(html[start_i:end_i])
                    for it in items:
                        pid = str(it.get("id", ""))
                        if pid:
                            CrawlerEngine._place_info_cache[pid] = {
                                "name": it.get("name", ""),
                                "category": it.get("category", ""),
                                "tel": it.get("tel", ""),
                                "virtualTel": it.get("virtualTel", ""),
                                "address": it.get("address", ""),
                                "roadAddress": it.get("roadAddress", ""),
                            }
                except Exception:
                    pass

            pids = list(dict.fromkeys(re.findall(r'/place/(\d{5,})', html)))
            if not total:
                tm = re.search(r'totalCount["\s:]+(\d+)', html)
                if tm:
                    total = int(tm.group(1))
            return set(pids), total
        except Exception as e:
            logger.debug(f"mobile map search error: {e}")
        return set(), 0

    # ── 전국 행정구역 데이터 (오프라인, API 불필요) ──
    _DISTRICTS_DATA = None  # 지연 로딩

    @classmethod
    def _load_districts_data(cls):
        """districts_data.json 로드 (최초 1회)"""
        if cls._DISTRICTS_DATA is not None:
            return cls._DISTRICTS_DATA
        json_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "districts_data.json")
        try:
            with open(json_path, "r", encoding="utf-8") as f:
                cls._DISTRICTS_DATA = json.load(f)
                logger.debug(f"districts_data.json loaded: {len(cls._DISTRICTS_DATA)} cities")
        except Exception as e:
            logger.warning(f"districts_data.json 로드 실패: {e}")
            cls._DISTRICTS_DATA = {}
        return cls._DISTRICTS_DATA

    def _fetch_sub_districts(self, city_name):
        """전국 행정구역 데이터에서 구/군 + 동 목록 반환 (오프라인).
        반환: {구이름: [동1, 동2, ...], ...}
        """
        data = self._load_districts_data()
        return data.get(city_name, {})

    # 시 → 하위 구 매핑 (하위 구가 있는 시)
    _CITY_TO_GU = {
        "안양": ["만안구", "동안구"],
        "수원": ["장안구", "권선구", "팔달구", "영통구"],
        "성남": ["수정구", "중원구", "분당구"],
        "안산": ["상록구", "단원구"],
        "용인": ["처인구", "기흥구", "수지구"],
        "고양": ["덕양구", "일산동구", "일산서구"],
        "천안": ["동남구", "서북구"],
        "청주": ["상당구", "서원구", "흥덕구", "청원구"],
        "전주": ["완산구", "덕진구"],
        "창원": ["의창구", "성산구", "마산합포구", "마산회원구", "진해구"],
        "포항": ["남구", "북구"],
    }

    def _fetch_sub_districts_for_area(self, city_name, area_name):
        """특정 시/구 영역의 하위 구/동만 가져오기 (오프라인).
        예: city=경기, area=안양 → 만안구(안양동,석수동,박달동), 동안구(비산동,관양동,평촌동,호계동)
        """
        data = self._load_districts_data()
        city_data = data.get(city_name, {})
        if not city_data:
            return {}

        filtered = {}

        # 1차: 시→하위구 매핑 사전으로 정확 매칭
        sub_gus = self._CITY_TO_GU.get(area_name, [])
        if sub_gus:
            for gu_name in sub_gus:
                if gu_name in city_data and city_data[gu_name]:
                    filtered[gu_name] = city_data[gu_name]
            if filtered:
                return filtered

        # 2차: area_name이 구/군과 직접 매칭
        for dist_name, dong_list in city_data.items():
            if area_name in dist_name or dist_name.startswith(area_name):
                if dong_list:
                    filtered[dist_name] = dong_list

        return filtered

    def _mobile_map_search(self, keyword):
        """m.map.naver.com에서 Place ID 대량 수집.
        1차: 원본 키워드 (75건)
        2차: 구/군 분할 (75건씩)
        3차: 동 단위 분할 (75건씩) — 98%+ 커버
        반환: (items list, total int)
        """
        all_pids = set()

        # 1차: 원본 키워드
        pids, total = self._mobile_map_search_single(keyword)
        all_pids.update(pids)
        self.callback("log", f"  원본 검색: {len(pids)}건 (전체 {total}건)")

        # 2차+3차: 구/군 + 동 단위 분할
        if total > len(pids) and self.running:
            location, biz_type = self._parse_keyword_location(keyword)
            city = self._find_city_for_location(location)
            if city and biz_type:
                # location이 시/구 단위면 (안양, 수원 등) → 해당 지역만 분할
                # city가 광역시/도면 (경기, 충남 등) location 기준 분할
                search_area = location if location != city else city
                self.callback("log", f"  📍 {search_area} 행정구역 분할 수집 중...")

                if location != city:
                    # "안양" → 안양시 하위 구/동만 가져오기
                    dist_dongs = self._fetch_sub_districts_for_area(city, location)
                    if not dist_dongs:
                        # 하위 구/동 없으면 city 전체로 폴백 (서면→부산 전체)
                        dist_dongs = self._fetch_sub_districts(city)
                        search_area = city
                        self.callback("log", f"  📍 {location} 하위 없음 → {city} 전체 분할")
                else:
                    dist_dongs = self._fetch_sub_districts(city)
                if not dist_dongs and city in self._CITY_DISTRICTS:
                    # JSON 없으면 하드코딩 폴백 (구 단위만)
                    dist_dongs = {d: [] for d in self._CITY_DISTRICTS[city]}

                for dist_name, dong_list in dist_dongs.items():
                    if not self.running:
                        break
                    # 구 단위 검색 (search_area 기준)
                    sub_kw = f"{search_area} {dist_name} {biz_type}"
                    sub_pids, _ = self._mobile_map_search_single(sub_kw)
                    new = sub_pids - all_pids
                    if new:
                        all_pids.update(new)
                        self.callback("log", f"  {dist_name}: +{len(new)}건 (누적 {len(all_pids)})")
                    self._random_delay(0.3, 0.6)

                    # 동 단위 검색 (아직 total 미달이면)
                    if len(all_pids) < total and dong_list:
                        for dong in dong_list:
                            if not self.running:
                                break
                            dong_kw = f"{search_area} {dist_name} {dong} {biz_type}"
                            dong_pids, _ = self._mobile_map_search_single(dong_kw)
                            dong_new = dong_pids - all_pids
                            if dong_new:
                                all_pids.update(dong_new)
                                self.callback("log", f"    {dong}: +{len(dong_new)}건 (누적 {len(all_pids)})")
                            self._random_delay(0.3, 0.6)

                    # 이미 total 도달하면 조기 종료
                    if len(all_pids) >= total:
                        self.callback("log", f"  ✅ 전체 {total}건 도달!")
                        break

        self.callback("log", f"  📊 최종: {len(all_pids)}/{total}건 ({len(all_pids)*100//max(total,1)}%)")

        # items 생성
        items = []
        for pid in all_pids:
            items.append({
                "id": pid, "name": "", "category": "", "phone": "",
                "virtualPhone": "", "address": "", "roadAddress": "",
                "reviewCount": 0, "blogCafeReviewCount": 0,
            })
        return items, total or len(items)

    def _html_search_fallback(self, keyword, start):
        """HTML 파싱 폴백: 네이버 검색에서 place ID 추출 (스트리밍).
        반환: (items list of dicts, total int estimate).
        """
        page_start = start
        url = (
            f"https://m.search.naver.com/search.naver"
            f"?query={quote(keyword)}&sm=tab_jum&where=nexearch&start={page_start}"
        )
        # 스트리밍: Place ID가 포함된 부분만 받고 끊기 (속도 ~2x)
        fp = self._get_fingerprint()
        headers = {
            "User-Agent": fp["ua"],
            "Referer": "https://m.search.naver.com/",
            "Accept-Encoding": "gzip, deflate",
        }
        html = ""
        try:
            resp = self.session.get(url, headers=headers, timeout=self.timeout, stream=True)
            resp.encoding = "utf-8"
            raw_bytes = b""
            ids_found = set()
            for chunk in resp.iter_content(chunk_size=16384):
                if chunk:
                    raw_bytes += chunk
                    # Place ID 실시간 탐지 (바이트에서 ASCII 패턴)
                    for m in re.finditer(rb'/place/(\d{5,})', raw_bytes[-20000:]):
                        ids_found.add(m.group(1).decode())
                    for m in re.finditer(rb'"id"\s*:\s*"(\d{5,})"', raw_bytes[-20000:]):
                        ids_found.add(m.group(1).decode())
                    # 충분한 ID 수집 → 조기 종료 (ID는 ~220KB 지점)
                    if len(ids_found) >= 10:
                        break
            html = raw_bytes.decode("utf-8", errors="replace")
            resp.close()
        except Exception as e:
            logger.debug(f"stream fetch error: {e}")
        if not html:
            return [], 0
        # place ID 추출 — /place/ 우선 (가장 정확), 보조 패턴 추가
        ids = re.findall(r'/place/(\d{5,})', html)
        ids += re.findall(r'"id"\s*:\s*"(\d{5,})"', html)
        ids += re.findall(r'placeId[=:]["]*(\d{5,})', html)
        # 중복 제거, 순서 유지
        unique_ids = list(dict.fromkeys(ids))
        items = []
        for pid in unique_ids:
            items.append({"id": pid, "name": "", "category": "", "phone": "",
                          "virtualPhone": "", "address": "", "roadAddress": "",
                          "reviewCount": 0, "blogCafeReviewCount": 0})
        # total 추정 (HTML에서 정확한 total 없음 → 결과 있으면 999로 설정)
        total = 999 if items else 0
        return items, total

    def _stat_inc(self, key, n=1):
        """thread-safe stats 증가"""
        with self._stats_lock:
            self.stats[key] = self.stats.get(key, 0) + n

    def _detail_crawl_one(self, r, total, idx):
        """1건 상세 크롤링 — GraphQL 우선, place HTML 폴백 (thread-safe)"""
        name = r.get("상호명", "")
        pid = r.get("고유번호", "")
        log_items = []

        # PID 없으면 검색으로 재시도
        if not pid and name:
            addr = r.get("업체주소", "")
            pid = self._api_name_to_place_id(name, addr)
            if pid:
                r["고유번호"] = pid
                r["플레이스URL"] = f"https://m.place.naver.com/place/{pid}"

        # ═══ 1차: GraphQL API (전화/홈페이지/리뷰 정확도 최고) ═══
        gql = {}
        if pid:
            gql = self._fetch_place_graphql(pid)

        # ═══ 2차: place HTML 폴백 (GraphQL 실패 시) ═══
        parsed_place = {}
        need_hp = not (r.get("홈페이지URL") or "").strip() or (r.get("홈페이지URL") or "").strip() == "X"
        need_nid = not (r.get("네이버아이디") or "").strip()
        if pid and not gql and (need_hp or need_nid):
            place_html = self._fetch_place_html(pid, name)
            parsed_place = self._parse_place_html(place_html)

        # ═══ 안심번호 ═══
        current_phone = (r.get("안심번호") or "").strip()
        if current_phone:
            log_items.append(f"tel:{current_phone}")
            self._stat_inc("phone")
        elif gql.get("phone"):
            r["안심번호"] = gql["phone"]
            log_items.append(f"tel:{gql['phone']}")
            self._stat_inc("phone")
        elif parsed_place.get("phone"):
            r["안심번호"] = parsed_place["phone"]
            log_items.append(f"tel:{parsed_place['phone']}")
            self._stat_inc("phone")

        # ═══ 방문자/블로그 리뷰 ═══
        cur_vr = str(r.get("방문자리뷰수") or "0").strip()
        cur_br = str(r.get("블로그리뷰수") or "0").strip()
        if gql.get("visitor_review") and gql["visitor_review"] != "0":
            r["방문자리뷰수"] = gql["visitor_review"]
        elif cur_vr == "0" and parsed_place.get("visitor_review"):
            r["방문자리뷰수"] = parsed_place["visitor_review"]

        # 블로그 리뷰: GraphQL fsasReviews가 0이면 summary API로 보충
        if gql.get("blog_review") and gql["blog_review"] != "0":
            r["블로그리뷰수"] = gql["blog_review"]
        elif cur_br == "0":
            if pid:
                summary = self._fetch_place_summary(pid)
                if summary.get("blog_review") and summary["blog_review"] != "0":
                    r["블로그리뷰수"] = summary["blog_review"]
            if (r.get("블로그리뷰수") or "0") == "0" and parsed_place.get("blog_review"):
                r["블로그리뷰수"] = parsed_place["blog_review"]

        # ═══ 홈페이지 ═══
        if need_hp:
            if gql.get("homepage"):
                r["홈페이지URL"] = gql["homepage"]
                self._stat_inc("hp")
                log_items.append(f"hp:{gql['homepage']}")
            elif parsed_place.get("homepage"):
                r["홈페이지URL"] = parsed_place["homepage"]
                self._stat_inc("hp")
                log_items.append(f"hp:{parsed_place['homepage']}")

        # ═══ 네이버 톡톡 URL ═══
        if not (r.get("톡톡URL") or "").strip():
            tt = gql.get("talktalk") or parsed_place.get("talktalk") or ""
            if tt:
                r["톡톡URL"] = tt
                r["톡톡"] = "O"
                self._stat_inc("talktalk")
                log_items.append("talk:O")

        # ═══ 카카오톡 (O/X) ═══
        if not (r.get("카카오톡") or "").strip():
            kk = gql.get("kakao") or parsed_place.get("kakao") or ""
            if kk:
                r["카카오톡"] = "O"
                self._stat_inc("kakao")
                log_items.append("kakao:O")

        # ═══ 인스타그램 (O/X) ═══
        if not (r.get("인스타그램") or "").strip():
            ig = gql.get("instagram") or parsed_place.get("instagram") or ""
            if ig:
                r["인스타그램"] = "O"
                self._stat_inc("instagram")
                log_items.append("ig:O")

        # ═══ 신규업체 ═══
        if not (r.get("신규업체") or "").strip():
            if gql.get("is_new") == "O" or parsed_place.get("is_new") == "O":
                r["신규업체"] = "O"
                self._stat_inc("new_biz")
                log_items.append("new:O")

        # ═══ 이메일 + 홈페이지 분석 (반응형/전화버튼/웹빌더/카카오/인스타 — HTML 1회 재활용) ═══
        need_email = not (r.get("업체이메일") or "").strip()
        need_resp = not (r.get("홈페이지반응형") or "").strip()
        need_phone_btn = not (r.get("전화버튼없음") or "").strip()
        need_webbuilder = not (r.get("웹빌더") or "").strip()
        need_kakao_hp = not (r.get("카카오톡") or "").strip()
        need_insta_hp = not (r.get("인스타그램") or "").strip()
        hp_url = (r.get("홈페이지URL") or "").strip()
        if hp_url and hp_url != "X" and hp_url.startswith("http") and (
            need_email or need_resp or need_phone_btn or need_webbuilder or need_kakao_hp or need_insta_hp
        ):
            html = self._fetch(hp_url, referer=f"https://www.google.com/search?q={quote(name)}", timeout=3)
            if need_email:
                emails = self.extract_emails(html)
                if emails:
                    r["업체이메일"] = emails[0]
                    self._stat_inc("email")
                    log_items.append(f"email:{emails[0]}")
            if need_resp:
                resp_flag = self.detect_responsive(html) if html else "X"
                r["홈페이지반응형"] = resp_flag
                if resp_flag == "O":
                    self._stat_inc("responsive")
                log_items.append(f"resp:{resp_flag}")
            if need_phone_btn:
                has_btn = self.detect_phone_button(html) if html else False
                flag = "X" if has_btn else "O"
                r["전화버튼없음"] = flag
                if flag == "O":
                    self._stat_inc("no_phone_btn")
                log_items.append(f"nophonebtn:{flag}")
            if need_webbuilder:
                wb_flag = self.detect_webbuilder(html, url=hp_url)
                r["웹빌더"] = wb_flag
                if wb_flag == "O":
                    self._stat_inc("webbuilder")
                log_items.append(f"wb:{wb_flag}")
            if need_kakao_hp and html:
                kk2 = self.extract_kakao_url(html)
                if kk2:
                    r["카카오톡"] = "O"
                    self._stat_inc("kakao")
                    log_items.append("kakao:O(hp)")
            if need_insta_hp and html:
                ig2 = self.extract_instagram_url(html)
                if ig2:
                    r["인스타그램"] = "O"
                    self._stat_inc("instagram")
                    log_items.append("ig:O(hp)")
        # place HTML에서도 이메일 시도
        if not (r.get("업체이메일") or "").strip() and parsed_place.get("email"):
            r["업체이메일"] = parsed_place["email"]
            self._stat_inc("email")
            log_items.append(f"email:{parsed_place['email']}(place)")

        # ═══ 네이버아이디 ═══
        if need_nid and name:
            # GraphQL homepage가 블로그면 ID 추출
            nid = None
            hp_for_nid = gql.get("homepage", "") or ""
            if "blog.naver.com" in hp_for_nid:
                import re
                # blogId= 파라미터 우선 (profile/intro.naver?blogId=xxx)
                m = re.search(r'blogId=([a-zA-Z0-9_.-]+)', hp_for_nid)
                if not m:
                    m = re.search(r'blog\.naver\.com/([a-zA-Z0-9_.-]+)', hp_for_nid)
                if m and m.group(1) not in ("profile", "intro", "PostList", "PostView"):
                    nid = m.group(1) + "@naver.com"
            # etc 링크에서도 블로그 찾기
            if not nid and gql.get("etc_links"):
                for link in gql["etc_links"]:
                    if "blog.naver.com" in (link.get("url") or ""):
                        import re
                        m = re.search(r'blogId=([a-zA-Z0-9_.-]+)', link["url"])
                        if not m:
                            m = re.search(r'blog\.naver\.com/([a-zA-Z0-9_.-]+)', link["url"])
                        if m and m.group(1) not in ("profile", "intro", "PostList", "PostView"):
                            nid = m.group(1) + "@naver.com"
                            break
            # 기존 방식 폴백
            if not nid:
                nid = self._find_naver_id(name, parsed_place or gql)
            if nid:
                r["네이버아이디"] = nid
                self.stats["naver_id"] += 1
                log_items.append(f"nid:{nid}")
        elif not need_nid:
            naver_id = (r.get("네이버아이디") or "").replace("@naver.com", "").strip()
            if naver_id and "@" not in (r.get("네이버아이디") or ""):
                r["네이버아이디"] = naver_id + "@naver.com"
                self.stats["naver_id"] += 1

        # 미감지 항목은 X로 확정 (톡톡URL은 빈값 유지 — 복사 편의)
        if not (r.get("톡톡") or "").strip():
            r["톡톡"] = "X"
        if not (r.get("카카오톡") or "").strip():
            r["카카오톡"] = "X"
        if not (r.get("인스타그램") or "").strip():
            r["인스타그램"] = "X"
        if not (r.get("신규업체") or "").strip():
            r["신규업체"] = "X"
        if not (r.get("홈페이지반응형") or "").strip():
            r["홈페이지반응형"] = "X"
        if not (r.get("전화버튼없음") or "").strip():
            # 홈페이지 자체가 없거나 검사 실패 → 영업 타겟 아님 (X)
            r["전화버튼없음"] = "X"
        if not (r.get("웹빌더") or "").strip():
            # 홈페이지 없음/검사 실패 → 웹빌더 아님 = X (영업 가능 추정)
            r["웹빌더"] = "X"

        # 리뷰합계 계산 (방문자 + 블로그)
        r["리뷰합계"] = self._compute_review_total(r)

        # 로그
        log_msg = f"[{idx}/{total}] {name}"
        if log_items:
            log_msg += " -> " + " ".join(log_items)
        self.callback("log", log_msg)
        self.callback("progress", idx)
        self.callback("stats", dict(self.stats))

        # 업체간 딜레이
        self._random_delay()

    def run_keyword_search(self, keyword, output_file, start_page=1, max_pages=0, progress_file=None, rank_mode=False, naver_id=None, naver_pw=None):
        """
        키워드로 네이버 플레이스 검색 → 모든 방법 동원하여 업체 수집.

        수집 순서 (rank_mode=False):
        1. 모바일 지도 — 75건/페이지
        2. Playwright (map.naver.com) — 30건/페이지
        3. Map API — 30건/페이지
        4. HTML 파싱 폴백 — 보충

        수집 순서 (rank_mode=True):
        0. Selenium으로 실제 순위 수집
        1~4. 동일 (상세정보 보강)
        → 최종 출력 시 실제 순위 기준 정렬

        Args:
            keyword: 검색 키워드 (예: "수원 피부과")
            output_file: 출력 파일 경로 (.csv 또는 .xlsx)
            start_page: 시작 페이지 (기본 1)
            max_pages: 최대 페이지 수 (0 = 무제한, 끝까지)
            progress_file: 진행 저장 파일 (None이면 자동 생성)
            rank_mode: True이면 Selenium으로 실제 검색 순위 수집
        """
        self.running = True
        self.stats = {
            "place_id": 0, "category": 0, "name": 0, "address": 0,
            "phone": 0, "visitor_review": 0, "blog_review": 0,
            "hp": 0, "email": 0, "naver_id": 0,
            "responsive": 0, "talktalk": 0, "kakao": 0, "instagram": 0,
            "no_phone_btn": 0, "webbuilder": 0, "new_biz": 0,
            "blocked": 0, "success": 0,
        }

        if not progress_file:
            progress_file = output_file.rsplit(".", 1)[0] + "-progress.json"

        self.callback("log", f"키워드: {keyword}")
        self.callback("log", f"출력: {output_file}")
        self.callback("log", f"딜레이: {self.delay_min}~{self.delay_max}초 (랜덤)")
        if self.proxies:
            self.callback("log", f"프록시: {len(self.proxies)}개")

        # 이전 진행 복구 (같은 키워드일 때만)
        rows = []
        try:
            with open(progress_file, "r", encoding="utf-8") as f:
                prog = json.load(f)
                if prog.get("keyword") == keyword:
                    rows = prog.get("rows", [])
                    if rows:
                        self.callback("log", f"이전 진행 복구: {len(rows)}건")
                else:
                    os.remove(progress_file)
        except Exception:
            pass

        existing_pids = {r.get("고유번호") for r in rows if r.get("고유번호")}

        # ═══ 네이버 로그인 자동 로드 (naver_login.json) ═══
        if not naver_id or not naver_pw:
            _login_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "naver_login.json")
            if os.path.isfile(_login_file):
                try:
                    with open(_login_file, "r", encoding="utf-8") as _lf:
                        _login = json.load(_lf)
                        naver_id = _login.get("naver_id", "")
                        naver_pw = _login.get("naver_pw", "")
                        if naver_id and naver_pw:
                            self.callback("log", f"🔑 네이버 로그인 정보 로드: {naver_id[:3]}***")
                except Exception:
                    pass

        # selenium_ranks: Selenium 실행 시 자동 채워짐
        selenium_ranks = {}  # pid → rank

        # ═══ PHASE 1: Place ID 수집 ═══
        self.callback("log", "━━━ PHASE 1: 업체 목록 수집 ━━━")
        all_items = []
        seen_pids = set(existing_pids)

        def _add_items(items, source=""):
            added = 0
            for it in items:
                pid = str(it.get("id", ""))
                if pid and pid not in seen_pids:
                    seen_pids.add(pid)
                    all_items.append(it)
                    added += 1
            if added:
                self.callback("log", f"  → {source}: +{added}건 (총 {len(all_items)}건)")

        # 방법 1: Selenium 스크롤+페이지네이션 (가장 정확, 순위 보존)
        self.callback("log", "🌐 [1/4] Selenium 스크롤 수집 시도...")
        if self.running:
            sel_max = max_pages if max_pages > 0 else 0
            sel_results = self._selenium_rank_search(keyword, max_pages=sel_max, naver_id=naver_id, naver_pw=naver_pw)
            if sel_results:
                sel_items = []
                for sr in sel_results:
                    sel_items.append({
                        "id": str(sr["id"]),
                        "name": sr.get("name", ""),
                        "category": "",
                        "phone": "",
                        "virtualPhone": "",
                        "address": "",
                        "roadAddress": "",
                        "reviewCount": 0,
                        "blogCafeReviewCount": 0,
                    })
                    selenium_ranks[str(sr["id"])] = sr["rank"]
                _add_items(sel_items, f"Selenium ({len(sel_results)}건)")
            else:
                self.callback("log", "  ⚠️ Selenium 실패 — 모바일 지도로 폴백")

        # 방법 2: 모바일 지도 검색 (Selenium 부족분 보충)
        if self.running and len(all_items) < 50:
            self.callback("log", "📱 [2/4] 모바일 지도 검색 보충...")
            items, total = self._mobile_map_search(keyword)
            if items:
                _add_items(items, f"모바일 지도 ({total}건 중)")

        # 방법 3: Playwright (추가 보충)
        if self.running and len(all_items) < 10:
            self.callback("log", "🎭 [3/4] Playwright 시도...")
            items, total = self._playwright_search(keyword)
            if items:
                _add_items(items, "Playwright")

        # 방법 4: HTML 파싱 폴백 (5건 미만일 때만)
        if self.running and len(all_items) < 5:
            self.callback("log", "🔍 [4/4] HTML 파싱 폴백...")
            html_items, _ = self._html_search_fallback(keyword, 1)
            _add_items(html_items, "HTML 폴백")

        if not all_items and not rows:
            self.callback("log", "❌ 수집된 업체 없음")
            return

        resumed_count = len(rows)  # 이전 진행 복구된 건수 (이어하기 offset)
        total_expected = len(all_items) + resumed_count  # 전체 예상 건수
        self.callback("log", f"━━━ PHASE 1 완료: 총 {len(all_items)}건 신규 수집 ━━━")
        if resumed_count > 0:
            self.callback("log", f"  (기존 {resumed_count}건 + 신규 {len(all_items)}건 = 총 {total_expected}건)")
        self.callback("log", "")
        self.callback("log", "━━━ PHASE 2: 상세 정보 크롤링 ━━━")
        self.callback("total", total_expected)

        # ═══ PHASE 2: 상세 크롤링 ═══
        for idx, item in enumerate(all_items):
            if not self.running:
                break

            pid = str(item.get("id", ""))
            # 캐시된 기본정보 활용 (m.map JSON에서)
            cached = CrawlerEngine._place_info_cache.get(pid, {})
            name = item.get("name", "") or cached.get("name", "")
            category = item.get("category", "") or cached.get("category", "")
            v_phone = item.get("virtualPhone", "") or item.get("phone", "") or cached.get("virtualTel", "") or cached.get("tel", "")
            road_addr = item.get("roadAddress", "") or item.get("address", "") or cached.get("roadAddress", "") or cached.get("address", "")
            visitor_review = str(item.get("reviewCount", 0) or 0)
            blog_review = str(item.get("blogCafeReviewCount", 0) or 0)

            row = {
                "고유번호": pid,
                "상호명": name,
                "업종(키워드)": category,
                "안심번호": v_phone,
                "업체주소": road_addr,
                "방문자리뷰수": visitor_review,
                "블로그리뷰수": blog_review,
                "리뷰합계": 0,
                "플레이스URL": f"https://m.place.naver.com/place/{pid}" if pid else "",
                "업데이트날짜": time.strftime("%Y-%m-%d"),
                "홈페이지URL": "",
                "홈페이지반응형": "",
                "전화버튼없음": "",
                "웹빌더": "",
                "톡톡": "",
                "톡톡URL": "",
                "카카오톡": "",
                "인스타그램": "",
                "신규업체": "",
                "업체이메일": "",
                "네이버아이디": "",
            }

            # 기본 정보 부족 시 place HTML에서 보충
            self._last_parsed_place = {}
            if pid and (not name or not category):
                place_html = self._fetch_place_html(pid, keyword)
                parsed = self._parse_place_html(place_html)
                self._last_parsed_place = parsed
                if parsed.get("name") and not name:
                    row["상호명"] = parsed["name"]
                    name = parsed["name"]
                if parsed.get("category") and not category:
                    row["업종(키워드)"] = parsed["category"]
                if parsed.get("address") and not road_addr:
                    row["업체주소"] = parsed["address"]
                if parsed.get("phone") and not v_phone:
                    row["안심번호"] = parsed["phone"]
                if parsed.get("visitor_review") and visitor_review == "0":
                    row["방문자리뷰수"] = parsed["visitor_review"]
                if parsed.get("blog_review") and blog_review == "0":
                    row["블로그리뷰수"] = parsed["blog_review"]
                # 톡톡/카카오/인스타/신규는 place HTML 파싱 시점에 같이 채움
                if parsed.get("talktalk"):
                    row["톡톡URL"] = parsed["talktalk"]
                    row["톡톡"] = "O"
                    self._stat_inc("talktalk")
                if parsed.get("kakao"):
                    row["카카오톡"] = "O"
                    self._stat_inc("kakao")
                if parsed.get("instagram"):
                    row["인스타그램"] = "O"
                    self._stat_inc("instagram")
                if parsed.get("is_new") == "O":
                    row["신규업체"] = "O"
                    self._stat_inc("new_biz")
                self._random_delay()

            rows.append(row)

            # 상세 크롤링 (홈페이지/이메일/네이버ID) — 번호는 기존건 포함한 전체 기준
            self._detail_crawl_one(row, total_expected, resumed_count + idx + 1)

            # 매건 실시간 저장 (업종 정렬 적용)
            try:
                self.save_output(rows, output_file, keyword=keyword)
            except Exception as e:
                logger.debug(f"save error: {e}")

            # 진행 저장
            try:
                with open(progress_file, "w", encoding="utf-8") as f:
                    json.dump({"keyword": keyword, "rows": rows}, f, ensure_ascii=False)
            except Exception:
                pass

            self.callback("progress", len(rows))

        if not self.running:
            # 사용자 중지 — progress_file 보존 (이어하기용)
            if rows:
                self.save_output(rows, output_file, keyword=keyword)
                self.callback("log", f"⏸ 사용자 중지 — {len(rows)}건 저장됨 (이어하기 가능)")
                self.callback("done", output_file)
            else:
                self.callback("log", "사용자 중지 (수집 데이터 없음)")
            return  # progress_file 삭제하지 않고 리턴!

        # 이미 파이프라인에서 상세 크롤링 완료

        # 순위 모드: Selenium 순위 기준 정렬
        if rank_mode and selenium_ranks:
            self.callback("log", "📊 Selenium 순위 기준 정렬 중...")
            ranked = []
            unranked = []
            for r in rows:
                pid = r.get("고유번호", "")
                if pid in selenium_ranks:
                    r["_rank"] = selenium_ranks[pid]
                    ranked.append(r)
                else:
                    unranked.append(r)
            ranked.sort(key=lambda x: x["_rank"])
            # _rank 제거 (내부용)
            for r in ranked:
                r.pop("_rank", None)
            rows = ranked + unranked
            self.callback("log", f"  → 순위 확보 {len(ranked)}건 + 순위 없음 {len(unranked)}건")

        # 최종 저장 (업종 우선 정렬 적용 — rank_mode 시에는 순위 정렬 우선)
        if rank_mode and selenium_ranks:
            self.save_output(rows, output_file)  # 이미 순위 정렬됨, 업종 정렬 스킵
        else:
            self.save_output(rows, output_file, keyword=keyword)
        try:
            os.remove(progress_file)
        except Exception:
            pass

        # 최종 통계
        total = len(rows)

        def _pct(field, empty_vals=("", "X")):
            cnt = sum(1 for r in rows if (r.get(field) or "").strip() not in empty_vals)
            return cnt, round(cnt / total * 100) if total else 0

        c_cat, p_cat = _pct("업종(키워드)")
        c_name, p_name = _pct("상호명")
        c_email, p_email = _pct("업체이메일")
        c_phone, p_phone = _pct("안심번호")
        c_addr, p_addr = _pct("업체주소")
        c_hp, p_hp = _pct("홈페이지URL")
        c_resp, p_resp = _pct("홈페이지반응형", empty_vals=("", "X"))
        c_nopb, p_nopb = _pct("전화버튼없음", empty_vals=("", "X"))
        c_wb, p_wb = _pct("웹빌더", empty_vals=("", "X"))
        c_talk, p_talk = _pct("톡톡", empty_vals=("", "X"))
        c_kakao, p_kakao = _pct("카카오톡", empty_vals=("", "X"))
        c_insta, p_insta = _pct("인스타그램", empty_vals=("", "X"))
        c_new, p_new = _pct("신규업체", empty_vals=("", "X"))
        c_visit, p_visit = _pct("방문자리뷰수")
        c_blog, p_blog = _pct("블로그리뷰수")
        c_nid, p_nid = _pct("네이버아이디")
        c_pid, p_pid = _pct("고유번호")
        c_purl, p_purl = _pct("플레이스URL")
        c_date, p_date = _pct("업데이트날짜")

        s = self.stats
        summary = (
            f"\n=== 최종 결과 ({total}건) ===\n"
            f"  업종(키워드):    {c_cat}/{total} ({p_cat}%)\n"
            f"  상호명:          {c_name}/{total} ({p_name}%)\n"
            f"  업체이메일:      {c_email}/{total} ({p_email}%)  +{s['email']}\n"
            f"  안심번호:        {c_phone}/{total} ({p_phone}%)\n"
            f"  업체주소:        {c_addr}/{total} ({p_addr}%)\n"
            f"  홈페이지URL:     {c_hp}/{total} ({p_hp}%)  +{s['hp']}\n"
            f"  홈페이지반응형:  {c_resp}/{total} ({p_resp}%)  +{s.get('responsive', 0)}\n"
            f"  전화버튼없음:    {c_nopb}/{total} ({p_nopb}%)  +{s.get('no_phone_btn', 0)}\n"
            f"  웹빌더:          {c_wb}/{total} ({p_wb}%)  +{s.get('webbuilder', 0)}\n"
            f"  톡톡:            {c_talk}/{total} ({p_talk}%)  +{s.get('talktalk', 0)}\n"
            f"  카카오톡:        {c_kakao}/{total} ({p_kakao}%)  +{s.get('kakao', 0)}\n"
            f"  인스타그램:      {c_insta}/{total} ({p_insta}%)  +{s.get('instagram', 0)}\n"
            f"  신규업체:        {c_new}/{total} ({p_new}%)  +{s.get('new_biz', 0)}\n"
            f"  방문자리뷰수:    {c_visit}/{total} ({p_visit}%)\n"
            f"  블로그리뷰수:    {c_blog}/{total} ({p_blog}%)\n"
            f"  네이버아이디:    {c_nid}/{total} ({p_nid}%)  +{s['naver_id']}\n"
            f"  고유번호:        {c_pid}/{total} ({p_pid}%)\n"
            f"  플레이스URL:     {c_purl}/{total} ({p_purl}%)\n"
            f"  업데이트날짜:    {c_date}/{total} ({p_date}%)\n"
            f"\n완료: {output_file}"
        )
        self.callback("log", summary)
        self.callback("done", output_file)
        self.running = False

    @staticmethod
    def _sort_by_category(rows, keyword):
        """검색 키워드 업종 우선 정렬 + 나머지 업종별 그룹핑
        
        예: "해운대피부과" → 피부과 상단, 나머지(한의원, 의원, 병원...) 업종별 묶어서 하단
        """
        if not rows or not keyword:
            return rows

        # 키워드에서 업종명 추출 (지역명 제거)
        # "해운대피부과" → "피부과", "수원 성형외과" → "성형외과"
        kw_lower = keyword.replace(" ", "")
        category_hint = ""
        # 흔한 업종 키워드 목록
        common_cats = [
            "피부과", "성형외과", "정형외과", "치과", "한의원", "내과", "안과", "이비인후과",
            "비뇨기과", "산부인과", "소아과", "소아청소년과", "신경외과", "외과", "통증의학과",
            "재활의학과", "가정의학과", "정신건강의학과", "영상의학과", "마취통증의학과",
            "요양병원", "병원", "의원", "약국", "동물병원", "네일", "미용실", "헬스",
            "필라테스", "요가", "카페", "맛집", "음식점", "학원", "어린이집",
        ]
        for cat in common_cats:
            if cat in kw_lower:
                category_hint = cat
                break

        if not category_hint:
            # 업종 키워드를 못 찾으면 정렬 없이 반환
            return rows

        # 분류: 키워드 업종 매칭 vs 나머지
        primary = []  # 검색 업종과 일치
        others = {}   # 업종별 그룹핑

        for row in rows:
            cat = row.get("업종(키워드)", "") or ""
            name = row.get("상호명", "") or ""
            # 업종 또는 상호명에 키워드 업종이 포함되면 primary
            if category_hint in cat or category_hint in name:
                primary.append(row)
            else:
                # 업종별 그룹핑
                cat_key = cat.strip() if cat.strip() else "기타"
                if cat_key not in others:
                    others[cat_key] = []
                others[cat_key].append(row)

        # 나머지 업종은 알파벳/가나다 순 정렬
        sorted_others = []
        for cat_key in sorted(others.keys()):
            sorted_others.extend(others[cat_key])

        return primary + sorted_others

    def stop(self):
        """크롤링 중지"""
        self.running = False
